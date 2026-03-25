"""
Запуск рядом с manage.py:

    python find_small_orders.py

Ищет ордера где:
    - количество USDT (amount) < 1
    ИЛИ
    - стоимость в рублях (cost) < 400

Сохраняет: small_orders_YYYY-MM-DD.xlsx
"""

import os
import sys
import django
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

try:
    django.setup()
except Exception as e:
    print(f"Ошибка инициализации Django: {e}")
    sys.exit(1)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.db.models import Q
from orders.models import Order

# ── Выборка ───────────────────────────────────────────────────────────────────

qs = (
    Order.objects
    .filter(Q(amount__lt=1) | Q(cost__lt=400))
    .select_related("user")
    .order_by("user__username", "-created_at")
)

total = qs.count()
print(f"Найдено ордеров: {total}")

if total == 0:
    print("Ничего не найдено.")
    sys.exit(0)

# ── Excel ─────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Малые ордера"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")
red_fill    = PatternFill("solid", fgColor="FFE0E0")
center      = Alignment(horizontal="center", vertical="center")
left        = Alignment(horizontal="left",   vertical="center")
thin        = Side(style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "Пользователь",
    "ID ордера",
    "Биржа",
    "Тип",
    "Кол-во USDT",
    "Стоимость (руб)",
    "Курс",
    "Дата",
    "Причина",
]

ws.row_dimensions[1].height = 22
col_widths = [20, 24, 10, 8, 14, 18, 12, 16, 22]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center
    cell.border    = border
    ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = col_widths[col_idx - 1]

for row_idx, order in enumerate(qs, 2):
    amount_small = float(order.amount or 0) < 1
    cost_small   = float(order.cost or 0) < 400

    if amount_small and cost_small:
        reason = "USDT < 1 и руб < 400"
        fill   = red_fill
    elif amount_small:
        reason = "USDT < 1"
        fill   = red_fill
    else:
        reason = "руб < 400"
        fill   = yellow_fill

    values = [
        order.user.username,
        order.external_id,
        order.exchange_type,
        order.operation_type,
        str(order.amount),
        str(order.cost),
        str(order.price),
        order.created_at.strftime("%d.%m.%Y %H:%M") if order.created_at else "",
        reason,
    ]

    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill      = fill
        cell.border    = border
        cell.alignment = center if col_idx != 1 else left

# Итого
last_row = total + 3
ws.cell(row=last_row, column=1, value=f"Всего: {total}").font = Font(bold=True)

filename = f"small_orders_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
wb.save(filepath)

print(f"Excel сохранён: {filepath}")
print("Запушь на git и скачай.")