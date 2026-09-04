# --- Стандартная библиотека ---
import concurrent.futures
import io
import json
import os
import zipfile
from datetime import datetime, date, time, timezone as dt_timezone, timedelta
import json as _json
from itertools import groupby as _groupby

# --- Django ---
from django.contrib import messages
from django.contrib.auth import authenticate, get_user_model, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from decimal import Decimal, ROUND_DOWN
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from django.contrib.auth.forms import AuthenticationForm
from django.conf import settings

# --- Проект ---
from .bybit_api import sync_bybit_orders
from .bybit_service import get_orders_parallel as get_bybit_orders
from .mexc_api import sync_mexc_orders
from .mexc_service import get_mexc_orders_parallel
from .models import BankDetail, Exchange, IgnoredOrder, Order, UnprocessedOrder, UserExpense, MonthlyManualEntry, BalanceCorrection
from .receipt_service import create_or_update_and_send_receipt
from zoneinfo import ZoneInfo
MSK = ZoneInfo('Europe/Moscow')
from .uvedomlenie_generator import build_uvedomlenie, PERIODS
# --- Инициализация ---
User = get_user_model()
SYSTEM_START = datetime(2026, 2, 1, 0, 0, 0, tzinfo=dt_timezone.utc)
from collections import defaultdict

from .nds_generator import build_nds_declaration, QUARTERS
from .uvedomlenie_generator import build_uvedomlenie, PERIODS as UVED_PERIODS
from .nds_generator import build_nds_declaration, QUARTERS as NDS_QUARTERS
from .models import DocumentSubmission
from .models import MonthlySystemProfitSummary
from .models import MonthlyUserProfitCache
from django.views.decorators.http import require_POST
# Причины, при которых пункт скрывается молча (не ошибка данных,
# а просто "документ не требуется в этом периоде").
_FNS_SKIP_MARKERS = ('равна нулю', 'не требуется', 'не формируется', 'не превышает порог')
#ПОЛЬЗАК ПАНЕЛЬ____________________________________________________________________________________________________________________


def safe_decimal(value):
    """
    Превращает '1,4' -> 1.4, '1 000' -> 1000.0
    Если пусто или ошибка -> возвращает 0
    """
    if not value:
        return 0
    clean_val = str(value).replace(' ', '').replace(',', '.').strip()
    try:
        return float(clean_val)
    except ValueError:
        return 0


def truncate(value, places=3):
    """
    Обрезает число до нужного кол-ва знаков БЕЗ округления.
    86.3429  -> 86.342
    523.1239 -> 523.123
    46123.4567 -> 46123.456
    """
    try:
        d = Decimal(str(float(value)))
        factor = Decimal(10) ** -places
        return float(d.quantize(factor, rounding=ROUND_DOWN))
    except Exception:
        return 0.0


def format_ru_number(value, decimals=4):
    """
    Число в привычном для трейдеров русском формате: запятая вместо точки,
    пробел как разделитель тысяч. LANGUAGE_CODE проекта — en-us, поэтому
    обычный {{ value }} в шаблоне печатает точку — используем это в местах,
    где число рендерится сервером (не через JS formatMoney на клиенте).
    78.1387 -> "78,1387"
    """
    try:
        text = f"{float(value):,.{decimals}f}"
        return text.replace(",", " ").replace(".", ",")
    except (TypeError, ValueError):
        return "0"


def _accessible_banks_qs(user):
    """Банки, доступные юзеру: публичные + приватные, куда явно добавлен."""
    return (
        BankDetail.objects.filter(is_deleted=False)
        .filter(Q(is_public=True) | Q(allowed_users=user))
        .distinct().order_by('id')
    )


def _accessible_exchanges_qs(user):
    """Биржи, доступные юзеру: публичные + приватные, куда явно добавлен."""
    return (
        Exchange.objects.filter(is_deleted=False)
        .filter(Q(is_public=True) | Q(allowed_users=user))
        .distinct().order_by('id')
    )


@login_required
def my_orders_list(request):
    # Грузим список банков (для фильтра — весь список, чтобы можно было
    # найти старые ордера даже по банку, который у юзера сейчас скрыт)
    default_banks = BankDetail.objects.filter(is_deleted=False).order_by('id')

    # Для формы создания/редактирования ордера — только то, что юзеру
    # разрешено видеть (см. User.visible_banks/visible_exchanges, уровень 2
    # доступа — что юзер сам выбрал себе в /settings/ из разрешённого админом).
    visible_banks     = request.user.visible_banks.filter(is_deleted=False).order_by('id')
    visible_exchanges = request.user.visible_exchanges.filter(is_deleted=False).order_by('id')

    # ЛОГИКА СОЗДАНИЯ
    if request.method == 'POST' and 'create_order' in request.POST:

        # Очищаем старую ошибку из сессии
        request.session.pop('order_error', None)

        raw_date = request.POST.get('created_at')
        if raw_date:
            naive_date = datetime.strptime(raw_date, '%Y-%m-%dT%H:%M')
            order_date = timezone.make_aware(naive_date)
        else:
            order_date = timezone.now()

        # 1. Сохраняем настройки формы В ТОМ ВИДЕ, КАК ВВЕЛ ЮЗЕР
        request.session['saved_order_form'] = {
            'operation_type':   request.POST.get('operation_type'),
            'exchange':         request.POST.get('exchange'),
            'details':          request.POST.get('details'),
            'commission_value': request.POST.get('commission_value'),
            'commission_type':  request.POST.get('commission_type'),
            'created_at':       raw_date,
            'currency':         request.POST.get('currency')
        }

        # ЗАЩИТА 1: external_id обязателен
        external_id = request.POST.get('external_id', '').strip()
        if not external_id:
            request.session['order_error'] = 'Номер ордера биржи обязателен — заполните поле "Номер ордера"'
            return redirect('my_orders')

        # ЗАЩИТА 2: проверяем дубль у этого пользователя
        exchange_val = request.POST.get('exchange', '').strip()
        already_exists = Order.objects.filter(
            user=request.user,
            external_id=external_id,
            exchange_type=exchange_val,
        ).exists()

        if already_exists:
            request.session['order_error'] = f'Ордер {external_id} ({exchange_val}) уже существует в вашей системе'
            return redirect('my_orders')

        bank_id = request.POST.get('details')
        bank_instance = None
        if bank_id:
            bank_instance = BankDetail.objects.filter(id=bank_id, is_deleted=False).first()

        # 2. Очищаем числа перед сохранением в БД (полная точность, без обрезки)
        price_val      = safe_decimal(request.POST.get('price'))
        amount_val     = safe_decimal(request.POST.get('amount'))
        cost_val       = safe_decimal(request.POST.get('cost'))
        commission_val = safe_decimal(request.POST.get('commission_value'))

        # 3. Определяем процент биржи
        exchange_name  = exchange_val.lower()
        user_comm_rate = 0.0

        if 'bybit' in exchange_name:
            user_comm_rate = request.user.bybit_commission
        elif 'htx' in exchange_name or 'huobi' in exchange_name:
            user_comm_rate = request.user.htx_commission
        elif 'mexc' in exchange_name:
            user_comm_rate = request.user.mexc_commission
        elif 'bitget' in exchange_name:
            user_comm_rate = request.user.bitget_commission
        elif 'telegram' in exchange_name:
            user_comm_rate = request.user.telegram_commission

        # 4. Создаём ордер (в БД храним полные значения без обрезки)
        order = Order.objects.create(
            user=request.user,
            external_id=external_id,
            price=price_val,
            amount=amount_val,
            cost=cost_val,
            commission=commission_val,
            operation_type=request.POST.get('operation_type'),
            exchange_type=exchange_val,
            bank_detail=bank_instance,
            commission_type=request.POST.get('commission_type'),
            exchange_commission_rate=user_comm_rate,
            screenshot=request.FILES.get('screenshot'),
            # Второй скрин ("после исполнения") заполняется только для
            # Telegram — форма для остальных бирж это поле не отправляет,
            # request.FILES.get('screenshot_after') тогда просто None.
            screenshot_after=request.FILES.get('screenshot_after'),
            created_at=order_date,
            currency=request.POST.get('currency', 'USDT').strip().upper(),
        )

        # 5. Удаляем из необработанных
        UnprocessedOrder.objects.filter(
            user=request.user,
            order_id=external_id
        ).delete()

        # 6. Логика чека (Эвотор)
        need_receipt = request.POST.get('need_receipt') == 'on'
        
        if 'intelion' in (request.POST.get('exchange') or '').lower():
            need_receipt = False
        
        if need_receipt:
            contact_email = request.POST.get('receipt_contact', '').strip()
            if not contact_email:
                contact_email = request.user.email
            currency = request.POST.get('currency', 'USDT').strip().upper()
            if currency not in ('USDT', 'TON', 'BTC'):
                currency = 'USDT'
            # У BTC нужна точность до сатоши (8 знаков) — 3, как для USDT/TON,
            # обнулили бы небольшие суммы (0.00034521 BTC -> 0.000).
            amount_places = 8 if currency == 'BTC' else 3
            receipt_data = {
                "contact": contact_email,
                "sum":    truncate(order.cost,   2),
                "price":  truncate(order.price,  2),
                "amount": truncate(order.amount, amount_places),
                "purpose": f"Цифровая валюта {currency}",
            }
            create_or_update_and_send_receipt(order, receipt_data)

        return redirect('my_orders')

    # =====================================================================
    # ФИЛЬТРЫ — читаем из GET параметров
    # =====================================================================
    filter_date_from  = request.GET.get('date_from', '')
    filter_date_to    = request.GET.get('date_to', '')
    filter_exchange   = request.GET.get('exchange', '')
    filter_bank       = request.GET.get('bank', '')
    filter_op_type    = request.GET.get('op_type', '')
    filter_currency   = request.GET.get('currency', '')
    filter_order_id   = request.GET.get('order_id', '')

    orders = Order.objects.filter(user=request.user).order_by('-created_at')

    if filter_date_from:
        orders = orders.filter(created_at__date__gte=filter_date_from)
    if filter_date_to:
        orders = orders.filter(created_at__date__lte=filter_date_to)
    if filter_exchange:
        orders = orders.filter(exchange_type__icontains=filter_exchange)
    if filter_bank:
        orders = orders.filter(bank_detail_id=filter_bank)
    if filter_op_type:
        orders = orders.filter(operation_type=filter_op_type)
    if filter_currency:
        orders = orders.filter(currency=filter_currency)
    if filter_order_id:
        orders = orders.filter(external_id__icontains=filter_order_id)

    # Общее кол-во после фильтрации
    total_count = orders.count()

    # =====================================================================
    # ПАГИНАЦИЯ — 50 ордеров на страницу
    # =====================================================================
    from django.core.paginator import Paginator

    paginator    = Paginator(orders, 50)
    page_number  = request.GET.get('page', 1)
    page_obj     = paginator.get_page(page_number)

    # =====================================================================
    # Достаём сохранённые настройки формы
    # =====================================================================
    saved_data = request.session.get('saved_order_form', {})

    if not saved_data:
        saved_data = {
            'operation_type':  'BUY',
            'exchange':        'Bybit',
            'commission_type': 'PERCENT'
        }

    if saved_data.get('created_at'):
        current_time = saved_data['created_at']
    else:
        current_time = timezone.now().strftime('%Y-%m-%dT%H:%M')

    return render(request, 'orders/my_orders.html', {
        'orders':           page_obj,          # теперь page_obj вместо всех ордеров
        'page_obj':         page_obj,
        'total_count':      total_count,
        'default_banks':    default_banks,
        'visible_banks':    visible_banks,
        'visible_exchanges': visible_exchanges,
        'current_time':     current_time,
        'saved_data':       saved_data,
        # фильтры обратно в шаблон чтобы форма не сбрасывалась
        'filter_date_from': filter_date_from,
        'filter_date_to':   filter_date_to,
        'filter_exchange':  filter_exchange,
        'filter_bank':      filter_bank,
        'filter_op_type':   filter_op_type,
        'filter_currency':  filter_currency,
        'filter_order_id':  filter_order_id,
    })


@login_required
def edit_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)

    if request.method == 'POST':

        # 1. Обработка банка (ищем по ID или по названию)
        bank_id_or_name = request.POST.get('details')
        if bank_id_or_name:
            if bank_id_or_name.isdigit():
                bank_instance = BankDetail.objects.filter(id=bank_id_or_name, is_deleted=False).first()
            else:
                bank_instance = BankDetail.objects.filter(name=bank_id_or_name, is_deleted=False).first()

            if bank_instance:
                order.bank_detail = bank_instance

        # 2. Основные поля
        order.external_id    = request.POST.get('external_id')
        order.price          = safe_decimal(request.POST.get('price'))
        order.amount         = safe_decimal(request.POST.get('amount'))
        order.cost           = safe_decimal(request.POST.get('cost'))
        order.commission     = safe_decimal(request.POST.get('commission_value'))
        order.operation_type = request.POST.get('operation_type')
        order.exchange_type  = request.POST.get('exchange')
        order.commission_type = request.POST.get('commission_type')

        # 3. Валюта (USDT / TON / BTC)
        currency = request.POST.get('currency', '').strip().upper()
        if currency in ('USDT', 'TON', 'BTC'):
            order.currency = currency

        # 4. Дата
        raw_date = request.POST.get('created_at')
        if raw_date:
            naive_date = datetime.strptime(raw_date, '%Y-%m-%dT%H:%M')
            order.created_at = timezone.make_aware(naive_date)

        order.save()
        return redirect('my_orders')

@login_required
def delete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    

    if order.screenshot:
        order.screenshot.delete(save=False)
        
    order.delete()
    return redirect('my_orders')

@login_required
def upload_screenshot(request, order_id):

    if request.method == 'POST' and request.FILES.get('screenshot'):
        order = get_object_or_404(Order, id=order_id, user=request.user)
        file = request.FILES.get('screenshot')

    
        order_key = str(order.external_id)
        file.name = f"{order_key}.png"


        if order.screenshot:
            order.screenshot.delete(save=False)


        order.screenshot = file
        order.save()

    return redirect('my_orders')


@login_required
def upload_screenshot_after(request, order_id):
    """
    Дозагрузка/замена ВТОРОГО скрина ("после исполнения сделки") для уже
    существующего ордера — до этого кнопка 📷 умела писать только в
    order.screenshot, для screenshot_after дозагрузки не было вообще
    (можно было приложить только при самом создании Telegram-ордера).
    """
    if request.method == 'POST' and request.FILES.get('screenshot_after'):
        order = get_object_or_404(Order, id=order_id, user=request.user)
        file = request.FILES.get('screenshot_after')

        order_key = str(order.external_id)
        file.name = f"{order_key}_after.png"

        if order.screenshot_after:
            order.screenshot_after.delete(save=False)

        order.screenshot_after = file
        order.save()

    return redirect('my_orders')


@login_required
def profile_settings(request):
    user = request.user

    if request.method == 'POST':
        # 1. Сохраняем основные данные
        user.evotor_login   = request.POST.get('evotor_login')
        user.evotor_password = request.POST.get('evotor_password')
        user.kkt_id         = request.POST.get('kkt_id')
        user.email          = request.POST.get('email')
        user.payment_address = request.POST.get('payment_address')
        user.tax_type       = request.POST.get('tax_type')
        user.inn            = request.POST.get('inn')
        user.last_name = (request.POST.get('last_name') or '').strip()
        user.first_name = (request.POST.get('first_name') or '').strip()
        user.middle_name = (request.POST.get('middle_name') or '').strip()
        user.oktmo = (request.POST.get('oktmo') or '').strip()
        user.kod_no = (request.POST.get('kod_no') or '').strip()
        user.phone = (request.POST.get('phone') or '').strip()
        # 2. Сохраняем ключи API
        user.htx_access_key  = request.POST.get('htx_key')
        user.htx_private_key = request.POST.get('htx_secret')

        # MEXC: если ключ изменился — сбрасываем флаг невалидности
        new_mexc_key    = request.POST.get('mexc_key')
        new_mexc_secret = request.POST.get('mexc_secret')
        if new_mexc_key and new_mexc_key != user.mexc_api_key:
            user.mexc_key_valid = True
        user.mexc_api_key    = new_mexc_key
        user.mexc_api_secret = new_mexc_secret
        
        # Bybit: если ключ изменился — сбрасываем флаг невалидности
        new_bybit_key    = request.POST.get('bybit_key')
        new_bybit_secret = request.POST.get('bybit_secret')
        if new_bybit_key and new_bybit_key != user.bybit_api_key:
            user.bybit_key_valid = True
        user.bybit_api_key    = new_bybit_key
        user.bybit_api_secret = new_bybit_secret

        # 3. Сохраняем комиссии бирж
        user.bybit_commission    = float(request.POST.get('bybit_commission') or 0.3)
        user.htx_commission      = float(request.POST.get('htx_commission') or 0)
        user.mexc_commission     = float(request.POST.get('mexc_commission') or 0)
        user.bitget_commission   = float(request.POST.get('bitget_commission') or 0)
        user.telegram_commission = float(request.POST.get('telegram_commission') or 0.9)

        # 3.5 Какие банки/биржи показывать при ручном пробитии — личный выбор
        # юзера, но только из того, что ему вообще доступно (публичное +
        # приватное, куда его явно добавил админ в каталоге).
        selected_bank_ids     = request.POST.getlist('visible_banks')
        selected_exchange_ids = request.POST.getlist('visible_exchanges')
        user.visible_banks.set(_accessible_banks_qs(user).filter(id__in=selected_bank_ids))
        user.visible_exchanges.set(_accessible_exchanges_qs(user).filter(id__in=selected_exchange_ids))

        # 4. Логика смены пароля
        new_password = request.POST.get('new_password')
        if new_password and new_password.strip():
            user.set_password(new_password)
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Пароль успешно изменен!")
        else:
            user.save()
            messages.success(request, "Настройки сохранены.")

        return redirect('settings')

    all_banks      = _accessible_banks_qs(user)
    all_exchanges  = _accessible_exchanges_qs(user)
    visible_bank_ids     = set(user.visible_banks.values_list('id', flat=True))
    visible_exchange_ids = set(user.visible_exchanges.values_list('id', flat=True))

    return render(request, 'orders/settings.html', {
        'user': user,
        'all_banks':             all_banks,
        'all_exchanges':         all_exchanges,
        'visible_bank_ids':      visible_bank_ids,
        'visible_exchange_ids':  visible_exchange_ids,
    })

UNPROCESSED_PAGE_SIZE = 200


@login_required
def unprocessed_orders_list(request):
    user = request.user
    bybit_ok = bool(user.bybit_api_key and user.bybit_api_secret)
    mexc_ok  = bool(
        getattr(user, 'mexc_api_key', None) and getattr(user, 'mexc_api_secret', None)
    )

    if not bybit_ok and not mexc_ok:
        messages.warning(request, "API ключи не настроены. Зайди в настройки.")

    unprocessed_qs = (
        UnprocessedOrder.objects
        .filter(user=user)
        .order_by('-created_at')
    )

    # Список бирж для фильтра. Bybit/MEXC — всегда (это два основных
    # интегрированных источника синка), даже если у юзера сейчас там пусто.
    # Остальные (Telegram/HTX/Gate/...) — только те, что реально встречаются.
    exchange_choices = sorted(set(
        UnprocessedOrder.objects.filter(user=user)
        .values_list('exchange_type', flat=True).distinct()
    ) | {'Bybit', 'MEXC'})

    exchange_filter = request.GET.get('exchange', '').strip()
    if exchange_filter:
        unprocessed_qs = unprocessed_qs.filter(exchange_type=exchange_filter)

    # ИСПРАВЛЕНО: раньше рендерился весь queryset без ограничения — у
    # некоторых юзеров тут тысячи строк (найдено до ~4200), страница была
    # тяжёлой и на сервере, и в браузере. Теперь пагинация.
    paginator = Paginator(unprocessed_qs, UNPROCESSED_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj':           page_obj,
        'unprocessed_orders': page_obj.object_list,
        'bybit_configured':   bybit_ok,
        'mexc_configured':    mexc_ok,
        'exchange_filter':    exchange_filter,
        'exchange_choices':   exchange_choices,
    }
    return render(request, 'orders/unprocessed.html', context)


@login_required
def delete_unprocessed_order(request, pk):
    """
    Удаляет ордер из буфера необработанных И навсегда добавляет в игнор,
    чтобы синхронизация с биржи больше никогда его не подтягивала.
    """
    try:
        order = UnprocessedOrder.objects.get(pk=pk, user=request.user)

        # Добавляем в игнор (get_or_create — защита от дублей)
        IgnoredOrder.objects.get_or_create(
            user=request.user,
            order_id=order.order_id,
            defaults={"exchange_type": order.exchange_type},
        )

        order.delete()
        messages.success(request, "Ордер удалён и добавлен в игнор-лист.")

    except UnprocessedOrder.DoesNotExist:
        messages.error(request, "Ордер не найден.")

    return redirect("unprocessed_orders")


@login_required
@require_POST
def bulk_ignore_unprocessed_orders(request):
    """
    Массовое игнорирование — то же самое, что delete_unprocessed_order,
    только за один запрос по списку id (галочки на странице "Необработанные").
    Специально ограничено ТЕКУЩЕЙ СТРАНИЦЕЙ (галочки видны только в пределах
    одной страницы пагинации) - юзер физически не может случайно выбрать
    и снести тысячи строк одним кликом, максимум — размер страницы.
    """
    ids = request.POST.getlist('order_ids')
    if not ids:
        messages.warning(request, "Ничего не выбрано.")
        return redirect("unprocessed_orders")

    orders = list(UnprocessedOrder.objects.filter(pk__in=ids, user=request.user))

    if not orders:
        messages.error(request, "Выбранные ордера не найдены.")
        return redirect("unprocessed_orders")

    IgnoredOrder.objects.bulk_create(
        [
            IgnoredOrder(user=request.user, order_id=o.order_id, exchange_type=o.exchange_type)
            for o in orders
        ],
        ignore_conflicts=True,
    )
    deleted_count = len(orders)
    UnprocessedOrder.objects.filter(pk__in=[o.pk for o in orders]).delete()

    messages.success(request, f"Проигнорировано ордеров: {deleted_count}.")
    return redirect("unprocessed_orders")



@login_required
def user_profit_view(request):

    user   = request.user
    action = request.GET.get('action')

    # =====================================================================
    # РАСХОДЫ
    # =====================================================================
    if action == 'get_expenses':
        expenses = UserExpense.objects.filter(user=user).order_by('-month', 'name')
        return JsonResponse({'expenses': [
            {'id': e.id, 'name': e.name, 'amount': float(e.amount), 'month': e.month}
            for e in expenses
        ]})

    if action == 'add_expense':
        try:
            body   = _json.loads(request.body)
            name   = str(body.get('name', '')).strip()
            amount = float(body.get('amount', 0))
            month  = str(body.get('month', ''))
            if not name or amount <= 0 or not month:
                return JsonResponse({'error': 'Некорректные данные'}, status=400)
            UserExpense.objects.create(user=user, name=name, amount=amount, month=month)
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    if action == 'delete_expense':
        try:
            body   = _json.loads(request.body)
            exp_id = int(body.get('id', 0))
            UserExpense.objects.filter(id=exp_id, user=user).delete()
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    # =====================================================================
    # КОРРЕКЦИЯ ОСТАТКА
    # =====================================================================
    # Ручное списание остатка (крипта выведена из оборота, не распродана) —
    # реализовано как обычный BUY Order с отрицательными amount/cost и
    # фиксированным external_id="Списание остатка". Он естественно проходит
    # через тот же LIFO-расчёт, что и обычные ордера (и на этой странице,
    # и в Excel-отчёте), без дублирования логики в отдельном месте.
    # BalanceCorrection хранит только "назначение" и даёт удобный список/удаление.
    BALANCE_CORRECTION_CURRENCIES = ('USDT', 'TON', 'BTC')

    if action == 'get_balance_corrections':
        corrections = (
            BalanceCorrection.objects
            .filter(user=user)
            .select_related('order')
            .order_by('-created_at')
        )
        return JsonResponse({'corrections': [
            {
                'id': c.id,
                'purpose': c.purpose,
                'amount': float(abs(c.order.amount)),
                'price': float(abs(c.order.price)),
                'currency': c.order.currency,
                'date': c.order.created_at.strftime('%d.%m.%Y'),
            }
            for c in corrections
        ]})

    if action == 'add_balance_correction':
        try:
            body     = _json.loads(request.body)
            amount   = abs(float(body.get('amount', 0)))
            price    = abs(float(body.get('price', 0)))
            currency = str(body.get('currency', '')).strip().upper()
            purpose  = str(body.get('purpose', '')).strip()

            if amount <= 0 or price <= 0:
                return JsonResponse({'error': 'Количество и курс должны быть больше 0'}, status=400)
            if currency not in BALANCE_CORRECTION_CURRENCIES:
                return JsonResponse({'error': 'Некорректная валюта'}, status=400)
            if not purpose:
                return JsonResponse({'error': 'Укажите назначение списания'}, status=400)

            correction_order = Order.objects.create(
                user=user,
                external_id='Списание остатка',
                exchange_type='Коррекция остатка',
                operation_type='BUY',
                currency=currency,
                amount=Decimal(str(-amount)),
                price=Decimal(str(price)),
                cost=Decimal(str(-round(amount * price, 2))),
                is_verified=True,
                is_manual=True,
                created_at=timezone.now(),
            )
            correction = BalanceCorrection.objects.create(
                user=user, order=correction_order, purpose=purpose,
            )
            return JsonResponse({'ok': True, 'id': correction.id})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    if action == 'delete_balance_correction':
        try:
            body = _json.loads(request.body)
            cid  = int(body.get('id', 0))
            correction = BalanceCorrection.objects.filter(id=cid, user=user).first()
            if correction:
                correction.order.delete()  # BalanceCorrection удалится каскадом
            return JsonResponse({'ok': True})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    # =====================================================================
    # ОБЩАЯ ПРЕДЗАГРУЗКА ДАННЫХ
    # =====================================================================
    def _preload_user_data(year=None):
        from zoneinfo import ZoneInfo
        from collections import defaultdict
        MSK = ZoneInfo('Europe/Moscow')

        if year:
            cutoff = datetime(year + 1, 1, 1, tzinfo=MSK)
        else:
            cutoff = datetime(timezone.now().year + 1, 1, 1, tzinfo=MSK)

        raw_orders = (
            Order.objects
            .filter(user=user, created_at__gte=SYSTEM_START, created_at__lt=cutoff)
            .order_by('created_at')
            .values_list(
                'id', 'user_id', 'created_at', 'currency', 'operation_type',
                'amount', 'cost', 'price', 'commission', 'commission_type',
                'exchange_type', 'bank_detail_id', 'exchange_commission_rate',
            )
        )

        class _O:
            __slots__ = (
                'id', 'user_id', 'created_at', 'currency', 'operation_type',
                'amount', 'cost', 'price', 'commission', 'commission_type',
                'exchange_type', 'bank_detail_id', 'exchange_commission_rate',
            )
            def __init__(self, row):
                (self.id, self.user_id, self.created_at, self.currency,
                 self.operation_type, self.amount, self.cost, self.price,
                 self.commission, self.commission_type, self.exchange_type,
                 self.bank_detail_id, self.exchange_commission_rate) = row

        all_orders_by_user = defaultdict(list)
        for row in raw_orders:
            o = _O(row)
            all_orders_by_user[o.user_id].append(o)

        raw_expenses = (
            UserExpense.objects
            .filter(user=user)
            .values('user_id', 'month', 'name', 'amount')
        )
        expenses_by_user_month = defaultdict(float)
        expenses_list_by_user  = defaultdict(list)
        for e in raw_expenses:
            key = (e['user_id'], e['month'])
            expenses_by_user_month[key] += float(e['amount'])
            expenses_list_by_user[key].append({'name': e['name'], 'amount': float(e['amount'])})

        raw_manuals = (
            MonthlyManualEntry.objects
            .filter(user=user)
            .values('user_id', 'month', 'gross',
                     'buy_qty', 'buy_cost', 'buy_comm',
                     'sell_qty', 'sell_cost', 'sell_comm',
                     'exch_comm_rub')
        )

        class _Manual:
            __slots__ = ('gross', 'buy_qty', 'buy_cost', 'buy_comm',
                         'sell_qty', 'sell_cost', 'sell_comm', 'exch_comm_rub')
            def __init__(self, m):
                self.gross         = float(m.get('gross', 0) or 0)
                self.buy_qty       = float(m.get('buy_qty', 0) or 0)
                self.buy_cost      = float(m.get('buy_cost', 0) or 0)
                self.buy_comm      = float(m.get('buy_comm', 0) or 0)
                self.sell_qty      = float(m.get('sell_qty', 0) or 0)
                self.sell_cost     = float(m.get('sell_cost', 0) or 0)
                self.sell_comm     = float(m.get('sell_comm', 0) or 0)
                self.exch_comm_rub = float(m.get('exch_comm_rub', 0) or 0)

        manuals_by_user_month = {}
        for m in raw_manuals:
            manuals_by_user_month[(m['user_id'], m['month'])] = _Manual(m)

        return all_orders_by_user, expenses_by_user_month, expenses_list_by_user, manuals_by_user_month

    # =====================================================================
    # ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: полный расчёт по одной валюте
    # =====================================================================
    def _calc_currency_full(gross_raw, sell_cost, tax_type, share_percent,
                             month_expenses_share, ytd_base):
        gross = max(0.0, gross_raw - month_expenses_share)
        ndfl  = _calc_ndfl(gross, tax_type, sell_cost=sell_cost, ytd_base=ytd_base)
        after = gross - ndfl
        share = after * (share_percent / 100) if after > 0 else 0.0
        net   = after - share
        return {
            'gross': round(gross, 2),
            'ndfl':  round(ndfl,  2),
            'share': round(share, 2),
            'net':   round(net,   2),
        }

    # =====================================================================
    # РАСЧЁТ ПО МЕСЯЦАМ
    # =====================================================================
    if action == 'get_months':
        from zoneinfo import ZoneInfo
        MSK  = ZoneInfo('Europe/Moscow')
        uid  = user.id
        year = timezone.now().year

        all_orders_by_user, expenses_by_user_month, _, manuals_by_user_month = \
            _preload_user_data(year=year)

        tax_type    = str(getattr(user, 'tax_type', '') or '').strip().upper()
        user_shares = user.profit_shares if isinstance(user.profit_shares, dict) else {}
        tax_label   = 'УСН' if tax_type in ('USN_INCOME', 'USN_INCOME_OUTCOME') else 'НДФЛ'

        months_with_orders = set()
        for o in all_orders_by_user.get(uid, []):
            dt_msk = o.created_at.astimezone(MSK)
            months_with_orders.add((dt_msk.year, dt_msk.month))
        for (u_id, mo_key) in manuals_by_user_month:
            if u_id == uid:
                parts = mo_key.split('-')
                if len(parts) == 2:
                    months_with_orders.add((int(parts[0]), int(parts[1])))

        if not months_with_orders:
            return JsonResponse({'months': []})

        # Раньше на каждый месяц отдельно пересчитывалась вся цепочка
        # переноса остатка с начала (_calc_month_profit_from_orders +
        # _get_ytd_base_from_cache внутри неё ещё раз пересчитывали то же
        # самое для каждого предыдущего месяца) — у активных трейдеров
        # (десятки месяцев истории) страница грузилась секундами.
        # Теперь считаем один раз проходом вперёд по всем месяцам сразу
        # (см. _calc_all_months_profit_from_orders) — то же самое из тех
        # же живых ордеров, без сохранённого кэша, просто без повторной
        # работы.
        all_months_calc = _calc_all_months_profit_from_orders(uid, all_orders_by_user, months_with_orders)

        ytd_base_by_month = {}
        ytd_running = 0.0
        ytd_year_tracker = None
        for (yr_k, mo_k) in sorted(months_with_orders):
            if ytd_year_tracker != yr_k:
                ytd_running = 0.0
                ytd_year_tracker = yr_k
            ytd_base_by_month[(yr_k, mo_k)] = ytd_running
            calc_mo = all_months_calc[(yr_k, mo_k)]
            gross_mo = calc_mo['gross']
            if calc_mo['month_buy_qty'] == 0 and calc_mo['month_sell_qty'] == 0:
                manual_mo = manuals_by_user_month.get((uid, f"{yr_k}-{str(mo_k).zfill(2)}"))
                if manual_mo:
                    gross_mo = float(manual_mo.gross)
            expenses_mo = expenses_by_user_month.get((uid, f"{yr_k}-{str(mo_k).zfill(2)}"), 0.0)
            ytd_running += max(0.0, gross_mo - expenses_mo)

        months_data = []

        for (yr, mo) in sorted(months_with_orders):
            share_key = f"{yr}-{str(mo).zfill(2)}"

            calc = all_months_calc[(yr, mo)]
            usdt = calc['usdt']
            ton  = calc['ton']

            share_percent  = float(user_shares.get(share_key, 20.0))
            month_expenses = expenses_by_user_month.get((uid, share_key), 0.0)

            gross_raw_usdt = usdt['gross']
            gross_raw_ton  = ton['gross']
            gross_raw_all  = gross_raw_usdt + gross_raw_ton

            has_activity = (calc['month_buy_qty'] > 0 or calc['month_sell_qty'] > 0)
            manual = manuals_by_user_month.get((uid, share_key))
            if not has_activity and manual:
                gross_raw_all  = manual.gross
                gross_raw_usdt = gross_raw_all
                gross_raw_ton  = 0.0

            pos_usdt  = max(0.0, gross_raw_usdt)
            pos_ton   = max(0.0, gross_raw_ton)
            pos_total = pos_usdt + pos_ton

            if pos_total > 0:
                usdt_exp_share = (pos_usdt / pos_total) * month_expenses
                ton_exp_share  = (pos_ton  / pos_total) * month_expenses
            else:
                usdt_exp_share = 0.0
                ton_exp_share  = 0.0

            if tax_type not in ('USN_INCOME', 'USN_INCOME_OUTCOME'):
                ytd_base = ytd_base_by_month.get((yr, mo), 0.0)
            else:
                ytd_base = 0.0

            ytd_usdt = ytd_base * (pos_usdt / pos_total) if pos_total > 0 else ytd_base
            ytd_ton  = ytd_base * (pos_ton  / pos_total) if pos_total > 0 else 0.0

            # sell_cost для ИТОГО-колонки и налога — из ручной записи, если
            # активности по ордерам не было (так же, как в admin_profit_view:
            # sell_cost_all = calc['month_sell_cost'] if has_activity else
            # manual_sell_cost). Раньше тут всегда стоял calc['month_sell_cost']
            # (0 при фолбэке на manual), из-за чего налог УСН (1% от sell_cost)
            # на личной странице занижался до нуля при ручных записях.
            sell_cost_all = calc['month_sell_cost'] if has_activity else (manual.sell_cost if manual else 0.0)

            r_usdt = _calc_currency_full(
                gross_raw_usdt, usdt['month_sell_cost'], tax_type, share_percent,
                usdt_exp_share, ytd_usdt
            )
            r_ton = _calc_currency_full(
                gross_raw_ton, ton['month_sell_cost'], tax_type, share_percent,
                ton_exp_share, ytd_ton
            )
            r_all = _calc_currency_full(
                gross_raw_all, sell_cost_all, tax_type, share_percent,
                month_expenses, ytd_base
            )

            usdt_bank_comm = usdt['month_buy_comm'] + usdt['month_sell_comm']
            ton_bank_comm  = ton['month_buy_comm']  + ton['month_sell_comm']
            all_bank_comm  = usdt_bank_comm + ton_bank_comm
            usdt_exch_comm = usdt['month_exch_comm_rub']
            ton_exch_comm  = ton['month_exch_comm_rub']
            all_exch_comm  = usdt_exch_comm + ton_exch_comm

            # Те же "финальные" ВСЕ-значения для отображения (buy/sell cost,
            # комиссии) — из manual при фолбэке, иначе занижаются до нуля
            # точно как sell_cost_all выше.
            final_buy_cost  = calc['month_buy_cost']  if has_activity else (manual.buy_cost  if manual else 0.0)
            final_sell_cost = sell_cost_all
            final_bank_comm = all_bank_comm           if has_activity else (
                (manual.buy_comm + manual.sell_comm) if manual else 0.0
            )
            final_exch_comm = all_exch_comm           if has_activity else (manual.exch_comm_rub if manual else 0.0)

            months_data.append({
                'key':           share_key,
                'year':          yr,
                'month':         mo,
                'tax_label':     tax_label,
                'share_percent': share_percent,
                'month_expenses': round(month_expenses, 2),

                # ── ВСЕ ──────────────────────────────────────────
                'remainder_qty':     round(usdt['remainder_qty_display'], 2),
                'ton_remainder_qty': round(ton['remainder_qty_display'],  2),
                'buy_cost':          round(final_buy_cost,  2),
                'sell_cost':         round(final_sell_cost, 2),
                'bank_comm':         round(final_bank_comm, 2),
                'exch_comm_rub':     round(final_exch_comm, 2),
                'gross':             r_all['gross'],
                'ndfl':              r_all['ndfl'],
                'share_amount':      r_all['share'],
                'net_profit':        r_all['net'],

                # ── USDT ─────────────────────────────────────────
                'usdt_remainder':  round(usdt['remainder_qty_display'], 2),
                'usdt_buy_cost':   round(usdt['month_buy_cost'],  2),
                'usdt_sell_cost':  round(usdt['month_sell_cost'], 2),
                'usdt_bank_comm':  round(usdt_bank_comm, 2),
                'usdt_exch_comm':  round(usdt_exch_comm, 2),
                'usdt_gross':      r_usdt['gross'],
                'usdt_ndfl':       r_usdt['ndfl'],
                'usdt_share':      r_usdt['share'],
                'usdt_net':        r_usdt['net'],

                # ── TON ──────────────────────────────────────────
                'ton_remainder':   round(ton['remainder_qty_display'], 2),
                'ton_buy_cost':    round(ton['month_buy_cost'],  2),
                'ton_sell_cost':   round(ton['month_sell_cost'], 2),
                'ton_bank_comm':   round(ton_bank_comm, 2),
                'ton_exch_comm':   round(ton_exch_comm, 2),
                'ton_gross':       r_ton['gross'],
                'ton_ndfl':        r_ton['ndfl'],
                'ton_share':       r_ton['share'],
                'ton_net':         r_ton['net'],
            })

        return JsonResponse({'months': months_data})

    # =====================================================================
    # ДЕТАЛЬНЫЙ РАСЧЁТ ЗА ПЕРИОД (вкладка "Сегодня")
    # =====================================================================
    if action == 'get_turnover':
        from zoneinfo import ZoneInfo
        MSK      = ZoneInfo('Europe/Moscow')
        uid      = user.id
        exchange = request.GET.get('exchange', '')
        bank_id  = request.GET.get('bank', '')

        start_date_str = request.GET.get('start')
        end_date_str   = request.GET.get('end')

        if start_date_str and end_date_str:
            try:
                naive_start  = datetime.strptime(start_date_str, '%Y-%m-%d')
                naive_end    = datetime.strptime(end_date_str,   '%Y-%m-%d')
                naive_end    = naive_end.replace(hour=23, minute=59, second=59)
                period_start = timezone.make_aware(naive_start)
                period_end   = timezone.make_aware(naive_end)
            except ValueError:
                return JsonResponse({'error': 'Неверный формат даты'}, status=400)
        else:
            # ИСПРАВЛЕНО: timezone.now() — время в UTC, а .replace(hour=0,...)
            # обнулял часы В UTC, а не в МСК (границы суток были сдвинуты на
            # 3 часа — сделки с полуночи до 3:00 МСК попадали в "вчера", а
            # с 23:00 до 2:59 следующего дня МСК — ложно в "сегодня"). Ордера
            # у трейдеров пробиваются по МСК-времени биржи, поэтому сначала
            # переводим "сейчас" в МСК и только потом обнуляем часы.
            now_msk      = timezone.now().astimezone(MSK)
            period_start = now_msk.replace(hour=0,  minute=0,  second=0, microsecond=0)
            period_end   = now_msk.replace(hour=23, minute=59, second=59, microsecond=999999)

        year      = period_start.year
        month     = period_start.month
        share_key = f"{year}-{str(month).zfill(2)}"
        ms_start  = datetime(year, month, 1, tzinfo=MSK)
        ms_end    = (datetime(year + 1, 1, 1, tzinfo=MSK)
                     if month == 12 else datetime(year, month + 1, 1, tzinfo=MSK))

        all_orders_by_user, expenses_by_user_month, _, manuals_by_user_month = \
            _preload_user_data(year=year)

        tax_type      = str(getattr(user, 'tax_type', '') or '').strip().upper()
        user_shares   = user.profit_shares if isinstance(user.profit_shares, dict) else {}
        share_percent = float(user_shares.get(share_key, 20.0))
        tax_label     = 'УСН' if tax_type in ('USN_INCOME', 'USN_INCOME_OUTCOME') else 'НДФЛ'

        calc = _calc_month_profit_from_orders(
            uid, ms_start, ms_end, all_orders_by_user,
            exchange_filter=exchange,
            bank_filter_id=bank_id,
        )
        usdt = calc['usdt']
        ton  = calc['ton']

        month_expenses = expenses_by_user_month.get((uid, share_key), 0.0)

        gross_raw_usdt = usdt['gross']
        gross_raw_ton  = ton['gross']
        gross_raw_all  = gross_raw_usdt + gross_raw_ton

        has_activity = (calc['month_buy_qty'] > 0 or calc['month_sell_qty'] > 0)
        if not has_activity:
            manual = manuals_by_user_month.get((uid, share_key))
            if manual:
                gross_raw_all  = float(manual.gross)
                gross_raw_usdt = gross_raw_all
                gross_raw_ton  = 0.0

        pos_usdt  = max(0.0, gross_raw_usdt)
        pos_ton   = max(0.0, gross_raw_ton)
        pos_total = pos_usdt + pos_ton

        if pos_total > 0:
            usdt_exp_share = (pos_usdt / pos_total) * month_expenses
            ton_exp_share  = (pos_ton  / pos_total) * month_expenses
        else:
            usdt_exp_share = 0.0
            ton_exp_share  = 0.0

        if tax_type not in ('USN_INCOME', 'USN_INCOME_OUTCOME'):
            ytd_base = _get_ytd_base_from_cache(
                uid, year, month,
                all_orders_by_user,
                expenses_by_user_month,
                manuals_by_user_month,
            )
        else:
            ytd_base = 0.0

        ytd_usdt = ytd_base * (pos_usdt / pos_total) if pos_total > 0 else ytd_base
        ytd_ton  = ytd_base * (pos_ton  / pos_total) if pos_total > 0 else 0.0

        r_usdt = _calc_currency_full(
            gross_raw_usdt, usdt['month_sell_cost'], tax_type, share_percent,
            usdt_exp_share, ytd_usdt
        )
        r_ton = _calc_currency_full(
            gross_raw_ton, ton['month_sell_cost'], tax_type, share_percent,
            ton_exp_share, ytd_ton
        )
        r_all = _calc_currency_full(
            gross_raw_all, calc['month_sell_cost'], tax_type, share_percent,
            month_expenses, ytd_base
        )

        usdt_balance = usdt['remainder_qty_display']
        ton_balance  = ton['remainder_qty_display']

        period_orders_qs = (
            Order.objects
            .filter(user=user, created_at__range=(period_start, period_end))
            .select_related('bank_detail')
            .order_by('-created_at')
        )
        if exchange:
            period_orders_qs = period_orders_qs.filter(exchange_type__icontains=exchange)
        if bank_id:
            period_orders_qs = period_orders_qs.filter(bank_detail_id=bank_id)

        buy_orders  = [o for o in period_orders_qs if o.operation_type == 'BUY']
        sell_orders = [o for o in period_orders_qs if o.operation_type == 'SELL']

        usdt_buy_sum  = sum(float(o.cost or 0) for o in buy_orders  if o.currency == 'USDT')
        usdt_sell_sum = sum(float(o.cost or 0) for o in sell_orders if o.currency == 'USDT')
        ton_buy_sum   = sum(float(o.cost or 0) for o in buy_orders  if o.currency == 'TON')
        ton_sell_sum  = sum(float(o.cost or 0) for o in sell_orders if o.currency == 'TON')
        all_buy_sum   = usdt_buy_sum  + ton_buy_sum
        all_sell_sum  = usdt_sell_sum + ton_sell_sum

        period_buy_crypto  = sum(float(o.amount or 0) for o in buy_orders)
        period_sell_crypto = sum(float(o.amount or 0) for o in sell_orders)
        crypto_delta       = period_buy_crypto - period_sell_crypto

        def serialize(orders, is_sell=False):
            res = []
            for o in orders:
                display_id = o.external_id if o.external_id else str(o.id)
                comm_val   = float(o.commission or 0)
                bank_comm  = (float(o.cost or 0) * comm_val / 100
                              if o.commission_type == 'PERCENT' else comm_val)
                exch_comm  = (float(o.amount or 0) * float(o.exchange_commission_rate or 0) / 100
                              if is_sell else 0.0)
                res.append({
                    'id':            o.id,
                    'external_id':   display_id,
                    'price':         float(o.price  or 0),
                    'amount':        float(o.amount or 0),
                    'cost':          float(o.cost   or 0),
                    'date':          o.created_at.strftime('%d.%m.%Y %H:%M'),
                    'exchange':      getattr(o, 'exchange_type', ''),
                    'bank':          str(o.bank_detail) if o.bank_detail else '',
                    'bank_comm':     round(bank_comm, 2),
                    'exchange_comm': round(exch_comm, 2),
                    'currency':      getattr(o, 'currency', 'USDT'),
                })
            return res

        return JsonResponse({
            # ── Суммарно ─────────────────────────────────────────
            'buy_sum':        round(all_buy_sum,  2),
            'buy_count':      len(buy_orders),
            'sell_sum':       round(all_sell_sum, 2),
            'sell_count':     len(sell_orders),
            'gross':          r_all['gross'],
            'ndfl':           r_all['ndfl'],
            'share_amount':   r_all['share'],
            'net_profit':     r_all['net'],

            # ── USDT ─────────────────────────────────────────────
            'usdt_buy_sum':   round(usdt_buy_sum,  2),
            'usdt_sell_sum':  round(usdt_sell_sum, 2),
            'usdt_gross':     r_usdt['gross'],
            'usdt_ndfl':      r_usdt['ndfl'],
            'usdt_share':     r_usdt['share'],
            'usdt_net':       r_usdt['net'],

            # ── TON ──────────────────────────────────────────────
            'ton_buy_sum':    round(ton_buy_sum,  2),
            'ton_sell_sum':   round(ton_sell_sum, 2),
            'ton_gross':      r_ton['gross'],
            'ton_ndfl':       r_ton['ndfl'],
            'ton_share':      r_ton['share'],
            'ton_net':        r_ton['net'],

            # ── Общее ────────────────────────────────────────────
            'tax_label':      tax_label,
            'share_percent':  share_percent,
            'month_expenses': round(month_expenses, 2),
            'usdt_balance':   round(usdt_balance, 4),
            'ton_balance':    round(ton_balance,  4),
            'usdt_avg_price': round(usdt.get('avg_price', 0) or 0, 4),
            'ton_avg_price':  round(ton.get('avg_price', 0) or 0, 4),
            'crypto_balance': round(usdt_balance + ton_balance, 4),
            'crypto_delta':   round(crypto_delta, 4),
            'buy_orders':     serialize(buy_orders,  is_sell=False),
            'sell_orders':    serialize(sell_orders, is_sell=True),
            'user_shares':    user_shares,
        })

    # =====================================================================
    # РЕНДЕР СТРАНИЦЫ
    # =====================================================================
    default_banks = BankDetail.objects.filter(is_deleted=False)
    return render(request, 'orders/profit.html', {'default_banks': default_banks})

# Добавление в orders/views.py (или views_fns_user_v2.py)


def _deadline_25th(year, end_month):
    """
    Срок подачи документа: 25-е число месяца, следующего за концом
    отчётного периода. Если попадает на выходной (суббота/воскресенье) -
    переносится на ближайший понедельник (перенос по НК РФ).
    Праздничные дни НЕ учитываются (упрощение) - если понадобится точный
    производственный календарь, нужен отдельный справочник праздников.
    """
    next_month = end_month + 1
    next_year = year
    if next_month > 12:
        next_month = 1
        next_year += 1
    d = date(next_year, next_month, 25)
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def get_required_fns_documents(user, year, for_user_id=None):
    """
    ОБЩАЯ ФУНКЦИЯ - единый источник правды о том, какие документы ФНС
    сейчас актуальны для пользователя.

    ИЗМЕНЕНИЯ (группировка по кварталам для страницы /fns/):
      - для уведомлений добавлено поле 'amount' - сумма налога к уплате
        (сумма всех КБК-блоков из info['blocks']);
      - для всех документов добавлено 'deadline' (datetime.date) и
        'deadline_str' (ДД.ММ.ГГГГ) - срок подачи;
      - добавлено 'quarter_num' - номер квартала для группировки
        в шаблоне (1/2/3/4), считается от sort_month.
    """
    today = timezone.now()
    documents = []

    tax_type = (getattr(user, "tax_type", "") or "").upper()
    is_osno = tax_type in ("OSNO", "OCH")
    uved_title = ("Уведомление об авансовом платеже НДФЛ" if is_osno
                  else "Уведомление об авансовом платеже УСН")

    for period_key, period in UVED_PERIODS.items():
        last_month = period['last_month']
        period_ended = (year < today.year) or (year == today.year and today.month > last_month)
        if not period_ended:
            continue
        try:
            filename, xml_bytes, info = build_uvedomlenie(user, year, period_key)
        except ValueError as e:
            msg = str(e)
            if any(m in msg for m in _FNS_SKIP_MARKERS):
                continue
            documents.append({
                'kind': 'uvedomlenie', 'period_key': period_key, 'sort_month': last_month,
                'quarter_num': (last_month + 2) // 3,
                'title': uved_title,
                'url': None, 'ok': False, 'error': msg,
                'deadline': _deadline_25th(year, last_month),
            })
            continue

        total_amount = sum(a for _, a in info['blocks'])
        documents.append({
            'kind': 'uvedomlenie', 'period_key': period_key, 'sort_month': last_month,
            'quarter_num': (last_month + 2) // 3,
            'title': uved_title,
            'url': (f"/admin-panel/export/uvedomlenie/?user_id={for_user_id}&year={year}&period={period_key}"
                    if for_user_id else f"/fns/export/uvedomlenie/?year={year}&period={period_key}"),
            'ok': True, 'amount': total_amount,
            'deadline': _deadline_25th(year, last_month),
        })

    # НДС показываем ОСНО всегда, и УСН — начиная с квартала, где
    # накопительный с начала года sell-оборот превышает 20 000 000 ₽
    # (build_nds_declaration сам решает, нужна ли декларация конкретному
    # кварталу конкретного УСН-щика; если порог ещё не пробит - кидает
    # ValueError с "не превышает порог", это уже есть в _FNS_SKIP_MARKERS
    # и корректно скипается ниже, не попадая в список как ошибка).
    # Патент сюда не входит - НДС не платит вообще ни при каких условиях.
    show_nds_block = is_osno or tax_type in ("USN_INCOME", "USN_INCOME_OUTCOME")

    for q_key, q in (NDS_QUARTERS.items() if show_nds_block else []):
        if q_key == 'Q1':
            continue
        end_month = q['end_month']
        period_ended = (year < today.year) or (year == today.year and today.month > end_month)
        if not period_ended:
            continue
        try:
            build_nds_declaration(user, year, q_key)
        except ValueError as e:
            msg = str(e)
            if any(m in msg for m in _FNS_SKIP_MARKERS):
                continue
            documents.append({
                'kind': 'nds', 'period_key': q_key, 'sort_month': end_month,
                'quarter_num': (end_month + 2) // 3,
                'title': "Декларация по НДС",
                'url': None, 'ok': False, 'error': msg,
                'deadline': _deadline_25th(year, end_month),
            })
            continue
        documents.append({
            'kind': 'nds', 'period_key': q_key, 'sort_month': end_month,
            'quarter_num': (end_month + 2) // 3,
            'title': "Декларация по НДС",
            'url': (f"/admin-panel/export/nds/?user_id={for_user_id}&year={year}&quarter={q_key}"
                    if for_user_id else f"/fns/export/nds/?year={year}&quarter={q_key}"),
            'ok': True,
            'deadline': _deadline_25th(year, end_month),
        })

    # Форматируем deadline_str для шаблона (Django не умеет .strftime в тегах без фильтра date)
    for d in documents:
        d['deadline_str'] = d['deadline'].strftime('%d.%m.%Y')

    # Свежий квартал сверху (убывание по месяцу), внутри квартала
    # уведомление показываем раньше декларации НДС.
    documents.sort(key=lambda d: (-d['sort_month'], 0 if d['kind'] == 'uvedomlenie' else 1))
    return documents




def is_fns_all_submitted(user, year):
    """
    ЧИСТЫЙ расчёт статуса ФНС (без побочных эффектов, без кэша).
    Возвращает одну из трёх строк:
      'ok'      - есть обязательные документы за год, и ВСЕ отмечены
                  отправленными;
      'pending' - есть обязательные документы, но не все отмечены
                  отправленными;
      'none'    - у пользователя за этот год вообще нет ни одного
                  обязательного документа (не было торговой активности).
                  ВАЖНО: раньше это тоже считалось "зелёным"/True наравне
                  с 'ok' - из-за этого у неактивных пользователей всегда
                  горела галочка "отправлено", хотя отправлять им было
                  нечего. Теперь это отдельное (серое) состояние.

    Расчёт тяжёлый (строит XML уведомления/НДС на всю глубину года) —
    поэтому сама эта функция больше НЕ вызывается напрямую со страницы
    списка пользователей. См. refresh_fns_status_cached() и
    User.fns_status_cached — как bybit_key_valid/mexc_key_valid, обычное
    поле на юзере, которое обновляет фоновая задача
    tasks.recompute_fns_status_task, а шаблон читает напрямую.
    """
    docs = get_required_fns_documents(user, year)
    if not docs:
        return 'none'

    submitted_set = set(
        DocumentSubmission.objects
        .filter(user=user, year=year, submitted=True)
        .values_list('doc_type', 'period_key')
    )
    for d in docs:
        if not d['ok'] or (d['kind'], d['period_key']) not in submitted_set:
            return 'pending'

    return 'ok'


def refresh_fns_status_cached(user, year=None):
    """
    Пересчитывает is_fns_all_submitted и сохраняет в User.fns_status_cached.
    Вызывается: фоновой задачей recompute_fns_status_task (для всех
    юзеров, по расписанию) и toggle_document_submission (сразу для
    текущего юзера, чтобы после клика по чекбоксу не ждать следующего
    прогона фоновой задачи).
    """
    if year is None:
        year = timezone.now().year
    status = is_fns_all_submitted(user, year)
    if user.fns_status_cached != status:
        user.fns_status_cached = status
        user.save(update_fields=['fns_status_cached'])
    return status
 


@login_required(login_url='login')
def my_fns_documents(request):
    """Личная страница пользователя: список документов, требующих подачи."""
    user = request.user
    year = int(request.GET.get('year') or timezone.now().year)
 
    documents = get_required_fns_documents(user, year)
 
    submissions = {
        (s.doc_type, s.period_key): s
        for s in DocumentSubmission.objects.filter(user=user, year=year)
    }
    for d in documents:
        if d['ok']:
            sub = submissions.get((d['kind'], d['period_key']))
            d['submitted'] = bool(sub and sub.submitted)
 
    return render(request, 'orders/fns.html', {
        'year': year,
        'documents': documents,
    })
 
 
@login_required(login_url='login')
def user_export_uvedomlenie(request):
    """Скачивание СВОЕГО уведомления. user_id не принимается - только request.user."""
    year = int(request.GET.get('year') or timezone.now().year)
    period_key = request.GET.get('period', 'Q1')
 
    if period_key not in UVED_PERIODS:
        return JsonResponse({'error': 'Неверный период.'}, status=400)
 
    try:
        filename, xml_bytes, info = build_uvedomlenie(request.user, year, period_key)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
 
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=windows-1251')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
 
 
@login_required(login_url='login')
def user_export_nds(request):
    """Скачивание СВОЕЙ декларации НДС. user_id не принимается - только request.user."""
    year = int(request.GET.get('year') or timezone.now().year)
    quarter_key = request.GET.get('quarter', 'Q2')
    correction_number = int(request.GET.get('correction') or 0)

    if quarter_key not in NDS_QUARTERS or quarter_key == 'Q1':
        return JsonResponse({'error': 'Неверный квартал.'}, status=400)

    try:
        filename, xml_bytes, info = build_nds_declaration(request.user, year, quarter_key, correction_number)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)
 
    response = HttpResponse(xml_bytes, content_type='application/xml; charset=windows-1251')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
 
 
@login_required(login_url='login')
@require_POST
def toggle_document_submission(request):
    """
    Сохраняет отметку "Отправлено" для конкретного документа текущего
    пользователя. Сразу сбрасывает кэш точки статуса в админке, чтобы
    изменение отразилось без задержки в 5 минут.
    """
    doc_type = request.POST.get('doc_type')
    period_key = request.POST.get('period_key')
    year = request.POST.get('year')
    submitted = request.POST.get('submitted') == 'true'
 
    if not (doc_type and period_key and year):
        return JsonResponse({'error': 'Некорректные данные.'}, status=400)
    if doc_type not in ('uvedomlenie', 'nds'):
        return JsonResponse({'error': 'Некорректный тип документа.'}, status=400)
 
    year = int(year)
    obj, _ = DocumentSubmission.objects.get_or_create(
        user=request.user, doc_type=doc_type, period_key=period_key, year=year
    )
    obj.submitted = submitted
    obj.submitted_at = timezone.now() if submitted else None
    obj.save()
 
    refresh_fns_status_cached(request.user, year)
 
    return JsonResponse({'ok': True, 'submitted': obj.submitted})
 



# --- АДМИН ПАНЕЛЬ ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

def _get_celery_queue_depths():
    """
    Длины celery-очередей напрямую из Redis (LLEN) — быстрый health-индикатор
    для админки. Раньше затор в очереди "sync" (218 застрявших задач)
    обнаружили только руками через redis-cli — с этим на виду это будет
    заметно сразу на странице юзеров, а не только при ручной диагностике.
    Порог "warn" подобран по факту застревания.
    """
    import redis as redis_lib
    try:
        r = redis_lib.from_url(settings.CELERY_BROKER_URL, socket_timeout=2)
        depths = {q: r.llen(q) for q in ("sync", "receipt", "celery")}
        depths["warn"] = depths.get("sync", 0) > 50 or depths.get("receipt", 0) > 50
        depths["ok"] = True
        return depths
    except Exception as e:
        print(f"_get_celery_queue_depths failed: {e}")
        return {"ok": False}


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_users_list(request):
    """Список пользователей и управление ими"""
    if not request.user.is_superuser:
        return redirect('admin_login')

    # === 1. Редактирование пользователя ===
    if request.method == 'POST' and request.POST.get('action') == 'edit_user':
        user_id = request.POST.get('user_id')
        try:
            user_to_edit = User.objects.get(id=user_id)

            new_username = request.POST.get('username')
            if new_username:
                user_to_edit.username = new_username

            new_password = request.POST.get('password')
            if new_password:
                user_to_edit.set_password(new_password)

            user_to_edit.htx_access_key  = request.POST.get('htx_key')
            user_to_edit.htx_private_key = request.POST.get('htx_secret')

            # Bybit: если ключ изменился — сбрасываем флаг невалидности
            new_bybit_key    = request.POST.get('bybit_key')
            new_bybit_secret = request.POST.get('bybit_secret')
            if new_bybit_key and new_bybit_key != user_to_edit.bybit_api_key:
                user_to_edit.bybit_key_valid = True
            user_to_edit.bybit_api_key    = new_bybit_key
            user_to_edit.bybit_api_secret = new_bybit_secret

            # MEXC: если ключ изменился — сбрасываем флаг невалидности
            new_mexc_key    = request.POST.get('mexc_key')
            new_mexc_secret = request.POST.get('mexc_secret')
            if new_mexc_key and new_mexc_key != user_to_edit.mexc_api_key:
                user_to_edit.mexc_key_valid = True
            user_to_edit.mexc_api_key    = new_mexc_key
            user_to_edit.mexc_api_secret = new_mexc_secret

            user_to_edit.evotor_login    = request.POST.get('evotor_login')
            user_to_edit.evotor_password = request.POST.get('evotor_password')

            user_to_edit.save()
            messages.success(request, f"Пользователь {user_to_edit.username} успешно обновлен.")
        except User.DoesNotExist:
            messages.error(request, "Пользователь не найден.")

        return redirect('admin_users')

    # === 2. Создание пользователя ===
    if request.method == 'POST' and request.POST.get('action') == 'create_user':
        username = request.POST.get('username')
        password = request.POST.get('password')
        if username and password:
            new_user = User.objects.create(
                username=username,
                password=make_password(password),
                bybit_api_key=request.POST.get('bybit_key'),
                bybit_api_secret=request.POST.get('bybit_secret'),
                is_staff=True
            )

            # Дефолт для новых юзеров (по задаче): из банков — только
            # Сбербанк, из бирж — всё кроме Intelion показано сразу. Банки/
            # биржи — общий каталог на всю систему, юзер сам донастраивает
            # в /settings/ (может добавить что угодно из каталога, это уже
            # не ограничено админом).
            sberbank = BankDetail.objects.filter(is_deleted=False, name__icontains='сбер').first()
            if sberbank:
                new_user.visible_banks.set([sberbank])

            default_exchanges = list(
                Exchange.objects.filter(is_deleted=False, is_public=True).exclude(name='Intelion')
            )
            new_user.visible_exchanges.set(default_exchanges)

            messages.success(request, f"Пользователь {username} создан.")
        return redirect('admin_users')

    # === 3. Корректировка начального остатка ===
    if request.method == 'POST' and request.POST.get('action') == 'edit_balance':
        user_id        = request.POST.get('user_id')
        adjustment_str = request.POST.get('balance_adjustment', '0')
        rate_str       = request.POST.get('balance_rate', '0')

        try:
            user_to_edit = User.objects.get(id=user_id)

            amount_val = float(str(adjustment_str).replace(',', '.'))
            rate_val   = float(str(rate_str).replace(',', '.'))

            init_order = Order.objects.filter(
                user=user_to_edit,
                exchange_type="Остаток (до 1 фев)"
            ).first()

            if amount_val == 0:
                if init_order:
                    init_order.delete()
                messages.success(
                    request,
                    f"Начальный остаток пользователя {user_to_edit.username} удалён."
                )
            else:
                if rate_val <= 0:
                    messages.error(request, "Укажите курс закупки больше нуля.")
                    return redirect('admin_users')

                op_type    = 'BUY' if amount_val > 0 else 'SELL'
                abs_amount = abs(amount_val)
                cost       = abs_amount * rate_val

                if init_order:
                    init_order.operation_type = op_type
                    init_order.amount         = abs_amount
                    init_order.price          = rate_val
                    init_order.cost           = cost
                    init_order.save()
                    messages.success(
                        request,
                        f"Начальный остаток пользователя {user_to_edit.username} "
                        f"изменён: {amount_val} USDT по курсу {rate_val} ₽ "
                        f"(= {cost:,.2f} ₽)."
                    )
                else:
                    feb_first = datetime(2026, 2, 1, 0, 0, 0, tzinfo=dt_timezone.utc)
                    Order.objects.create(
                        user=user_to_edit,
                        external_id=f"INIT-{user_to_edit.id}",
                        operation_type=op_type,
                        amount=abs_amount,
                        price=rate_val,
                        cost=cost,
                        exchange_type="Остаток (до 1 фев)",
                        created_at=feb_first
                    )
                    messages.success(
                        request,
                        f"Начальный остаток внесён для {user_to_edit.username}: "
                        f"{amount_val} USDT по курсу {rate_val} ₽ "
                        f"(= {cost:,.2f} ₽)."
                    )

        except User.DoesNotExist:
            messages.error(request, "Пользователь не найден.")
        except ValueError:
            messages.error(request, "Некорректное значение количества или курса.")

        return redirect('admin_users')

    # === 4. Настройка доли пользователя по месяцам ===
    if request.method == 'POST' and request.POST.get('action') == 'edit_share':
        user_id = request.POST.get('user_id')
        year    = request.POST.get('share_year')

        try:
            user_to_edit = User.objects.get(id=user_id)

            if not isinstance(user_to_edit.profit_shares, dict):
                user_to_edit.profit_shares = {}

            for i in range(1, 13):
                month_str = f"{i:02d}"
                key       = f"{year}-{month_str}"
                val       = request.POST.get(f"share_{month_str}")

                if val is not None and val.strip() != "":
                    user_to_edit.profit_shares[key] = float(str(val).replace(',', '.'))

            user_to_edit.save()
            messages.success(
                request,
                f"Доля для пользователя {user_to_edit.username} обновлена на {year} год."
            )
        except User.DoesNotExist:
            messages.error(request, "Пользователь не найден.")
        except ValueError:
            messages.error(request, "Ошибка в формате числа доли. Используйте цифры.")

        return redirect('admin_users')

    # === 5. Вывод списка пользователей ===
    users = User.objects.all().order_by('id')

    # Один запрос — все user_id у которых есть manual entries
    users_with_manual = set(
        MonthlyManualEntry.objects
        .values_list('user_id', flat=True)
        .distinct()
    )

    # Один запрос — все user_id у которых есть init_order
    init_orders_map = {
        o.user_id: o
        for o in Order.objects.filter(exchange_type="Остаток (до 1 фев)")
    }

    # Один запрос — суммы BUY и SELL по всем пользователям
    from django.db.models import Sum, Case, When, FloatField
    balance_qs = (
        Order.objects
        .values('user_id')
        .annotate(
            total_buy=Sum(
                Case(When(operation_type='BUY',  then='amount'), default=0, output_field=FloatField())
            ),
            total_sell=Sum(
                Case(When(operation_type='SELL', then='amount'), default=0, output_field=FloatField())
            ),
        )
    )
    balance_map = {row['user_id']: row for row in balance_qs}

    for u in users:
        bal = balance_map.get(u.id, {})
        bought = float(bal.get('total_buy',  0) or 0)
        sold   = float(bal.get('total_sell', 0) or 0)
        u.real_balance = bought - sold

        init_order = init_orders_map.get(u.id)
        if init_order:
            u.has_initial_balance = True
            u.init_amount = (
                float(init_order.amount)
                if init_order.operation_type == 'BUY'
                else -float(init_order.amount)
            )
            u.init_rate = float(init_order.price)
        else:
            u.has_initial_balance = False
            u.init_amount = 0
            u.init_rate   = 0

        u.shares_json       = json.dumps(u.profit_shares) if u.profit_shares else "{}"
        u.has_manual_entries = u.id in users_with_manual

    return render(request, 'custom_admin/users_list.html', {
        'users': users,
        'fns_year': timezone.now().year,
        'queue_depths': _get_celery_queue_depths(),
    })


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_catalog(request):
    """
    Каталог банков/бирж — общий список на всю систему (BankDetail/Exchange).
    Раньше редактировался только через встроенную /admin/ — теперь то же
    самое доступно прямо в кастомной админке.

    ВАЖНО про переименование: у BankDetail это безопасно (Order.bank_detail —
    FK, переименование меняет отображение везде разом). У Exchange — НЕТ:
    Order.exchange_type — обычная строка, не FK на Exchange, и по её
    значению (точное совпадение/подстрока) завязана уйма бизнес-логики —
    комиссии (views.py/api_views.py: 'bybit' in ex_lower и т.п.), место
    расчётов в чеке (evotor_atol.py), верификация через API биржи
    (exchange_api.py: exchange_name == "Bybit"/"MEXC"). Переименование биржи
    в каталоге НЕ обновит эти проверки и НЕ тронет уже созданные ордера —
    новые ордера с этим новым именем тихо перестанут матчиться с нужной
    логикой. Поэтому переименование бирж сознательно не реализовано —
    только добавление новых и скрытие/восстановление существующих.
    """
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_bank':
            name = (request.POST.get('name') or '').strip()
            if name:
                BankDetail.objects.get_or_create(name=name)
                messages.success(request, f'Банк "{name}" добавлен.')
            else:
                messages.error(request, 'Название банка не может быть пустым.')

        elif action == 'rename_bank':
            bank     = BankDetail.objects.filter(id=request.POST.get('id')).first()
            new_name = (request.POST.get('name') or '').strip()
            if bank and new_name:
                bank.name = new_name
                bank.save(update_fields=['name'])
                messages.success(request, 'Банк переименован.')

        elif action == 'toggle_bank':
            bank = BankDetail.objects.filter(id=request.POST.get('id')).first()
            if bank:
                bank.is_deleted = not bank.is_deleted
                bank.save(update_fields=['is_deleted'])
                messages.success(
                    request,
                    f'Банк "{bank.name}" {"скрыт" if bank.is_deleted else "восстановлен"}.'
                )

        elif action == 'add_exchange':
            name = (request.POST.get('name') or '').strip()
            if name:
                Exchange.objects.get_or_create(name=name)
                messages.success(request, f'Биржа "{name}" добавлена.')
            else:
                messages.error(request, 'Название биржи не может быть пустым.')

        elif action == 'toggle_exchange':
            exchange = Exchange.objects.filter(id=request.POST.get('id')).first()
            if exchange:
                exchange.is_deleted = not exchange.is_deleted
                exchange.save(update_fields=['is_deleted'])
                messages.success(
                    request,
                    f'Биржа "{exchange.name}" {"скрыта" if exchange.is_deleted else "восстановлена"}.'
                )

        elif action == 'toggle_bank_public':
            bank = BankDetail.objects.filter(id=request.POST.get('id')).first()
            if bank:
                bank.is_public = not bank.is_public
                bank.save(update_fields=['is_public'])
                if not bank.is_public:
                    # Стал приватным — убираем его из личного показа у всех,
                    # кроме явно разрешённых (на этот момент разрешённых нет),
                    # иначе юзер продолжит видеть его в /settings/ по старой памяти.
                    for u in bank.visible_users.exclude(id__in=bank.allowed_users.all()):
                        u.visible_banks.remove(bank)
                messages.success(
                    request,
                    f'Банк "{bank.name}" теперь {"публичный" if bank.is_public else "приватный"}.'
                )

        elif action == 'toggle_exchange_public':
            exchange = Exchange.objects.filter(id=request.POST.get('id')).first()
            if exchange:
                exchange.is_public = not exchange.is_public
                exchange.save(update_fields=['is_public'])
                if not exchange.is_public:
                    for u in exchange.visible_users.exclude(id__in=exchange.allowed_users.all()):
                        u.visible_exchanges.remove(exchange)
                messages.success(
                    request,
                    f'Биржа "{exchange.name}" теперь {"публичная" if exchange.is_public else "приватная"}.'
                )

        elif action == 'add_bank_access':
            bank = BankDetail.objects.filter(id=request.POST.get('id')).first()
            target_user = User.objects.filter(id=request.POST.get('user_id')).first()
            if bank and target_user:
                bank.allowed_users.add(target_user)
                messages.success(request, f'{target_user.username} добавлен(а) в доступ к банку "{bank.name}".')

        elif action == 'remove_bank_access':
            bank = BankDetail.objects.filter(id=request.POST.get('id')).first()
            target_user = User.objects.filter(id=request.POST.get('user_id')).first()
            if bank and target_user:
                bank.allowed_users.remove(target_user)
                target_user.visible_banks.remove(bank)
                messages.success(request, f'{target_user.username} убран(а) из доступа к банку "{bank.name}".')

        elif action == 'add_exchange_access':
            exchange = Exchange.objects.filter(id=request.POST.get('id')).first()
            target_user = User.objects.filter(id=request.POST.get('user_id')).first()
            if exchange and target_user:
                exchange.allowed_users.add(target_user)
                messages.success(request, f'{target_user.username} добавлен(а) в доступ к бирже "{exchange.name}".')

        elif action == 'remove_exchange_access':
            exchange = Exchange.objects.filter(id=request.POST.get('id')).first()
            target_user = User.objects.filter(id=request.POST.get('user_id')).first()
            if exchange and target_user:
                exchange.allowed_users.remove(target_user)
                target_user.visible_exchanges.remove(exchange)
                messages.success(request, f'{target_user.username} убран(а) из доступа к бирже "{exchange.name}".')

        return redirect('admin_catalog')

    return render(request, 'custom_admin/catalog.html', {
        'banks':     BankDetail.objects.all().order_by('is_deleted', 'name').prefetch_related('allowed_users'),
        'exchanges': Exchange.objects.all().order_by('is_deleted', 'name').prefetch_related('allowed_users'),
    })


def _month_name(n):
    return ["Январь","Февраль","Март","Апрель","Май","Июнь",
            "Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"][n - 1]


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def export_excel_report(request):
    """
    Экспорт отчёта в Excel по логике КУДиР.

    Торговый результат включает остаток с предыдущего месяца:
      data_start фиксируется ДО строки переноса остатка —
      чтобы SUM диапазон захватывал строку переноса.

    Остаток = =E{tr}-I{tr}-M{tr}
    Курс остатка — LIFO (последние закупы первыми), считается в Python.
      При отрицательном остатке курс всё равно берётся по LIFO —
      себестоимость остатка будет отрицательной и уменьшит прибыль.
    Себестоимость остатка = =E{ost}*F{ost}
    Реализованный = G{tr} - G{ost}
    Прибыль = K{rr} - G{rr} - H{tr} - L{tr}

    Все даты и группировка по месяцам — в МСК (Europe/Moscow).
    """
    user_id    = request.GET.get('user_id')
    start_date = request.GET.get('start')
    end_date   = request.GET.get('end')
    bank_id    = request.GET.get('bank_id')
    op_type    = request.GET.get('type')
    currency   = request.GET.get('currency', '').strip().upper()  # '' | 'USDT' | 'TON'

    SYSTEM_START = datetime(2026, 2, 1, 0, 0, 0, tzinfo=dt_timezone.utc)

    # Трейдеры с реально импортированным январём (см. history.md §43.2/§45) —
    # по прямому решению показываем им январь построчно, вместо схлопнутой
    # строки "Остаток (до 1 фев)" (которая всё равно с этими ордерами не
    # согласована — отдельное число, вбитое независимо). Для всех остальных
    # трейдеров поведение не меняется вообще.
    has_january_orders = Order.objects.filter(
        user_id=user_id,
        created_at__lt=SYSTEM_START,
    ).exclude(exchange_type="Остаток (до 1 фев)").exists()

    if has_january_orders:
        orders = Order.objects.filter(
            user_id=user_id,
        ).exclude(exchange_type="Остаток (до 1 фев)").order_by('created_at')
    else:
        orders = Order.objects.filter(
            user_id=user_id,
            created_at__gte=SYSTEM_START
        ).order_by('created_at')

    if start_date: orders = orders.filter(created_at__date__gte=start_date)
    if end_date:   orders = orders.filter(created_at__date__lte=end_date)
    if bank_id and bank_id.isdigit(): orders = orders.filter(bank_detail_id=bank_id)
    if op_type:    orders = orders.filter(operation_type=op_type)

    # === ФИЛЬТР ПО ВАЛЮТЕ ===
    if currency == 'TON':
        orders = orders.filter(currency='TON')
    elif currency == 'USDT':
        orders = orders.filter(currency='USDT')

    orders_list = list(orders)

    # Исторический ордер — всегда первый (кроме TON), и только если январь
    # НЕ показывается построчно (иначе задвоение: и реальные январские
    # ордера, и несвязанный с ними "Остаток" одновременно).
    if currency != 'TON' and not has_january_orders:
        init_order = Order.objects.filter(
            user_id=user_id,
            exchange_type="Остаток (до 1 фев)"
        ).first()

        if init_order:
            existing_ids = {o.id for o in orders_list}
            if init_order.id not in existing_ids:
                orders_list = [init_order] + orders_list

    # Группировка по месяцам в МСК
    months_in_period = set()
    for o in orders_list:
        dt_msk = o.created_at.astimezone(MSK)
        months_in_period.add((dt_msk.year, dt_msk.month))
    multi_month = len(months_in_period) > 1

    # =====================================================================
    # Книга и стили
    # =====================================================================
    wb = openpyxl.Workbook()
    ws = wb.active

    if currency == 'TON':
        ws.title = "Отчет TON"
    elif currency == 'USDT':
        ws.title = "Отчет USDT"
    else:
        ws.title = "Отчет"

    thin_border  = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'),  bottom=Side(style='thin'))
    bold_font    = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left', vertical='center', wrap_text=True)

    cv_label = f" ({currency})" if currency else ""
    ws.merge_cells('A1:M1')
    ws['A1'] = f"Учет приобретенной и проданной цифровой валюты (ЦВ){cv_label}"
    ws['A1'].font      = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:A3'); ws['A2'] = "№ п/п"
    ws.merge_cells('B2:B3'); ws['B2'] = "Дата операции"
    ws.merge_cells('C2:C3'); ws['C2'] = "Номер документа"
    ws.merge_cells('D2:D3'); ws['D2'] = "Наименование ЦВ"
    ws.merge_cells('E2:H2'); ws['E2'] = "Приобретение цифровой валюты"
    ws.merge_cells('I2:M2'); ws['I2'] = "Продажа цифровой валюты"

    ws.append([
        "", "", "", "",
        "Кол-во", "Курс", "Стоимость", "Комиссия банка",
        "Кол-во", "Курс", "Стоимость", "Комиссия банка", "Комиссия биржи (USDT)"
    ])

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=13):
        for cell in row:
            cell.font      = bold_font
            cell.alignment = center_align
            cell.border    = thin_border

    # =====================================================================
    # LIFO расчёт курса остатка
    # Работает и при отрицательном остатке — курс всегда берётся по последним закупам.
    # prev_carry = (qty, price) — перенос остатка с прошлого месяца
    # =====================================================================
    def _calc_lifo_price(month_orders_list, remainder_qty, prev_carry=None):
        """
        Считает курс остатка по методу LIFO.
        Идём с конца списка покупок текущего месяца, если не хватает —
        берём из переноса остатка прошлого месяца (prev_carry).

        При отрицательном остатке курс всё равно считается —
        берём абсолютное значение для поиска курса, знак сохраняем в количестве.
        Это позволяет корректно посчитать себестоимость даже в минус.
        """
        # Используем abs для поиска курса — знак учтётся через кол-во в Excel
        abs_qty = abs(remainder_qty)

        if abs_qty == 0:
            return 0.0

        # Только BUY ордера текущего месяца — от последнего к первому.
        # qty <= 0 пропускаем — это не реальная покупка (например, строка
        # "Списание остатка" от коррекции), у неё нет своего LIFO-слоя
        # себестоимости, но её количество уже учтено в остатке через сумму.
        buys = []
        for o in reversed(month_orders_list):
            if o.operation_type == 'BUY':
                qty = float(o.amount or 0)
                if qty <= 0:
                    continue
                cost  = float(o.cost   or 0)
                price = float(o.price  or 0)
                unit_price = cost / qty if qty > 0 else price
                buys.append((qty, unit_price))

        # Если не хватает покупок текущего месяца — добавляем перенос остатка
        if prev_carry is not None:
            carry_qty, carry_price = prev_carry
            if carry_qty != 0 and carry_price > 0:
                buys.append((abs(carry_qty), carry_price))

        if not buys:
            return 0.0

        total_cost = 0.0
        need = abs_qty

        for qty, unit_price in buys:
            if need <= 0:
                break
            take = min(qty, need)
            total_cost += take * unit_price
            need -= take

        # Если закупов не хватило — берём курс последнего известного закупа
        if need > 0 and buys:
            last_price = buys[-1][1]
            total_cost += need * last_price

        return round(total_cost / abs_qty, 4) if abs_qty > 0 else 0.0

    # =====================================================================
    # Запись строки ордера
    # =====================================================================
    def _get_cv_name(o):
        if currency == 'TON':
            return 'TON'
        elif currency == 'USDT':
            return 'USDT'
        if getattr(o, 'currency', '') == 'TON':
            return 'TON'
        return 'USDT'

    def write_order_row(o, idx, is_historical=False):
        row_num    = ws.max_row + 1
        comm_val   = float(o.commission or 0)
        total_comm = float(o.cost) * comm_val / 100 if o.commission_type == 'PERCENT' else comm_val
        amount_f   = float(o.amount or 0)
        price_f    = float(o.price  or 0)
        exch_rate  = float(o.exchange_commission_rate or 0)

        if is_historical:
            label = "-"
        elif o.created_at:
            label = o.created_at.astimezone(MSK).strftime('%d.%m.%Y')
        else:
            label = ""

        num     = 0 if is_historical else idx
        cv_name = _get_cv_name(o)

        is_telegram = 'telegram' in str(getattr(o, 'exchange_type', '') or '').lower()

        if o.operation_type == 'BUY':
            buy_cost_val = float(o.cost or 0) if is_telegram else f"=E{row_num}*F{row_num}"
            ws.append([
                num, label, o.external_id, cv_name,
                amount_f,
                price_f,
                buy_cost_val,
                total_comm if total_comm > 0 else 0,
                0, 0, 0, 0, 0
            ])
        else:
            exch_comm_usdt = amount_f * exch_rate / 100
            sell_cost_val = float(o.cost or 0) if is_telegram else f"=I{row_num}*J{row_num}"
            ws.append([
                num, label, o.external_id, cv_name,
                0, 0, 0, 0,
                amount_f,
                price_f,
                sell_cost_val,
                total_comm if total_comm > 0 else 0,
                exch_comm_usdt if exch_rate > 0 else 0
            ])

    # =====================================================================
    # Перенос остатка предыдущего месяца
    # =====================================================================
    def write_carry_row(label, prev_ost_row):
        cv = currency if currency in ('USDT', 'TON') else 'USDT'
        ws.append([
            0, "-", label, cv,
            f"=E{prev_ost_row}",
            f"=F{prev_ost_row}",
            f"=G{prev_ost_row}",
            0,
            0, 0, 0, 0, 0
        ])

    # =====================================================================
    # Итоговые строки после блока ордеров
    # =====================================================================
    def write_summary(label_suffix, data_start, data_end,
                      last_sell_price=0.0,
                      month_orders_list=None,
                      prev_carry=None):
        ws.append([])  # разделитель

        tr = ws.max_row + 1
        ws.append([
            f"Торговый результат{label_suffix}:", "", "", "",
            f"=SUM(E{data_start}:E{data_end})",
            "",
            f"=SUM(G{data_start}:G{data_end})",
            f"=SUM(H{data_start}:H{data_end})",
            f"=SUM(I{data_start}:I{data_end})",
            "",
            f"=SUM(K{data_start}:K{data_end})",
            f"=SUM(L{data_start}:L{data_end})",
            f"=SUM(M{data_start}:M{data_end})",
        ])
        tr = ws.max_row

        # ── LIFO курс остатка — работает и при отрицательном остатке ──
        if month_orders_list is not None:
            total_buy_qty  = sum(float(o.amount or 0) for o in month_orders_list if o.operation_type == 'BUY')
            total_sell_qty = sum(float(o.amount or 0) for o in month_orders_list if o.operation_type == 'SELL')
            total_exch     = sum(
                float(o.amount or 0) * float(o.exchange_commission_rate or 0) / 100
                for o in month_orders_list if o.operation_type == 'SELL'
            )
            carry_qty  = float(prev_carry[0]) if prev_carry else 0.0
            remainder_qty = total_buy_qty + carry_qty - total_sell_qty - total_exch
            lifo_price    = _calc_lifo_price(month_orders_list, remainder_qty, prev_carry=prev_carry)
        else:
            lifo_price    = None
            remainder_qty = None
        # ──────────────────────────────────────────────────────────────

        ost = ws.max_row + 1
        ws.append([
            f"Остаток (нереализованная ЦВ){label_suffix}:", "", "", "",
            f"=E{tr}-I{tr}-M{tr}",
            lifo_price if lifo_price is not None else f"=IFERROR(G{tr}/E{tr},0)",
            f"=E{ost}*F{ost}",
            0, "", "", "", "", ""
        ])
        ost = ws.max_row

        rr = ws.max_row + 1
        ws.append([
            f"Реализованный результат{label_suffix}:", "", "", "",
            f"=I{tr}",
            "",
            f"=G{tr}-G{ost}",
            "",
            f"=I{tr}",
            last_sell_price,
            f"=K{tr}",
            "", ""
        ])
        rr = ws.max_row

        ws.append([
            f"Прибыль{label_suffix}:", "", "", "",
            f"=K{rr}-G{rr}-H{tr}-L{tr}-M{tr}*J{rr}",
            "", "", "", "", "", "", "", ""
        ])

        ws.append([])  # разделитель

        return ost, remainder_qty, lifo_price

    # =====================================================================
    # Обход ордеров — группировка по МСК месяцам
    # =====================================================================
    idx             = 0
    last_sell_price = 0.0
    prev_ost_row    = None
    prev_carry      = None  # (qty, price) перенос остатка с прошлого месяца

    def get_month_key(o):
        dt_msk = o.created_at.astimezone(MSK)
        return (dt_msk.year, dt_msk.month)

    if multi_month:
        is_first_month = True

        for (yr, mo), month_orders in _groupby(orders_list, key=get_month_key):
            month_list = list(month_orders)

            data_start = ws.max_row + 1

            if not is_first_month and prev_ost_row is not None:
                carry_label = f"Остаток с {_month_name(mo - 1 if mo > 1 else 12)} {yr if mo > 1 else yr - 1}"
                write_carry_row(carry_label, prev_ost_row)

            for o in month_list:
                is_hist = o.exchange_type == "Остаток (до 1 фев)"
                price_f = float(o.price or 0)

                if not is_hist:
                    idx += 1

                write_order_row(o, idx if not is_hist else 0, is_historical=is_hist)

                if o.operation_type == 'SELL' and price_f > 0:
                    last_sell_price = price_f

            data_end = ws.max_row
            prev_ost_row, new_remainder, new_price = write_summary(
                f" за {_month_name(mo)} {yr}",
                data_start, data_end,
                last_sell_price,
                month_orders_list=month_list,
                prev_carry=prev_carry
            )

            # Сохраняем перенос для следующего месяца
            if new_remainder is not None and new_price is not None and new_price > 0:
                prev_carry = (new_remainder, new_price)
            elif prev_carry is not None:
                prev_carry = prev_carry  # оставляем прошлый если не смогли посчитать

            is_first_month = False

    else:
        data_start = ws.max_row + 1

        for o in orders_list:
            is_hist = o.exchange_type == "Остаток (до 1 фев)"
            price_f = float(o.price or 0)

            if not is_hist:
                idx += 1

            write_order_row(o, idx if not is_hist else 0, is_historical=is_hist)

            if o.operation_type == 'SELL' and price_f > 0:
                last_sell_price = price_f

        data_end = ws.max_row
        write_summary(
            "",
            data_start, data_end,
            last_sell_price,
            month_orders_list=orders_list,
            prev_carry=None
        )

    # =====================================================================
    # Оформление
    # =====================================================================
    summary_kw = ('торговый', 'реализованный', 'прибыль', 'остаток')

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=13):
        for cell in row:
            cell.border = thin_border
            val = str(cell.value or '').lower()
            if cell.column == 1 and any(k in val for k in summary_kw):
                cell.alignment = left_align
                cell.font      = bold_font
            else:
                cell.alignment = Alignment(horizontal='center')

    col_widths = {
        'A': 40, 'B': 14, 'C': 24, 'D': 10,
        'E': 14, 'F': 12, 'G': 18, 'H': 18,
        'I': 14, 'J': 12, 'K': 18, 'L': 18, 'M': 20
    }
    for k, v in col_widths.items():
        ws.column_dimensions[k].width = v

    currency_suffix = f"_{currency}" if currency else ""
    filename = f"report_{user_id}{currency_suffix}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def admin_login(request):
    """Вход в админку"""
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('admin_users')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user.is_superuser:
                login(request, user)
                return redirect(request.GET.get('next', 'admin_users'))
            else:
                messages.error(request, "Нет прав администратора.")
        else:
            # Ошибки формы автоматически будут доступны в шаблоне
            pass
    else:
        form = AuthenticationForm()
    
    # Шаблон лежит в custom_admin/login.html
    return render(request, 'custom_admin/login.html', {'form': form})


@login_required(login_url='admin_login')
def admin_logout(request):
    """Выход из админки"""
    logout(request)
    return redirect('admin_login')


def _progressive_ndfl(base: float) -> float:
    """
    Прогрессивная шкала НДФЛ — БЕЗ ИЗМЕНЕНИЙ.
    """
    if base <= 0:
        return 0.0

    brackets = [
        (2_400_000,    0.13),
        (5_000_000,    0.15),
        (20_000_000,   0.18),
        (50_000_000,   0.20),
        (float('inf'), 0.22),
    ]

    ndfl = 0.0
    prev = 0.0
    for limit, rate in brackets:
        if base <= prev:
            break
        taxable = min(base, limit) - prev
        ndfl += taxable * rate
        prev = limit

    return ndfl


def _calc_ndfl(gross: float, tax_type: str,
               sell_cost: float = 0.0,
               ytd_base: float = 0.0) -> float:
    """
    Расчёт налога за текущий месяц — БЕЗ ИЗМЕНЕНИЙ.
    """
    t = str(tax_type or '').strip().upper()

    if t in ('USN_INCOME', 'USN_INCOME_OUTCOME'):
        return max(0.0, sell_cost) * 0.01

    if gross <= 0:
        return 0.0

    ndfl_total = _progressive_ndfl(ytd_base + gross)
    ndfl_prev  = _progressive_ndfl(ytd_base)
    return max(0.0, ndfl_total - ndfl_prev)


def _get_ytd_base(user, year: int, month: int, share_key: str) -> float:
    """
    Накопленная налоговая база (gross - расходы) за все месяцы года ДО текущего.
    Используется для нарастающего итога НДФЛ по ОСН.
    """
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo('Europe/Moscow')

    ytd_base = 0.0

    for mo in range(1, month):
        ms_start = datetime(year, mo, 1, tzinfo=MSK)
        ms_end   = datetime(year, mo + 1, 1, tzinfo=MSK) if mo < 12 else datetime(year + 1, 1, 1, tzinfo=MSK)
        mo_key   = f"{year}-{str(mo).zfill(2)}"

        calc    = _calc_month_profit_filtered(user, ms_start, ms_end)
        gross_mo = calc['gross']

        if calc['month_buy_qty'] == 0 and calc['month_sell_qty'] == 0:
            manual = MonthlyManualEntry.objects.filter(user=user, month=mo_key).first()
            if manual:
                gross_mo = float(manual.gross)

        expenses_mo = float(
            UserExpense.objects.filter(user=user, month=mo_key)
            .aggregate(Sum('amount'))['amount__sum'] or 0
        )

        # База = gross - расходы (то же что gross_net в текущем месяце)
        ytd_base += max(0.0, gross_mo - expenses_mo)

    return ytd_base


def _calc_lifo_price(buy_orders, remainder_qty, prev_avg_price=0.0):
    """
    Курс остатка по методу LIFO (последние закупки первыми) — как в Excel-отчёте.
    buy_orders — список BUY ордеров месяца в хронологическом порядке.
    remainder_qty — итоговый остаток (qty) на конец месяца.
    prev_avg_price — курс переходящего остатка с прошлого месяца (если не хватает покупок месяца).
    """
    abs_qty = abs(remainder_qty)
    if abs_qty == 0:
        return 0.0

    buys = []
    for o in reversed(buy_orders):
        qty = float(o.amount or 0)
        if qty <= 0:
            continue
        cost = float(o.cost or 0)
        price = float(o.price or 0)
        unit_price = cost / qty if qty > 0 else price
        buys.append((qty, unit_price))

    if prev_avg_price > 0:
        buys.append((float('inf'), prev_avg_price))

    if not buys:
        return 0.0

    total_cost = 0.0
    need = abs_qty
    for qty, unit_price in buys:
        if need <= 0:
            break
        take = min(qty, need)
        total_cost += take * unit_price
        need -= take

    if need > 0 and buys:
        total_cost += need * buys[-1][1]

    return round(total_cost / abs_qty, 4) if abs_qty > 0 else 0.0


def _calc_currency(user, month_start, month_end,
                   currency='USDT',
                   exchange_filter='', bank_filter_id=''):
    """
    Расчёт прибыли за месяц по одной валюте (USDT или TON).
    Средневзвешенный курс остатка.
    Комиссия биржи считается в рублях по курсу каждого ордера отдельно.
    """
    prev_orders = Order.objects.filter(
        user=user,
        created_at__gte=SYSTEM_START,
        created_at__lt=month_start,
        currency=currency
    ).order_by('created_at')

    prev_buy_qty   = 0.0
    prev_buy_cost  = 0.0
    prev_sell_qty  = 0.0
    prev_exch_comm = 0.0

    for o in prev_orders:
        amt = float(o.amount or 0)
        if o.operation_type == 'BUY':
            prev_buy_qty  += amt
            prev_buy_cost += float(o.cost or 0)
        else:
            exch_rate = float(o.exchange_commission_rate or 0)
            prev_sell_qty  += amt
            prev_exch_comm += amt * exch_rate / 100

    prev_balance_qty  = prev_buy_qty - prev_sell_qty - prev_exch_comm
    prev_avg_price    = (prev_buy_cost / prev_buy_qty) if prev_buy_qty > 0 else 0.0
    prev_balance_cost = prev_balance_qty * prev_avg_price

    month_orders = Order.objects.filter(
        user=user,
        created_at__gte=month_start,
        created_at__lt=month_end,
        currency=currency
    )
    if exchange_filter:
        month_orders = month_orders.filter(exchange_type__icontains=exchange_filter)
    if bank_filter_id:
        month_orders = month_orders.filter(bank_detail_id=bank_filter_id)

    month_orders = month_orders.order_by('created_at')

    month_buy_qty       = 0.0; month_buy_cost  = 0.0
    month_buy_comm      = 0.0
    month_sell_qty      = 0.0; month_sell_cost = 0.0
    month_sell_comm     = 0.0
    month_exch_comm     = 0.0
    month_exch_comm_rub = 0.0

    for o in month_orders:
        amt        = float(o.amount or 0)
        cost       = float(o.cost   or 0)
        price      = float(o.price  or 0)
        comm_val   = float(o.commission or 0)
        total_comm = cost * comm_val / 100 if o.commission_type == 'PERCENT' else comm_val

        if o.operation_type == 'BUY':
            month_buy_qty  += amt
            month_buy_cost += cost
            month_buy_comm += total_comm
        else:
            exch_rate   = float(o.exchange_commission_rate or 0)
            exch_crypto = amt * exch_rate / 100

            month_sell_qty      += amt
            month_sell_cost     += cost
            month_sell_comm     += total_comm
            month_exch_comm     += exch_crypto
            month_exch_comm_rub += exch_crypto * price

    total_buy_qty  = prev_balance_qty  + month_buy_qty
    total_buy_cost = prev_balance_cost + month_buy_cost

    remainder_qty_display = total_buy_qty - month_sell_qty - month_exch_comm

    # ── LIFO курс остатка (как в Excel-отчёте) ──────────────────────
    month_buy_orders_list = [o for o in month_orders if o.operation_type == 'BUY']
    lifo_price = _calc_lifo_price(
        month_buy_orders_list, remainder_qty_display, prev_avg_price=prev_avg_price
    )
    avg_price      = lifo_price
    remainder_cost = remainder_qty_display * avg_price
    eq_buy_cost    = total_buy_cost - remainder_cost
    # ──────────────────────────────────────────────────────────────────

    if month_sell_qty == 0:
        gross = 0.0
    else:
        gross = month_sell_cost - eq_buy_cost - month_buy_comm - month_sell_comm - month_exch_comm_rub

    return {
        'currency':              currency,
        'prev_balance_qty':      prev_balance_qty,
        'prev_balance_cost':     prev_balance_cost,
        'prev_avg_price':        round(prev_avg_price, 4),
        'month_buy_qty':         month_buy_qty,
        'month_buy_cost':        month_buy_cost,
        'month_buy_comm':        month_buy_comm,
        'month_sell_qty':        month_sell_qty,
        'month_sell_cost':       month_sell_cost,
        'month_sell_comm':       month_sell_comm,
        'month_exch_comm':       month_exch_comm,
        'month_exch_comm_rub':   month_exch_comm_rub,
        'total_buy_qty':         total_buy_qty,
        'total_buy_cost':        total_buy_cost,
        'remainder_qty_display': remainder_qty_display,
        'remainder_cost':        remainder_cost,
        'avg_price':             round(avg_price, 4),
        'eq_buy_cost':           eq_buy_cost,
        'gross':                 gross,
    }


def _calc_month_profit_filtered(user, month_start, month_end,
                                exchange_filter='', bank_filter_id=''):
    """
    Считает прибыль за месяц раздельно по USDT и TON.
    """
    usdt = _calc_currency(user, month_start, month_end,
                          currency='USDT',
                          exchange_filter=exchange_filter,
                          bank_filter_id=bank_filter_id)

    ton  = _calc_currency(user, month_start, month_end,
                          currency='TON',
                          exchange_filter=exchange_filter,
                          bank_filter_id=bank_filter_id)

    gross = usdt['gross'] + ton['gross']

    return {
        'prev_balance_qty':      usdt['prev_balance_qty'],
        'prev_balance_cost':     usdt['prev_balance_cost'],
        'month_buy_qty':         usdt['month_buy_qty']   + ton['month_buy_qty'],
        'month_buy_cost':        usdt['month_buy_cost']  + ton['month_buy_cost'],
        'month_buy_comm':        usdt['month_buy_comm']  + ton['month_buy_comm'],
        'month_sell_qty':        usdt['month_sell_qty']  + ton['month_sell_qty'],
        'month_sell_cost':       usdt['month_sell_cost'] + ton['month_sell_cost'],
        'month_sell_comm':       usdt['month_sell_comm'] + ton['month_sell_comm'],
        'month_exch_usdt':       usdt['month_exch_comm'] + ton['month_exch_comm'],
        'total_buy_qty':         usdt['total_buy_qty'],
        'total_buy_cost':        usdt['total_buy_cost'],
        'remainder_qty_display': usdt['remainder_qty_display'],
        'remainder_cost':        usdt['remainder_cost'],
        'eq_buy_cost':           usdt['eq_buy_cost'],
        'gross':                 gross,
        'usdt':                  usdt,
        'ton':                   ton,
    }



def _calc_lifo_price_excel(month_buy_orders, remainder_qty, prev_carry=None):
    """
    Точная копия логики из export_excel_report._calc_lifo_price.
    month_buy_orders — BUY ордера ТЕКУЩЕГО месяца (объекты с .amount/.cost/.price).
    prev_carry — (qty, price) перенос остатка с предыдущего месяца.
    """
    abs_qty = abs(remainder_qty)
    if abs_qty == 0:
        return 0.0

    buys = []
    for o in reversed(month_buy_orders):
        qty = float(o.amount or 0)
        if qty <= 0:
            continue
        cost  = float(o.cost or 0)
        price = float(o.price or 0)
        unit_price = cost / qty if qty > 0 else price
        buys.append((qty, unit_price))

    if prev_carry is not None:
        carry_qty, carry_price = prev_carry
        if carry_qty != 0 and carry_price > 0:
            buys.append((abs(carry_qty), carry_price))

    if not buys:
        return 0.0

    total_cost = 0.0
    need = abs_qty
    for qty, unit_price in buys:
        if need <= 0:
            break
        take = min(qty, need)
        total_cost += take * unit_price
        need -= take

    if need > 0 and buys:
        total_cost += need * buys[-1][1]

    return round(total_cost / abs_qty, 4) if abs_qty > 0 else 0.0


def _calc_all_months_currency_from_orders(user_id, all_orders_by_user, currency, extra_month_keys=None):
    """
    Оптимизированный аналог _calc_currency_from_orders для страницы
    /profit/ (action=get_months): та же математика переноса остатка
    (LIFO по месяцам), но одним проходом ВПЕРЁД по всем месяцам сразу,
    вместо пересчёта всей цепочки с нуля для каждого целевого месяца
    отдельно. У активных трейдеров (десятки месяцев истории) это убирает
    квадратичный/кубический рост времени загрузки страницы.

    Никакого сохранённого кэша — считает заново из all_orders_by_user
    при каждом вызове, так что правка любого прошлого ордера сразу
    учитывается на следующей же загрузке страницы, как и раньше.

    extra_month_keys — (year, month), для которых нужен результат, даже
    если в этом месяце нет ни одного ордера этой валюты (например месяц
    с ручной записью без реальных сделок) — получают "пустой" результат
    с перенесённым остатком, как раньше делал фолбэк result is None
    в _calc_currency_from_orders.

    ВАЖНО: используется только там, где exchange_filter/bank_filter_id
    не нужны (get_months их не читает) — в отличие от оригинала фильтры
    тут не применяются вообще.
    """
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo('Europe/Moscow')

    orders = all_orders_by_user.get(user_id, [])
    orders_currency = sorted(
        [o for o in orders if o.currency == currency],
        key=lambda o: o.created_at
    )

    by_month = {}
    for o in orders_currency:
        dt_msk = o.created_at.astimezone(MSK)
        key = (dt_msk.year, dt_msk.month)
        by_month.setdefault(key, []).append(o)

    all_keys = set(by_month.keys()) | set(extra_month_keys or [])
    if not all_keys:
        return {}

    results = {}
    prev_carry = None  # (qty, price)

    for key in sorted(all_keys):
        month_list = by_month.get(key, [])

        buy_qty = sum(float(o.amount or 0) for o in month_list if o.operation_type == 'BUY')
        buy_cost = sum(float(o.cost or 0) for o in month_list if o.operation_type == 'BUY')
        buy_comm = sum(
            (float(o.cost or 0) * float(o.commission or 0) / 100
             if o.commission_type == 'PERCENT' else float(o.commission or 0))
            for o in month_list if o.operation_type == 'BUY'
        )
        sell_qty = sum(float(o.amount or 0) for o in month_list if o.operation_type == 'SELL')
        sell_cost = sum(float(o.cost or 0) for o in month_list if o.operation_type == 'SELL')
        sell_comm = sum(
            (float(o.cost or 0) * float(o.commission or 0) / 100
             if o.commission_type == 'PERCENT' else float(o.commission or 0))
            for o in month_list if o.operation_type == 'SELL'
        )
        exch_qty = sum(
            float(o.amount or 0) * float(o.exchange_commission_rate or 0) / 100
            for o in month_list if o.operation_type == 'SELL'
        )
        last_sell_price = 0.0
        for o in month_list:
            if o.operation_type == 'SELL' and float(o.price or 0) > 0:
                last_sell_price = float(o.price or 0)

        carry_qty = float(prev_carry[0]) if prev_carry else 0.0
        remainder_qty = buy_qty + carry_qty - sell_qty - exch_qty

        buy_orders_month = [o for o in month_list if o.operation_type == 'BUY']
        lifo_price = _calc_lifo_price_excel(buy_orders_month, remainder_qty, prev_carry=prev_carry)

        remainder_cost = remainder_qty * lifo_price
        total_buy_for_month = buy_cost + (carry_qty * (prev_carry[1] if prev_carry else 0.0))
        eq_buy_cost_month = total_buy_for_month - remainder_cost

        exch_comm_rub_month = exch_qty * last_sell_price

        if sell_qty == 0:
            gross_month = 0.0
        else:
            gross_month = sell_cost - eq_buy_cost_month - buy_comm - sell_comm - exch_comm_rub_month

        results[key] = {
            'currency':              currency,
            'prev_balance_qty':      carry_qty,
            'prev_balance_cost':     carry_qty * (prev_carry[1] if prev_carry else 0.0),
            'prev_avg_price':        round(prev_carry[1], 4) if prev_carry else 0.0,
            'month_buy_qty':         buy_qty,
            'month_buy_cost':        buy_cost,
            'month_buy_comm':        buy_comm,
            'month_sell_qty':        sell_qty,
            'month_sell_cost':       sell_cost,
            'month_sell_comm':       sell_comm,
            'month_exch_comm':       exch_qty,
            'month_exch_comm_rub':   exch_comm_rub_month,
            'total_buy_qty':         buy_qty + carry_qty,
            'total_buy_cost':        total_buy_for_month,
            'remainder_qty_display': remainder_qty,
            'remainder_cost':        remainder_cost,
            'avg_price':             round(lifo_price, 4),
            'eq_buy_cost':           eq_buy_cost_month,
            'gross':                 gross_month,
        }

        # Прокидываем перенос на следующий месяц (та же логика, что в оригинале)
        if remainder_qty != 0 and lifo_price > 0:
            prev_carry = (remainder_qty, lifo_price)
        elif prev_carry is not None:
            pass  # оставляем прошлый carry если не смогли посчитать

    return results


_EMPTY_CURRENCY_RESULT = {
    'currency': '', 'prev_balance_qty': 0.0, 'prev_balance_cost': 0.0,
    'prev_avg_price': 0.0, 'month_buy_qty': 0.0, 'month_buy_cost': 0.0,
    'month_buy_comm': 0.0, 'month_sell_qty': 0.0, 'month_sell_cost': 0.0,
    'month_sell_comm': 0.0, 'month_exch_comm': 0.0, 'month_exch_comm_rub': 0.0,
    'total_buy_qty': 0.0, 'total_buy_cost': 0.0, 'remainder_qty_display': 0.0,
    'remainder_cost': 0.0, 'avg_price': 0.0, 'eq_buy_cost': 0.0, 'gross': 0.0,
}


def _calc_all_months_profit_from_orders(user_id, all_orders_by_user, month_keys):
    """
    Оптимизированный аналог _calc_month_profit_from_orders для get_months —
    считает USDT и TON одним проходом каждую (через
    _calc_all_months_currency_from_orders) и отдаёт готовые комбинированные
    результаты сразу по всем month_keys, а не по одному месяцу за вызов.
    """
    usdt_all = _calc_all_months_currency_from_orders(user_id, all_orders_by_user, 'USDT', extra_month_keys=month_keys)
    ton_all  = _calc_all_months_currency_from_orders(user_id, all_orders_by_user, 'TON', extra_month_keys=month_keys)

    result = {}
    for key in month_keys:
        usdt = usdt_all.get(key, _EMPTY_CURRENCY_RESULT)
        ton  = ton_all.get(key, _EMPTY_CURRENCY_RESULT)
        result[key] = {
            'prev_balance_qty':      usdt['prev_balance_qty'],
            'prev_balance_cost':     usdt['prev_balance_cost'],
            'month_buy_qty':         usdt['month_buy_qty']   + ton['month_buy_qty'],
            'month_buy_cost':        usdt['month_buy_cost']  + ton['month_buy_cost'],
            'month_buy_comm':        usdt['month_buy_comm']  + ton['month_buy_comm'],
            'month_sell_qty':        usdt['month_sell_qty']  + ton['month_sell_qty'],
            'month_sell_cost':       usdt['month_sell_cost'] + ton['month_sell_cost'],
            'month_sell_comm':       usdt['month_sell_comm'] + ton['month_sell_comm'],
            'month_exch_usdt':       usdt['month_exch_comm'] + ton['month_exch_comm'],
            'total_buy_qty':         usdt['total_buy_qty'],
            'total_buy_cost':        usdt['total_buy_cost'],
            'remainder_qty_display': usdt['remainder_qty_display'],
            'remainder_cost':        usdt['remainder_cost'],
            'eq_buy_cost':           usdt['eq_buy_cost'],
            'gross':                 usdt['gross'] + ton['gross'],
            'usdt':                  usdt,
            'ton':                   ton,
        }
    return result


def _calc_currency_from_orders(user_id, month_start, month_end,
                                all_orders_by_user,
                                currency='USDT',
                                exchange_filter='', bank_filter_id=''):
    """
    Расчёт прибыли по одной валюте — ТОЧНАЯ копия логики Excel-отчёта.
    LIFO по месяцам с rolling-переносом остатка (prev_carry),
    комиссия биржи считается по ПОСЛЕДНЕМУ курсу продажи месяца (как в Excel).

    Чтобы получить prev_carry на начало month_start, рекурсивно проходим
    все месяцы от SYSTEM_START.
    """
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo('Europe/Moscow')

    orders = all_orders_by_user.get(user_id, [])
    orders_currency = sorted(
        [o for o in orders if o.currency == currency],
        key=lambda o: o.created_at
    )

    if not orders_currency:
        return {
            'currency': currency, 'prev_balance_qty': 0.0, 'prev_balance_cost': 0.0,
            'prev_avg_price': 0.0, 'month_buy_qty': 0.0, 'month_buy_cost': 0.0,
            'month_buy_comm': 0.0, 'month_sell_qty': 0.0, 'month_sell_cost': 0.0,
            'month_sell_comm': 0.0, 'month_exch_comm': 0.0, 'month_exch_comm_rub': 0.0,
            'total_buy_qty': 0.0, 'total_buy_cost': 0.0, 'remainder_qty_display': 0.0,
            'remainder_cost': 0.0, 'avg_price': 0.0, 'eq_buy_cost': 0.0, 'gross': 0.0,
        }

    # Группируем ВСЕ ордера валюты по месяцам (МСК) от начала до month_end
    by_month = {}
    for o in orders_currency:
        if o.created_at >= month_end:
            break
        dt_msk = o.created_at.astimezone(MSK)
        key = (dt_msk.year, dt_msk.month)
        by_month.setdefault(key, []).append(o)

    target_key = (month_start.year, month_start.month)

    prev_carry = None  # (qty, price)
    result = None

    for key in sorted(by_month.keys()):
        month_list = by_month[key]
        is_target  = (key == target_key)

        # Фильтры применяются только к целевому месяцу (как в оригинале)
        filtered_list = month_list
        if is_target:
            filtered_list = [
                o for o in month_list
                if (not exchange_filter or exchange_filter.lower() in (o.exchange_type or '').lower())
                and (not bank_filter_id or str(getattr(o, 'bank_detail_id', '')) == str(bank_filter_id))
            ]

        buy_qty = sum(float(o.amount or 0) for o in filtered_list if o.operation_type == 'BUY')
        buy_cost = sum(float(o.cost or 0) for o in filtered_list if o.operation_type == 'BUY')
        buy_comm = sum(
            (float(o.cost or 0) * float(o.commission or 0) / 100
             if o.commission_type == 'PERCENT' else float(o.commission or 0))
            for o in filtered_list if o.operation_type == 'BUY'
        )
        sell_qty = sum(float(o.amount or 0) for o in filtered_list if o.operation_type == 'SELL')
        sell_cost = sum(float(o.cost or 0) for o in filtered_list if o.operation_type == 'SELL')
        sell_comm = sum(
            (float(o.cost or 0) * float(o.commission or 0) / 100
             if o.commission_type == 'PERCENT' else float(o.commission or 0))
            for o in filtered_list if o.operation_type == 'SELL'
        )
        exch_qty = sum(
            float(o.amount or 0) * float(o.exchange_commission_rate or 0) / 100
            for o in filtered_list if o.operation_type == 'SELL'
        )
        last_sell_price = 0.0
        for o in filtered_list:
            if o.operation_type == 'SELL' and float(o.price or 0) > 0:
                last_sell_price = float(o.price or 0)

        carry_qty = float(prev_carry[0]) if prev_carry else 0.0
        remainder_qty = buy_qty + carry_qty - sell_qty - exch_qty

        buy_orders_month = [o for o in filtered_list if o.operation_type == 'BUY']
        lifo_price = _calc_lifo_price_excel(buy_orders_month, remainder_qty, prev_carry=prev_carry)

        remainder_cost = remainder_qty * lifo_price
        total_buy_for_month = buy_cost + (carry_qty * (prev_carry[1] if prev_carry else 0.0))
        eq_buy_cost_month = total_buy_for_month - remainder_cost

        exch_comm_rub_month = exch_qty * last_sell_price

        if sell_qty == 0:
            gross_month = 0.0
        else:
            gross_month = sell_cost - eq_buy_cost_month - buy_comm - sell_comm - exch_comm_rub_month

        if is_target:
            result = {
                'currency':              currency,
                'prev_balance_qty':      carry_qty,
                'prev_balance_cost':     carry_qty * (prev_carry[1] if prev_carry else 0.0),
                'prev_avg_price':        round(prev_carry[1], 4) if prev_carry else 0.0,
                'month_buy_qty':         buy_qty,
                'month_buy_cost':        buy_cost,
                'month_buy_comm':        buy_comm,
                'month_sell_qty':        sell_qty,
                'month_sell_cost':       sell_cost,
                'month_sell_comm':       sell_comm,
                'month_exch_comm':       exch_qty,
                'month_exch_comm_rub':   exch_comm_rub_month,
                'total_buy_qty':         buy_qty + carry_qty,
                'total_buy_cost':        total_buy_for_month,
                'remainder_qty_display': remainder_qty,
                'remainder_cost':        remainder_cost,
                'avg_price':             round(lifo_price, 4),
                'eq_buy_cost':           eq_buy_cost_month,
                'gross':                 gross_month,
            }
            break

        # Прокидываем перенос на следующий месяц
        if remainder_qty != 0 and lifo_price > 0:
            prev_carry = (remainder_qty, lifo_price)
        elif prev_carry is not None:
            pass  # оставляем прошлый carry если не смогли посчитать

    if result is None:
        # ИСПРАВЛЕНО: целевой период (например "сегодня"/текущий месяц) мог
        # не попасть ни в один ключ by_month, если в нём ещё нет ни одного
        # ордера (юзер просто не успел ничего пробить в этом периоде).
        # Раньше это приводило к полному обнулению остатка, хотя реальный
        # перенос (prev_carry) из последнего месяца с активностью уже был
        # посчитан циклом выше и просто отбрасывался. Берём его.
        carry_qty   = float(prev_carry[0]) if prev_carry else 0.0
        carry_price = float(prev_carry[1]) if prev_carry else 0.0
        result = {
            'currency': currency, 'prev_balance_qty': carry_qty,
            'prev_balance_cost': carry_qty * carry_price,
            'prev_avg_price': round(carry_price, 4) if prev_carry else 0.0,
            'month_buy_qty': 0.0, 'month_buy_cost': 0.0,
            'month_buy_comm': 0.0, 'month_sell_qty': 0.0, 'month_sell_cost': 0.0,
            'month_sell_comm': 0.0, 'month_exch_comm': 0.0, 'month_exch_comm_rub': 0.0,
            'total_buy_qty': carry_qty, 'total_buy_cost': carry_qty * carry_price,
            'remainder_qty_display': carry_qty,
            'remainder_cost': carry_qty * carry_price,
            'avg_price': round(carry_price, 4) if prev_carry else 0.0,
            'eq_buy_cost': 0.0, 'gross': 0.0,
        }

    return result


def _calc_month_profit_from_orders(user_id, month_start, month_end,
                                    all_orders_by_user,
                                    exchange_filter='', bank_filter_id=''):
    """
    Аналог _calc_month_profit_filtered — работает без запросов к БД.
    Логика — ИДЕНТИЧНА оригиналу.
    """
    usdt = _calc_currency_from_orders(
        user_id, month_start, month_end, all_orders_by_user,
        currency='USDT',
        exchange_filter=exchange_filter,
        bank_filter_id=bank_filter_id,
    )
    ton = _calc_currency_from_orders(
        user_id, month_start, month_end, all_orders_by_user,
        currency='TON',
        exchange_filter=exchange_filter,
        bank_filter_id=bank_filter_id,
    )

    gross = usdt['gross'] + ton['gross']

    return {
        'prev_balance_qty':      usdt['prev_balance_qty'],
        'prev_balance_cost':     usdt['prev_balance_cost'],
        'month_buy_qty':         usdt['month_buy_qty']   + ton['month_buy_qty'],
        'month_buy_cost':        usdt['month_buy_cost']  + ton['month_buy_cost'],
        'month_buy_comm':        usdt['month_buy_comm']  + ton['month_buy_comm'],
        'month_sell_qty':        usdt['month_sell_qty']  + ton['month_sell_qty'],
        'month_sell_cost':       usdt['month_sell_cost'] + ton['month_sell_cost'],
        'month_sell_comm':       usdt['month_sell_comm'] + ton['month_sell_comm'],
        'month_exch_usdt':       usdt['month_exch_comm'] + ton['month_exch_comm'],
        'total_buy_qty':         usdt['total_buy_qty'],
        'total_buy_cost':        usdt['total_buy_cost'],
        'remainder_qty_display': usdt['remainder_qty_display'],
        'remainder_cost':        usdt['remainder_cost'],
        'eq_buy_cost':           usdt['eq_buy_cost'],
        'gross':                 gross,
        'usdt':                  usdt,
        'ton':                   ton,
    }


def _get_ytd_base_from_cache(user_id, year, month,
                              all_orders_by_user,
                              expenses_by_user_month,
                              manuals_by_user_month):
    """
    Аналог _get_ytd_base — без единого запроса к БД.
    Использует предзагруженные данные.
    Логика — ИДЕНТИЧНА оригиналу.
    """
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo('Europe/Moscow')

    ytd_base = 0.0

    for mo in range(1, month):
        ms_start = datetime(year, mo, 1, tzinfo=MSK)
        ms_end   = (datetime(year, mo + 1, 1, tzinfo=MSK)
                    if mo < 12 else datetime(year + 1, 1, 1, tzinfo=MSK))
        mo_key   = f"{year}-{str(mo).zfill(2)}"

        calc     = _calc_month_profit_from_orders(
            user_id, ms_start, ms_end, all_orders_by_user
        )
        gross_mo = calc['gross']

        # Если нет активности — берём manual (как в оригинале)
        if calc['month_buy_qty'] == 0 and calc['month_sell_qty'] == 0:
            manual = manuals_by_user_month.get((user_id, mo_key))
            if manual:
                gross_mo = float(manual.gross)

        # Расходы из кэша
        expenses_mo = expenses_by_user_month.get((user_id, mo_key), 0.0)

        ytd_base += max(0.0, gross_mo - expenses_mo)

    return ytd_base


def _calc_all_users_month_profit(users, year, month, all_orders_by_user,
                                  expenses_by_user_month, manuals_by_user_month):
    """
    Системная версия одной итерации основного цикла admin_profit_view —
    ТА ЖЕ математика (_calc_month_profit_from_orders + _calc_ndfl + доля),
    просто суммируется по всем юзерам сразу вместо построения таблицы
    user_stats. Существующая логика в admin_profit_view НЕ трогается —
    если поправить формулу тут, admin_profit_view не изменится, и наоборот
    (сознательное дублирование ради нулевого риска регресса на уже рабочей
    странице; оба места опираются на одни и те же module-level функции).

    Используется tasks.recompute_yearly_profit_summary_task для наполнения
    MonthlySystemProfitSummary — см. openspec/changes/add-monthly-profit-
    breakdown/design.md, решение №3.
    """
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo('Europe/Moscow')

    month_start = datetime(year, month, 1, tzinfo=MSK)
    month_end   = (datetime(year + 1, 1, 1, tzinfo=MSK)
                   if month == 12 else datetime(year, month + 1, 1, tzinfo=MSK))
    share_key = f"{year}-{str(month).zfill(2)}"

    totals = {
        'gross_all': 0.0, 'gross_usdt': 0.0, 'gross_ton': 0.0,
        'net_all':   0.0, 'net_usdt':   0.0, 'net_ton':   0.0,
        'share_all': 0.0, 'share_usdt': 0.0, 'share_ton': 0.0,
    }

    for u in users:
        uid = u.id
        calc = _calc_month_profit_from_orders(uid, month_start, month_end, all_orders_by_user)
        usdt = calc['usdt']
        ton  = calc['ton']

        tax_type      = str(getattr(u, 'tax_type', '') or '').strip().upper()
        share_percent = 20.0
        if u.profit_shares and isinstance(u.profit_shares, dict):
            share_percent = float(u.profit_shares.get(share_key, 20.0))

        month_expenses = expenses_by_user_month.get((uid, share_key), 0.0)

        gross_raw_usdt = usdt['gross']
        gross_raw_ton  = ton['gross']
        gross_raw_all  = gross_raw_usdt + gross_raw_ton

        has_activity = (calc['month_buy_qty'] > 0 or calc['month_sell_qty'] > 0)
        manual = manuals_by_user_month.get((uid, share_key))

        if not has_activity and manual:
            gross_raw_all  = manual.gross
            gross_raw_usdt = manual.gross
            gross_raw_ton  = 0.0

        pos_usdt  = max(0.0, gross_raw_usdt)
        pos_ton   = max(0.0, gross_raw_ton)
        pos_total = pos_usdt + pos_ton

        if pos_total > 0:
            usdt_expense_share = (pos_usdt / pos_total) * month_expenses
            ton_expense_share  = (pos_ton  / pos_total) * month_expenses
        else:
            usdt_expense_share = 0.0
            ton_expense_share  = 0.0

        gross_usdt = max(0.0, gross_raw_usdt - usdt_expense_share)
        gross_ton  = max(0.0, gross_raw_ton  - ton_expense_share)
        gross_all  = max(0.0, gross_raw_all  - month_expenses)

        if tax_type not in ('USN_INCOME', 'USN_INCOME_OUTCOME'):
            ytd_base = _get_ytd_base_from_cache(
                uid, year, month, all_orders_by_user,
                expenses_by_user_month, manuals_by_user_month,
            )
        else:
            ytd_base = 0.0

        sell_cost_all = calc['month_sell_cost'] if has_activity else (manual.sell_cost if manual else 0.0)

        ytd_usdt   = ytd_base * (pos_usdt / pos_total) if pos_total > 0 else ytd_base
        ndfl_usdt  = _calc_ndfl(gross_usdt, tax_type, sell_cost=usdt['month_sell_cost'], ytd_base=ytd_usdt)
        after_usdt = gross_usdt - ndfl_usdt
        share_usdt_val = after_usdt * (share_percent / 100) if after_usdt > 0 else 0.0
        net_usdt_val   = after_usdt - share_usdt_val

        ytd_ton   = ytd_base * (pos_ton / pos_total) if pos_total > 0 else 0.0
        ndfl_ton  = _calc_ndfl(gross_ton, tax_type, sell_cost=ton['month_sell_cost'], ytd_base=ytd_ton)
        after_ton = gross_ton - ndfl_ton
        share_ton_val = after_ton * (share_percent / 100) if after_ton > 0 else 0.0
        net_ton_val   = after_ton - share_ton_val

        ndfl_all  = _calc_ndfl(gross_all, tax_type, sell_cost=sell_cost_all, ytd_base=ytd_base)
        after_all = gross_all - ndfl_all
        share_all_val = after_all * (share_percent / 100) if after_all > 0 else 0.0
        net_all_val   = after_all - share_all_val

        totals['gross_all']  += gross_all
        totals['gross_usdt'] += gross_usdt
        totals['gross_ton']  += gross_ton
        totals['net_all']    += net_all_val
        totals['net_usdt']   += net_usdt_val
        totals['net_ton']    += net_ton_val
        totals['share_all']  += share_all_val
        totals['share_usdt'] += share_usdt_val
        totals['share_ton']  += share_ton_val

    return totals


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def export_uvedomlenie(request):
    """Скачивание XML-уведомления КНД 1110355 (сумма считается из ордеров)."""
    user_id = request.GET.get('user_id')
    year = int(request.GET.get('year') or datetime.now().year)
    period_key = request.GET.get('period', 'Q1')

    if period_key not in PERIODS:
        return JsonResponse({'error': 'Неверный период.'}, status=400)

    target = User.objects.filter(id=user_id).first()
    if not target:
        return JsonResponse({'error': 'Пользователь не найден.'}, status=404)

    try:
        filename, xml_bytes, info = build_uvedomlenie(target, year, period_key)
    except ValueError as e:
        # Понятная причина: не заполнено ОКТМО/код НО/ФИО, либо сумма = 0.
        return JsonResponse({'error': str(e)}, status=400)

    response = HttpResponse(xml_bytes, content_type='application/xml; charset=windows-1251')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _year_turnover_by_month(year, exchange_filter='', bank_filter_id=''):
    """
    Оборот (BUY/SELL сумма в рублях) по месяцам за год, по ВСЕМ юзерам сразу.
    Не считает прибыль/LIFO — только SUM(cost), это дешёво (12 запросов,
    один на месяц) в отличие от полного помесячного расчёта по каждому юзеру.
    """
    months_meta = [
        {'num': '01', 'name': 'Январь'},   {'num': '02', 'name': 'Февраль'},
        {'num': '03', 'name': 'Март'},     {'num': '04', 'name': 'Апрель'},
        {'num': '05', 'name': 'Май'},      {'num': '06', 'name': 'Июнь'},
        {'num': '07', 'name': 'Июль'},     {'num': '08', 'name': 'Август'},
        {'num': '09', 'name': 'Сентябрь'}, {'num': '10', 'name': 'Октябрь'},
        {'num': '11', 'name': 'Ноябрь'},   {'num': '12', 'name': 'Декабрь'},
    ]

    rows = []
    total = {'all_buy': 0.0, 'all_sell': 0.0, 'usdt_buy': 0.0, 'usdt_sell': 0.0, 'ton_buy': 0.0, 'ton_sell': 0.0}

    for m in months_meta:
        mo = int(m['num'])
        month_start = datetime(year, mo, 1, tzinfo=MSK)
        month_end   = (datetime(year + 1, 1, 1, tzinfo=MSK)
                       if mo == 12 else datetime(year, mo + 1, 1, tzinfo=MSK))

        qs = Order.objects.filter(
            created_at__gte=max(month_start, SYSTEM_START),
            created_at__lt=month_end,
        )
        if exchange_filter:
            qs = qs.filter(exchange_type__icontains=exchange_filter)
        if bank_filter_id:
            qs = qs.filter(bank_detail_id=bank_filter_id)

        row = {'num': m['num'], 'name': m['name'],
               'all_buy': 0.0, 'all_sell': 0.0, 'usdt_buy': 0.0, 'usdt_sell': 0.0, 'ton_buy': 0.0, 'ton_sell': 0.0}

        for grp in qs.values('operation_type', 'currency').annotate(s=Sum('cost')):
            val    = float(grp['s'] or 0)
            suffix = 'buy' if grp['operation_type'] == 'BUY' else 'sell'
            row[f'all_{suffix}'] += val
            total[f'all_{suffix}'] += val
            if grp['currency'] == 'USDT':
                row[f'usdt_{suffix}'] += val
                total[f'usdt_{suffix}'] += val
            elif grp['currency'] == 'TON':
                row[f'ton_{suffix}'] += val
                total[f'ton_{suffix}'] += val

        for key in ('all', 'usdt', 'ton'):
            row[f'{key}_turnover'] = round(row[f'{key}_buy'] + row[f'{key}_sell'], 2)
            row[f'{key}_buy']      = round(row[f'{key}_buy'], 2)
            row[f'{key}_sell']     = round(row[f'{key}_sell'], 2)

        rows.append(row)

    for key in ('all', 'usdt', 'ton'):
        total[f'{key}_turnover'] = round(total[f'{key}_buy'] + total[f'{key}_sell'], 2)
        total[f'{key}_buy']      = round(total[f'{key}_buy'], 2)
        total[f'{key}_sell']     = round(total[f'{key}_sell'], 2)

    return {'rows': rows, 'total': total}


def _build_user_month_stat(u, calc, ytd_base, year, month, month_start, month_end,
                            all_orders_by_user, expenses_by_user_month, expenses_list_by_user,
                            manuals_by_user_month, init_order_by_user,
                            exchange_filter='', bank_filter_id=''):
    """
    Финальная сборка одной строки admin-«Прибыли» (custom_admin/profit_list.html)
    из уже посчитанных calc (месячные суммы по USDT/TON) и ytd_base (накопленная
    база года для НДФЛ). Вынесено в отдельную функцию из admin_profit_view,
    чтобы её же использовала фоновая задача кэша (recompute_admin_profit_
    cache_task, см. history.md §46) — единая логика для живого расчёта
    (редкий путь с фильтром по бирже/банку) и для кэша (обычный путь без
    фильтра), чтобы они не разъехались, как уже случалось раньше с похожими
    параллельными расчётами (history.md §44.1).
    """
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo('Europe/Moscow')

    uid = u.id
    share_key = f"{year}-{str(month).zfill(2)}"

    usdt = calc['usdt']
    ton  = calc['ton']

    tax_type      = str(getattr(u, 'tax_type', '') or '').strip().upper()
    share_percent = 20.0
    if u.profit_shares and isinstance(u.profit_shares, dict):
        share_percent = float(u.profit_shares.get(share_key, 20.0))

    month_expenses = expenses_by_user_month.get((uid, share_key), 0.0)
    expenses_list  = expenses_list_by_user.get((uid, share_key), [])

    gross_raw_usdt = usdt['gross']
    gross_raw_ton  = ton['gross']
    gross_raw_all  = gross_raw_usdt + gross_raw_ton

    has_activity = (calc['month_buy_qty'] > 0 or calc['month_sell_qty'] > 0)

    # Берём manual запись если есть
    manual = manuals_by_user_month.get((uid, share_key))

    if not has_activity:
        if manual:
            gross_raw_all        = manual.gross
            gross_raw_usdt       = manual.gross
            gross_raw_ton        = 0.0
            manual_buy_cost      = manual.buy_cost
            manual_sell_cost     = manual.sell_cost
            manual_buy_comm      = manual.buy_comm
            manual_sell_comm     = manual.sell_comm
            manual_exch_comm_rub = manual.exch_comm_rub
            manual_buy_qty       = manual.buy_qty
            manual_sell_qty      = manual.sell_qty
        else:
            manual_buy_cost = manual_sell_cost = manual_buy_comm = 0.0
            manual_sell_comm = manual_exch_comm_rub = 0.0
            manual_buy_qty = manual_sell_qty = 0.0

        manual_crypto_balance = manual.remainder_qty if manual else init_order_by_user.get(uid, 0.0)
    else:
        manual_crypto_balance = None
        manual_buy_cost = manual_sell_cost = manual_buy_comm = 0.0
        manual_sell_comm = manual_exch_comm_rub = 0.0
        manual_buy_qty = manual_sell_qty = 0.0

    pos_usdt  = max(0.0, gross_raw_usdt)
    pos_ton   = max(0.0, gross_raw_ton)
    pos_total = pos_usdt + pos_ton

    if pos_total > 0:
        usdt_expense_share = (pos_usdt / pos_total) * month_expenses
        ton_expense_share  = (pos_ton  / pos_total) * month_expenses
    else:
        usdt_expense_share = 0.0
        ton_expense_share  = 0.0

    gross_usdt = max(0.0, gross_raw_usdt - usdt_expense_share)
    gross_ton  = max(0.0, gross_raw_ton  - ton_expense_share)
    gross_all  = max(0.0, gross_raw_all  - month_expenses)

    sell_cost_all = calc['month_sell_cost'] if has_activity else manual_sell_cost

    # ── USDT ──────────────────────────────────────────────────────
    ytd_usdt  = ytd_base * (pos_usdt / pos_total) if pos_total > 0 else ytd_base
    ndfl_usdt = _calc_ndfl(gross_usdt, tax_type,
                           sell_cost=usdt['month_sell_cost'],
                           ytd_base=ytd_usdt)
    after_usdt = gross_usdt - ndfl_usdt
    share_usdt = after_usdt * (share_percent / 100) if after_usdt > 0 else 0
    net_usdt   = after_usdt - share_usdt

    # ── TON ───────────────────────────────────────────────────────
    ytd_ton  = ytd_base * (pos_ton / pos_total) if pos_total > 0 else 0.0
    ndfl_ton = _calc_ndfl(gross_ton, tax_type,
                          sell_cost=ton['month_sell_cost'],
                          ytd_base=ytd_ton)
    after_ton = gross_ton - ndfl_ton
    share_ton = after_ton * (share_percent / 100) if after_ton > 0 else 0
    net_ton   = after_ton - share_ton

    # ── ИТОГО ─────────────────────────────────────────────────────
    ndfl_all  = _calc_ndfl(gross_all, tax_type,
                           sell_cost=sell_cost_all,
                           ytd_base=ytd_base)
    after_all = gross_all - ndfl_all
    share_all = after_all * (share_percent / 100) if after_all > 0 else 0
    net_all   = after_all - share_all

    bank_comm_rub      = calc['month_buy_comm'] + calc['month_sell_comm']
    usdt_bank_comm     = usdt['month_buy_comm'] + usdt['month_sell_comm']
    ton_bank_comm      = ton['month_buy_comm']  + ton['month_sell_comm']
    usdt_exch_comm_rub = usdt['month_exch_comm_rub']
    ton_exch_comm_rub  = ton['month_exch_comm_rub']
    all_exch_comm_rub  = usdt_exch_comm_rub + ton_exch_comm_rub

    has_activity_final = (
        has_activity
        or (uid, share_key) in manuals_by_user_month
    )

    # Топ-3 ордера по размеру комиссии банка за месяц (для тултипа)
    top_bank_comm_orders = []
    if has_activity:
        month_user_orders = [
            o for o in all_orders_by_user.get(uid, [])
            if month_start <= o.created_at < month_end
            and (not exchange_filter or exchange_filter.lower() in (o.exchange_type or '').lower())
            and (not bank_filter_id or str(o.bank_detail_id) == str(bank_filter_id))
        ]
        comm_rows = []
        for o in month_user_orders:
            cost     = float(o.cost or 0)
            comm_val = float(o.commission or 0)
            comm_rub = cost * comm_val / 100 if o.commission_type == 'PERCENT' else comm_val
            if comm_rub > 0:
                comm_rows.append({
                    'date':      o.created_at.astimezone(MSK).strftime('%d.%m %H:%M'),
                    'cost':      cost,
                    'comm':      comm_rub,
                    'op':        o.operation_type,
                    'comm_val':  comm_val,
                    'comm_type': o.commission_type,
                })
        comm_rows.sort(key=lambda r: r['comm'], reverse=True)
        top_bank_comm_orders = comm_rows[:3]

    # Финальные значения — из расчёта если есть ордера, иначе из manual
    final_buy_cost      = calc['month_buy_cost']  if has_activity else manual_buy_cost
    final_sell_cost     = calc['month_sell_cost'] if has_activity else manual_sell_cost
    final_buy_comm      = calc['month_buy_comm']  if has_activity else manual_buy_comm
    final_sell_comm     = calc['month_sell_comm'] if has_activity else manual_sell_comm
    final_exch_comm_rub = all_exch_comm_rub       if has_activity else manual_exch_comm_rub
    final_bank_comm     = bank_comm_rub            if has_activity else (manual_buy_comm + manual_sell_comm)
    final_buy_qty       = calc['month_buy_qty']   if has_activity else manual_buy_qty
    final_sell_qty      = calc['month_sell_qty']  if has_activity else manual_sell_qty

    return {
        'user':                  u,
        'crypto_balance':        (manual_crypto_balance
                                  if manual_crypto_balance is not None
                                  else calc['remainder_qty_display']),
        'prev_balance_qty':      calc['prev_balance_qty'],
        'prev_balance_cost':     calc['prev_balance_cost'],
        'month_buy_qty':         final_buy_qty,
        'month_buy_cost':        final_buy_cost,
        'month_sell_qty':        final_sell_qty,
        'month_sell_cost':       final_sell_cost,
        'turnover':              final_buy_cost + final_sell_cost,
        'usdt_balance_display':  (manual_crypto_balance
                                  if manual_crypto_balance is not None
                                  else usdt['remainder_qty_display']),
        'bank_comm_rub':         final_bank_comm,
        'exchange_comm_rub':     final_exch_comm_rub,
        'total_comm_rub':        final_bank_comm,
        'eq_buy_cost':           calc['eq_buy_cost'],
        'gross':                 gross_all,
        'gross_profit_positive': gross_all >= 0,
        'ndfl':                  ndfl_all,
        'tax_type':              tax_type,
        'share_percent':         share_percent,
        'share_amount':          share_all,
        'month_expenses':        month_expenses,
        'expenses_list':         expenses_list,
        'net_profit':            net_all,
        'net_profit_positive':   net_all >= 0,
        'has_activity':          has_activity_final,
        'usdt_bank_comm':        usdt_bank_comm,
        'ton_bank_comm':         ton_bank_comm,
        'usdt_exch_comm_rub':    usdt_exch_comm_rub,
        'ton_exch_comm_rub':     ton_exch_comm_rub,
        'all_exch_comm_rub':     final_exch_comm_rub,
        'top_bank_comm_orders':  top_bank_comm_orders,
        # Детализация USDT
        'usdt':       usdt,
        'usdt_avg_price_ru': format_ru_number(usdt.get('avg_price', 0)),
        'usdt_gross': gross_usdt,
        'usdt_ndfl':  ndfl_usdt,
        'usdt_share': share_usdt,
        'usdt_net':   net_usdt,
        # Детализация TON
        'ton':       ton,
        'ton_avg_price_ru': format_ru_number(ton.get('avg_price', 0)),
        'ton_gross': gross_ton,
        'ton_ndfl':  ndfl_ton,
        'ton_share': share_ton,
        'ton_net':   net_ton,
    }


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_profit_view(request):
    """
    Сводная таблица прибыли по всем пользователям за выбранный месяц.
    gross = выручка - себестоимость - комса банка - комса биржи - расходы пользователя
    ОСН — НДФЛ нарастающим итогом за год.
    УСН — 1% от оборота продаж.

    ОПТИМИЗАЦИЯ: вместо сотен запросов — 5 групповых запросов ДО цикла.
    """
    from zoneinfo import ZoneInfo
    from collections import defaultdict
    MSK = ZoneInfo('Europe/Moscow')

    import time
    t0 = time.time()
    now       = timezone.now()
    year_str  = request.GET.get('year',  str(now.year))
    month_str = request.GET.get('month', str(now.month).zfill(2))

    exchange_filter = request.GET.get('exchange', '')
    bank_filter_id  = request.GET.get('bank', '')

    # Вкладка "2026" (крайняя справа от месяцев) — сводный оборот по всей
    # системе (все юзеры сразу) помесячно за год. Отдельная лёгкая ветка —
    # не считает LIFO/прибыль по каждому юзеру (дорого, 142 юзера x 12
    # месяцев), только SUM(cost) по BUY/SELL за месяц, это и есть "оборот".
    if month_str == 'year':
        try:
            year = int(year_str)
        except ValueError:
            year = now.year
        year_data = _year_turnover_by_month(year, exchange_filter, bank_filter_id)

        # Прибыль (грязная/чистая трейдеров + доход системы) — НЕ считается
        # тут вживую (дорого, LIFO по каждому юзеру), а читается из кэша,
        # который наполняет tasks.recompute_yearly_profit_summary_task раз
        # в час. Если для месяца/валюты кэша ещё нет — has_profit=False,
        # шаблон должен показать явный плейсхолдер, а не молчаливый 0.
        profit_cache = {
            (row.month, row.currency): row
            for row in MonthlySystemProfitSummary.objects.filter(year=year)
        }
        profit_updated_at = None

        for row in year_data['rows']:
            mo = int(row['num'])
            for currency in ('all', 'usdt', 'ton'):
                cached = profit_cache.get((mo, currency))
                if cached:
                    row[f'{currency}_gross']      = float(cached.gross_traders)
                    row[f'{currency}_net']        = float(cached.net_traders)
                    row[f'{currency}_income']     = float(cached.system_income)
                    row[f'{currency}_has_profit'] = True
                    if profit_updated_at is None or cached.updated_at > profit_updated_at:
                        profit_updated_at = cached.updated_at
                else:
                    row[f'{currency}_gross']      = 0.0
                    row[f'{currency}_net']        = 0.0
                    row[f'{currency}_income']     = 0.0
                    row[f'{currency}_has_profit'] = False

        for currency in ('all', 'usdt', 'ton'):
            counted = [r for r in year_data['rows'] if r[f'{currency}_has_profit']]
            year_data['total'][f'{currency}_gross']  = round(sum(r[f'{currency}_gross']  for r in counted), 2)
            year_data['total'][f'{currency}_net']    = round(sum(r[f'{currency}_net']    for r in counted), 2)
            year_data['total'][f'{currency}_income'] = round(sum(r[f'{currency}_income'] for r in counted), 2)

        return render(request, 'custom_admin/profit_year.html', {
            'current_year':      year,
            'selected_exchange': exchange_filter,
            'selected_bank':     bank_filter_id,
            'default_banks':     BankDetail.objects.filter(is_deleted=False),
            'month_rows':        year_data['rows'],
            'year_total':        year_data['total'],
            'profit_updated_at': profit_updated_at,
        })

    try:
        year  = int(year_str)
        month = int(month_str)
    except ValueError:
        year  = now.year
        month = now.month

    month_str   = str(month).zfill(2)
    share_key   = f"{year}-{month_str}"
    month_start = datetime(year, month, 1, tzinfo=MSK)
    month_end   = (datetime(year + 1, 1, 1, tzinfo=MSK)
                   if month == 12 else datetime(year, month + 1, 1, tzinfo=MSK))
    year_start  = datetime(year, 1, 1, tzinfo=MSK)

    User  = get_user_model()
    users = list(User.objects.all().order_by('id'))

    # ================================================================
    #  КЭШ — до тяжёлой предзагрузки, чтобы знать, кого вообще ещё
    #  нужно считать вживую (см. history.md §46).
    # ================================================================
    # ОПТИМИЗАЦИЯ №2: без фильтра по бирже/банку (99% заходов на страницу)
    # читаем готовую строку из кэша (MonthlyUserProfitCache, наполняется
    # фоново раз в 15 минут — tasks.recompute_admin_profit_cache_task)
    # вместо полного пересчёта на каждый заход. Если кэша на юзера/месяц
    # ещё нет (юзер только что создан, фон не успел досчитать) — считаем
    # именно этого юзера вживую, не блокируя всю страницу.
    #
    # Фильтр по бирже/банку — редкий путь, там семантика другая (фильтр
    # только на целевой месяц, не на предыдущие для переноса остатка) —
    # кэш не используется вообще, все юзера считаются вживую, как раньше.
    has_filter = bool(exchange_filter or bank_filter_id)

    cache_rows = {}
    if not has_filter:
        cache_rows = {
            row.user_id: row.data
            for row in MonthlyUserProfitCache.objects.filter(year=year, month=month)
        }
    # Юзеры без готовой строки в кэше (для has_filter=True это ВСЕ юзеры —
    # cache_rows тогда пустой, что и даёт полную предзагрузку, как раньше).
    missing_user_ids = [u.id for u in users if u.id not in cache_rows]

    # ================================================================
    #  БЛОК ПРЕДЗАГРУЗКИ — только для missing_user_ids, а не для всех
    #  подряд. Если кэш покрывает всех (обычное дело) — missing_user_ids
    #  пуст, и Order.objects.filter(user_id__in=[]) — это ORM-заглушка
    #  без единого обращения к базе (Django сам коротко замыкает пустой
    #  __in). Раньше тут читалась вся база ордеров системы (~318 тыс.
    #  строк) на КАЖДЫЙ заход, независимо от того, нужны они были или нет.
    # ================================================================

    # 1. Ордера — только недостающих юзеров
    raw_orders = (
        Order.objects
        .filter(user_id__in=missing_user_ids, created_at__gte=SYSTEM_START, created_at__lt=month_end)
        .order_by('user_id', 'created_at')
        .values_list(
            'id', 'user_id', 'created_at', 'currency', 'operation_type',
            'amount', 'cost', 'price', 'commission', 'commission_type',
            'exchange_type', 'bank_detail_id', 'exchange_commission_rate',
        )
    )

    class _O:
        __slots__ = (
            'id', 'user_id', 'created_at', 'currency', 'operation_type',
            'amount', 'cost', 'price', 'commission', 'commission_type',
            'exchange_type', 'bank_detail_id', 'exchange_commission_rate',
        )
        def __init__(self, row):
            (self.id, self.user_id, self.created_at, self.currency,
             self.operation_type, self.amount, self.cost, self.price,
             self.commission, self.commission_type, self.exchange_type,
             self.bank_detail_id, self.exchange_commission_rate) = row

    all_orders_by_user = defaultdict(list)
    for row in raw_orders:
        o = _O(row)
        all_orders_by_user[o.user_id].append(o)

    # 2. UserExpense за текущий год — только недостающих юзеров
    raw_expenses = (
        UserExpense.objects
        .filter(user_id__in=missing_user_ids, month__startswith=str(year))
        .values('user_id', 'month', 'name', 'amount')
    )

    expenses_by_user_month = defaultdict(float)
    expenses_list_by_user  = defaultdict(list)
    for e in raw_expenses:
        key = (e['user_id'], e['month'])
        expenses_by_user_month[key] += float(e['amount'])
        expenses_list_by_user[key].append({'name': e['name'], 'amount': e['amount']})

    # 3. MonthlyManualEntry за текущий год — только недостающих юзеров
    raw_manuals = (
        MonthlyManualEntry.objects
        .filter(user_id__in=missing_user_ids, month__startswith=str(year))
        .values('user_id', 'month', 'gross',
                'buy_qty', 'buy_cost', 'buy_comm',
                'sell_qty', 'sell_cost', 'sell_comm',
                'exch_comm_rub', 'remainder_qty',
                'remainder_price', 'remainder_cost', 'source')
    )

    class _Manual:
        __slots__ = ('gross', 'buy_qty', 'buy_cost', 'buy_comm',
                     'sell_qty', 'sell_cost', 'sell_comm',
                     'exch_comm_rub', 'remainder_qty',
                     'remainder_price', 'remainder_cost', 'source')
        def __init__(self, m):
            self.gross           = float(m.get('gross', 0) or 0)
            self.buy_qty         = float(m.get('buy_qty', 0) or 0)
            self.buy_cost        = float(m.get('buy_cost', 0) or 0)
            self.buy_comm        = float(m.get('buy_comm', 0) or 0)
            self.sell_qty        = float(m.get('sell_qty', 0) or 0)
            self.sell_cost       = float(m.get('sell_cost', 0) or 0)
            self.sell_comm       = float(m.get('sell_comm', 0) or 0)
            self.exch_comm_rub   = float(m.get('exch_comm_rub', 0) or 0)
            self.remainder_qty   = float(m.get('remainder_qty', 0) or 0)
            self.remainder_price = float(m.get('remainder_price', 0) or 0)
            self.remainder_cost  = float(m.get('remainder_cost', 0) or 0)
            self.source          = m.get('source', 'manual')

    manuals_by_user_month = {}
    for m in raw_manuals:
        manuals_by_user_month[(m['user_id'], m['month'])] = _Manual(m)

    # 4. init_order — только недостающих юзеров
    init_orders = (
        Order.objects
        .filter(user_id__in=missing_user_ids, exchange_type="Остаток (до 1 фев)")
        .values('user_id', 'amount')
    )
    init_order_by_user = {o['user_id']: float(o['amount']) for o in init_orders}

    # 5. BankDetail — не связано с юзерами, всегда нужен (для фильтра на странице)
    default_banks = BankDetail.objects.all()

    # ================================================================
    #  ОСНОВНОЙ ЦИКЛ
    # ================================================================

    ytd_month_keys = {(year, m) for m in range(1, month + 1)}

    user_stats = []

    for u in users:
        uid = u.id

        if not has_filter and uid in cache_rows:
            stat = dict(cache_rows[uid])
            stat['user'] = u
            user_stats.append(stat)
            continue

        if has_filter:
            calc = _calc_month_profit_from_orders(
                uid, month_start, month_end, all_orders_by_user,
                exchange_filter=exchange_filter,
                bank_filter_id=bank_filter_id,
            )
        else:
            all_months_calc_u = _calc_all_months_profit_from_orders(uid, all_orders_by_user, ytd_month_keys)
            calc = all_months_calc_u[(year, month)]

        # YTD-база для НДФЛ считается тут (а не в _build_user_month_stat),
        # т.к. у has_filter-варианта свой источник (живой пересчёт с
        # фильтром), а у обычного — префиксная сумма по all_months_calc_u,
        # которая есть только в этой ветке.
        tax_type_for_ytd = str(getattr(u, 'tax_type', '') or '').strip().upper()
        if tax_type_for_ytd not in ('USN_INCOME', 'USN_INCOME_OUTCOME'):
            if has_filter:
                ytd_base = _get_ytd_base_from_cache(
                    uid, year, month,
                    all_orders_by_user,
                    expenses_by_user_month,
                    manuals_by_user_month,
                )
            else:
                ytd_base = 0.0
                for m in range(1, month):
                    calc_mo = all_months_calc_u[(year, m)]
                    gross_mo = calc_mo['gross']
                    if calc_mo['month_buy_qty'] == 0 and calc_mo['month_sell_qty'] == 0:
                        manual_mo = manuals_by_user_month.get((uid, f"{year}-{str(m).zfill(2)}"))
                        if manual_mo:
                            gross_mo = float(manual_mo.gross)
                    expenses_mo = expenses_by_user_month.get((uid, f"{year}-{str(m).zfill(2)}"), 0.0)
                    ytd_base += max(0.0, gross_mo - expenses_mo)
        else:
            ytd_base = 0.0

        stat = _build_user_month_stat(
            u, calc, ytd_base, year, month, month_start, month_end,
            all_orders_by_user, expenses_by_user_month, expenses_list_by_user,
            manuals_by_user_month, init_order_by_user,
            exchange_filter=exchange_filter, bank_filter_id=bank_filter_id,
        )
        user_stats.append(stat)

    # summary
    summary = {
        'buy_sum':    sum(s['month_buy_cost']  for s in user_stats),
        'sell_sum':   sum(s['month_sell_cost'] for s in user_stats),
        'turnover':   sum(s['month_sell_cost'] + s['month_buy_cost'] for s in user_stats),
        'gross':      sum(s['gross']           for s in user_stats),
        'ndfl':       sum(s['ndfl']            for s in user_stats),
        'our_share':  sum(s['share_amount']    for s in user_stats),
        'expenses':   sum(s['month_expenses']  for s in user_stats),
        'net':        sum(s['net_profit']      for s in user_stats),
        'usdt_sell':  sum(s['usdt']['month_sell_cost'] for s in user_stats),
        'usdt_buy':   sum(s['usdt']['month_buy_cost']  for s in user_stats),
        'usdt_gross': sum(s['usdt_gross']              for s in user_stats),
        'usdt_share': sum(s['usdt_share']              for s in user_stats),
        'usdt_net':   sum(s['usdt_net']                for s in user_stats),
        'ton_sell':   sum(s['ton']['month_sell_cost']  for s in user_stats),
        'ton_buy':    sum(s['ton']['month_buy_cost']   for s in user_stats),
        'ton_gross':  sum(s['ton_gross']               for s in user_stats),
        'ton_share':  sum(s['ton_share']               for s in user_stats),
        'ton_net':    sum(s['ton_net']                 for s in user_stats),
    }

    months_list = [
        {'num': '01', 'name': 'Янв'},  {'num': '02', 'name': 'Февр'},
        {'num': '03', 'name': 'Март'}, {'num': '04', 'name': 'Апр'},
        {'num': '05', 'name': 'Май'},  {'num': '06', 'name': 'Июнь'},
        {'num': '07', 'name': 'Июль'}, {'num': '08', 'name': 'Авг'},
        {'num': '09', 'name': 'Сент'}, {'num': '10', 'name': 'Окт'},
        {'num': '11', 'name': 'Нояб'}, {'num': '12', 'name': 'Дек'},
    ]

    current_month_name = next(
        (m['name'] for m in months_list if m['num'] == month_str), month_str
    )

    print(f"[profit_view] {time.time() - t0:.2f}s", flush=True)

    return render(request, 'custom_admin/profit_list.html', {
        'user_stats':          user_stats,
        'summary':             summary,
        'current_year':        year,
        'current_month':       month_str,
        'current_month_name':  current_month_name,
        'months_list':         months_list,
        'selected_exchange':   exchange_filter,
        'selected_bank':       bank_filter_id,
        'default_banks':       default_banks,
    })


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def api_user_yearly_profit(request):
    """
    AJAX endpoint: возвращает прибыль одного пользователя за все месяцы года.
    GET params: user_id, year
    """
    from zoneinfo import ZoneInfo
    from collections import defaultdict
    MSK = ZoneInfo('Europe/Moscow')

    user_id = request.GET.get('user_id')
    year    = int(request.GET.get('year', timezone.now().year))

    if not user_id:
        return JsonResponse({'error': 'no user_id'}, status=400)

    User = get_user_model()
    try:
        u = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'error': 'not found'}, status=404)

    # ── Предзагрузка данных пользователя ──────────────────────────
    year_end = datetime(year + 1, 1, 1, tzinfo=MSK)

    # Все ордера пользователя от SYSTEM_START до конца года
    raw_orders = (
        Order.objects
        .filter(user=u, created_at__gte=SYSTEM_START, created_at__lt=year_end)
        .order_by('created_at')
        .values_list(
            'id', 'user_id', 'created_at', 'currency', 'operation_type',
            'amount', 'cost', 'price', 'commission', 'commission_type',
            'exchange_type', 'bank_detail_id', 'exchange_commission_rate',
        )
    )

    class _O:
        __slots__ = (
            'id', 'user_id', 'created_at', 'currency', 'operation_type',
            'amount', 'cost', 'price', 'commission', 'commission_type',
            'exchange_type', 'bank_detail_id', 'exchange_commission_rate',
        )
        def __init__(self, row):
            (self.id, self.user_id, self.created_at, self.currency,
             self.operation_type, self.amount, self.cost, self.price,
             self.commission, self.commission_type, self.exchange_type,
             self.bank_detail_id, self.exchange_commission_rate) = row

    all_orders_by_user = defaultdict(list)
    for row in raw_orders:
        o = _O(row)
        all_orders_by_user[o.user_id].append(o)

    # Расходы за весь год
    raw_expenses = (
        UserExpense.objects
        .filter(user=u, month__startswith=str(year))
        .values('user_id', 'month', 'name', 'amount')
    )
    expenses_by_user_month = defaultdict(float)
    expenses_list_by_user  = defaultdict(list)
    for e in raw_expenses:
        key = (e['user_id'], e['month'])
        expenses_by_user_month[key] += float(e['amount'])
        expenses_list_by_user[key].append({'name': e['name'], 'amount': float(e['amount'])})

    # Manual entries за весь год — все поля
    raw_manuals = (
        MonthlyManualEntry.objects
        .filter(user=u, month__startswith=str(year))
        .values('user_id', 'month', 'gross',
                'buy_qty', 'buy_cost', 'buy_comm',
                'sell_qty', 'sell_cost', 'sell_comm',
                'exch_comm_rub', 'remainder_qty',
                'remainder_price', 'remainder_cost', 'source')
    )

    class _Manual:
        __slots__ = ('gross', 'buy_qty', 'buy_cost', 'buy_comm',
                     'sell_qty', 'sell_cost', 'sell_comm',
                     'exch_comm_rub', 'remainder_qty',
                     'remainder_price', 'remainder_cost', 'source')
        def __init__(self, m):
            self.gross           = float(m.get('gross', 0) or 0)
            self.buy_qty         = float(m.get('buy_qty', 0) or 0)
            self.buy_cost        = float(m.get('buy_cost', 0) or 0)
            self.buy_comm        = float(m.get('buy_comm', 0) or 0)
            self.sell_qty        = float(m.get('sell_qty', 0) or 0)
            self.sell_cost       = float(m.get('sell_cost', 0) or 0)
            self.sell_comm       = float(m.get('sell_comm', 0) or 0)
            self.exch_comm_rub   = float(m.get('exch_comm_rub', 0) or 0)
            self.remainder_qty   = float(m.get('remainder_qty', 0) or 0)
            self.remainder_price = float(m.get('remainder_price', 0) or 0)
            self.remainder_cost  = float(m.get('remainder_cost', 0) or 0)
            self.source          = m.get('source', 'manual')

    manuals_by_user_month = {}
    for m in raw_manuals:
        manuals_by_user_month[(m['user_id'], m['month'])] = _Manual(m)

    # init_order — остаток до 1 фев
    init_order = Order.objects.filter(
        user=u, exchange_type="Остаток (до 1 фев)"
    ).values('amount').first()
    init_order_amount = float(init_order['amount']) if init_order else 0.0

    months_list = [
        ('01', 'Январь'), ('02', 'Февраль'), ('03', 'Март'),
        ('04', 'Апрель'), ('05', 'Май'),     ('06', 'Июнь'),
        ('07', 'Июль'),   ('08', 'Август'),  ('09', 'Сентябрь'),
        ('10', 'Октябрь'),('11', 'Ноябрь'),  ('12', 'Декабрь'),
    ]

    tax_type      = str(getattr(u, 'tax_type', '') or '').strip().upper()
    uid           = u.id
    months_data   = []
    total_summary = {
        'sell': 0.0, 'buy': 0.0, 'gross': 0.0,
        'share': 0.0, 'net': 0.0,
    }

    for mo_num, mo_name in months_list:
        mo        = int(mo_num)
        share_key = f"{year}-{mo_num}"

        ms_start = datetime(year, mo, 1, tzinfo=MSK)
        ms_end   = (datetime(year + 1, 1, 1, tzinfo=MSK)
                    if mo == 12 else datetime(year, mo + 1, 1, tzinfo=MSK))

        calc = _calc_month_profit_from_orders(
            uid, ms_start, ms_end, all_orders_by_user
        )
        usdt = calc['usdt']
        ton  = calc['ton']

        share_percent = 20.0
        if u.profit_shares and isinstance(u.profit_shares, dict):
            share_percent = float(u.profit_shares.get(share_key, 20.0))

        month_expenses = expenses_by_user_month.get((uid, share_key), 0.0)

        gross_raw_usdt = usdt['gross']
        gross_raw_ton  = ton['gross']
        gross_raw_all  = gross_raw_usdt + gross_raw_ton

        has_activity = (calc['month_buy_qty'] > 0 or calc['month_sell_qty'] > 0)

        # Берём manual запись если есть
        manual = manuals_by_user_month.get((uid, share_key))

        if not has_activity:
            if manual:
                gross_raw_all        = manual.gross
                gross_raw_usdt       = manual.gross
                gross_raw_ton        = 0.0
                manual_buy_cost      = manual.buy_cost
                manual_sell_cost     = manual.sell_cost
                manual_buy_comm      = manual.buy_comm
                manual_sell_comm     = manual.sell_comm
                manual_exch_comm_rub = manual.exch_comm_rub
                manual_buy_qty       = manual.buy_qty
                manual_sell_qty      = manual.sell_qty
                manual_crypto_balance = manual.remainder_qty
            else:
                manual_buy_cost = manual_sell_cost = manual_buy_comm = 0.0
                manual_sell_comm = manual_exch_comm_rub = 0.0
                manual_buy_qty = manual_sell_qty = 0.0
                manual_crypto_balance = init_order_amount
        else:
            manual_crypto_balance = None
            manual_buy_cost = manual_sell_cost = manual_buy_comm = 0.0
            manual_sell_comm = manual_exch_comm_rub = 0.0
            manual_buy_qty = manual_sell_qty = 0.0

        gross_all = max(0.0, gross_raw_all - month_expenses)

        if tax_type not in ('USN_INCOME', 'USN_INCOME_OUTCOME'):
            ytd_base = _get_ytd_base_from_cache(
                uid, year, mo,
                all_orders_by_user,
                expenses_by_user_month,
                manuals_by_user_month,
            )
        else:
            ytd_base = 0.0

        # Финальные значения
        final_buy_cost      = calc['month_buy_cost']  if has_activity else manual_buy_cost
        final_sell_cost     = calc['month_sell_cost'] if has_activity else manual_sell_cost
        final_buy_comm      = calc['month_buy_comm']  if has_activity else manual_buy_comm
        final_sell_comm     = calc['month_sell_comm'] if has_activity else manual_sell_comm
        final_exch_comm_rub = (usdt['month_exch_comm_rub'] + ton['month_exch_comm_rub']) if has_activity else manual_exch_comm_rub
        final_bank_comm     = (calc['month_buy_comm'] + calc['month_sell_comm']) if has_activity else (manual_buy_comm + manual_sell_comm)

        sell_cost_all = final_sell_cost
        ndfl_all      = _calc_ndfl(gross_all, tax_type,
                                   sell_cost=sell_cost_all,
                                   ytd_base=ytd_base)
        after_all  = gross_all - ndfl_all
        share_all  = after_all * (share_percent / 100) if after_all > 0 else 0.0
        net_all    = after_all - share_all

        has_activity_final = (
            has_activity or (uid, share_key) in manuals_by_user_month
        )

        # Пропускаем месяцы без активности и без manual
        if not has_activity_final:
            continue

        months_data.append({
            'month_num':     mo_num,
            'month_name':    mo_name,
            'buy':           round(final_buy_cost, 2),
            'sell':          round(final_sell_cost, 2),
            'bank_comm':     round(final_bank_comm, 2),
            'exch_comm':     round(final_exch_comm_rub, 2),
            'expenses':      round(month_expenses, 2),
            'gross':         round(gross_all, 2),
            'gross_positive': gross_all >= 0,
            'tax_label':     'УСН' if tax_type in ('USN_INCOME', 'USN_INCOME_OUTCOME') else 'НДФЛ',
            'ndfl':          round(ndfl_all, 2),
            'share_percent': share_percent,
            'share':         round(share_all, 2),
            'net':           round(net_all, 2),
            'net_positive':  net_all >= 0,
            'usdt_balance':  round(manual_crypto_balance if manual_crypto_balance is not None else usdt['remainder_qty_display'], 4),
            'ton_balance':   round(ton['remainder_qty_display'], 4),
        })

        total_summary['sell']  += final_sell_cost
        total_summary['buy']   += final_buy_cost
        total_summary['gross'] += gross_all
        total_summary['share'] += share_all
        total_summary['net']   += net_all

    return JsonResponse({
        'username': u.username,
        'year':     year,
        'months':   months_data,
        'summary': {
            'sell_sum': round(total_summary['sell'],  2),
            'buy_sum':  round(total_summary['buy'],   2),
            'turnover': round(total_summary['sell'] + total_summary['buy'], 2),
            'gross':    round(total_summary['gross'], 2),
            'share':    round(total_summary['share'], 2),
            'net':      round(total_summary['net'],   2),
        },
    })

@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_orders_editor(request):
    """
    Редактор ордеров в админке.
    """
    order = None
    search_query = request.GET.get('order_id_search')
    all_banks = BankDetail.objects.filter(is_deleted=False).order_by('name')

    # Последний пользователь, для которого создавался ордер (из сессии)
    last_user_id = request.session.get('admin_last_order_user_id')
    last_user = None
    if last_user_id:
        last_user = User.objects.filter(id=last_user_id).first()

    # 1. Логика поиска ордера
    if search_query:
        search_query = search_query.strip()
        order = Order.objects.filter(external_id__iexact=search_query).first()
        if not order and search_query.isdigit():
            order = Order.objects.filter(id=int(search_query)).first()
        if not order:
            messages.error(request, f"Ордер '{search_query}' не найден.")

    # 2. Логика создания ордера администратором
    if request.method == 'POST' and 'create_order_admin' in request.POST:
        target_user_id = request.POST.get('target_user_id', '').strip()
        target_user = None
        if target_user_id:
            target_user = User.objects.filter(id=target_user_id).first()

        if not target_user:
            messages.error(request, "Пользователь не найден. Выберите пользователя из поиска.")
        else:
            external_id = request.POST.get('external_id', '').strip()
            if not external_id:
                messages.error(request, 'Номер ордера обязателен.')
            else:
                exchange_val = request.POST.get('exchange', '').strip()
                already_exists = Order.objects.filter(
                    user=target_user,
                    external_id=external_id,
                    exchange_type=exchange_val,
                ).exists()

                if already_exists:
                    messages.error(request, f'Ордер {external_id} ({exchange_val}) уже существует у пользователя {target_user.username}.')
                else:
                    raw_date = request.POST.get('created_at')
                    if raw_date:
                        naive_date = datetime.strptime(raw_date, '%Y-%m-%dT%H:%M')
                        order_date = timezone.make_aware(naive_date)
                    else:
                        order_date = timezone.now()

                    bank_id = request.POST.get('details')
                    bank_instance = BankDetail.objects.filter(id=bank_id, is_deleted=False).first() if bank_id else None

                    price_val      = safe_decimal(request.POST.get('price'))
                    amount_val     = safe_decimal(request.POST.get('amount'))
                    cost_val       = safe_decimal(request.POST.get('cost'))
                    commission_val = safe_decimal(request.POST.get('commission_value'))

                    exchange_name = exchange_val.lower()
                    user_comm_rate = 0.0
                    if 'bybit' in exchange_name:
                        user_comm_rate = target_user.bybit_commission
                    elif 'htx' in exchange_name or 'huobi' in exchange_name:
                        user_comm_rate = target_user.htx_commission
                    elif 'mexc' in exchange_name:
                        user_comm_rate = target_user.mexc_commission
                    elif 'bitget' in exchange_name:
                        user_comm_rate = target_user.bitget_commission
                    elif 'telegram' in exchange_name:
                        user_comm_rate = target_user.telegram_commission

                    Order.objects.create(
                        user=target_user,
                        external_id=external_id,
                        price=price_val,
                        amount=amount_val,
                        cost=cost_val,
                        commission=commission_val,
                        operation_type=request.POST.get('operation_type'),
                        exchange_type=exchange_val,
                        bank_detail=bank_instance,
                        commission_type=request.POST.get('commission_type'),
                        exchange_commission_rate=user_comm_rate,
                        created_at=order_date,
                        currency=request.POST.get('currency', 'USDT'),  # ← добавить эту строку
                    )

                    UnprocessedOrder.objects.filter(
                        user=target_user,
                        order_id=external_id,
                    ).delete()

                    # Запоминаем последнего пользователя
                    request.session['admin_last_order_user_id'] = target_user.id

                    messages.success(request, f'Ордер {external_id} успешно создан для пользователя {target_user.username}.')
                    return redirect('admin_orders_editor')

    # 3. Логика сохранения изменений
    if request.method == 'POST' and 'save_order' in request.POST:
        target_pk = request.POST.get('target_order_pk')
        try:
            current_order = Order.objects.get(pk=target_pk)

            current_order.external_id = request.POST.get('external_id')
            current_order.operation_type = request.POST.get('operation_type')

            bank_id = request.POST.get('bank_detail_id')
            if bank_id:
                current_order.bank_detail_id = bank_id

            def to_decimal(val):
                if not val: return Decimal('0')
                return Decimal(str(val).replace(',', '.'))

            current_order.price = to_decimal(request.POST.get('price'))
            current_order.amount = to_decimal(request.POST.get('amount'))
            current_order.cost = to_decimal(request.POST.get('cost'))
            current_order.commission = to_decimal(request.POST.get('commission'))
            current_order.commission_type = request.POST.get('commission_type')

            date_raw = request.POST.get('created_at')
            if date_raw:
                dt = parse_datetime(date_raw)
                if dt:
                    current_order.created_at = dt

            current_order.save()
            messages.success(request, f"Ордер #{current_order.id} успешно обновлен.")
            return redirect(f'/p2p-admin/orders/?order_id_search={current_order.external_id}')

        except Order.DoesNotExist:
            messages.error(request, "Ошибка: Ордер не найден в базе при сохранении.")
        except Exception as e:
            messages.error(request, f"Ошибка сохранения: {str(e)}")

    return render(request, 'custom_admin/orders_editor.html', {
        'order': order,
        'all_banks': all_banks,
        'search_query': search_query,
        'last_user': last_user,
        'current_time': timezone.now().strftime('%Y-%m-%dT%H:%M'),
    })



@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def export_nds(request):
    """Скачивание XML-декларации по НДС (сумма считается из ордеров)."""
    user_id = request.GET.get('user_id')
    year = int(request.GET.get('year') or datetime.now().year)
    quarter_key = request.GET.get('quarter', 'Q1')
    correction_number = int(request.GET.get('correction') or 0)

    if quarter_key not in QUARTERS:
        return JsonResponse({'error': 'Неверный квартал.'}, status=400)

    target = User.objects.filter(id=user_id).first()
    if not target:
        return JsonResponse({'error': 'Пользователь не найден.'}, status=404)

    try:
        filename, xml_bytes, info = build_nds_declaration(target, year, quarter_key, correction_number)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)

    response = HttpResponse(xml_bytes, content_type='application/xml; charset=windows-1251')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_fns_documents(request):
    """
    Вкладка ФНС в админке: список документов выбранного пользователя.
    Переключено на ОБЩУЮ функцию get_required_fns_documents - тот же
    источник правды, что и на личной странице пользователя /fns/
    (сумма налога, группировка по кварталам, срок сдачи, скрытие Q1 НДС -
    всё это теперь одинаково в обоих местах автоматически).
    """
    User = get_user_model()
    users = User.objects.all().order_by('username')

    selected_user_id = request.GET.get('user_id')
    year = int(request.GET.get('year') or timezone.now().year)

    selected_user = None
    documents = []

    if selected_user_id:
        selected_user = User.objects.filter(id=selected_user_id).first()

    if selected_user:
        documents = get_required_fns_documents(selected_user, year, for_user_id=selected_user.id)

    return render(request, 'custom_admin/fns_documents.html', {
        'users': users,
        'selected_user': selected_user,
        'selected_user_id': selected_user_id,
        'year': year,
        'documents': documents,
    })

@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_turnover_control(request):
    """
    Единая точка входа:
    1. Если action='search_users' -> ищет юзеров (JSON).
    2. Если action='get_turnover' -> считает оборот (JSON).
    3. Иначе -> отдает HTML страницу.
    """
    action = request.GET.get('action')

    # === ЛОГИКА 1: ПОИСК ПОЛЬЗОВАТЕЛЯ ===
    if action == 'search_users':
        q = request.GET.get('q', '').strip()
        if len(q) < 1:
            return JsonResponse([], safe=False)
        
        users = User.objects.filter(username__icontains=q)[:10]
        data = [{'id': u.id, 'username': u.username} for u in users]
        return JsonResponse(data, safe=False)

    # === ЛОГИКА 2: РАСЧЕТ ОБОРОТА ===
    if action == 'get_turnover':
        user_id = request.GET.get('target_user_id')
        start_date_str = request.GET.get('start')
        end_date_str = request.GET.get('end')

        if not user_id:
            return JsonResponse({'error': 'No user ID'}, status=400)

        orders = Order.objects.filter(user_id=user_id)

        if start_date_str and end_date_str:
            orders = orders.filter(created_at__range=[
                f"{start_date_str} 00:00:00", 
                f"{end_date_str} 23:59:59"
            ])

        buy_orders = orders.filter(operation_type='BUY').order_by('-created_at')
        sell_orders = orders.filter(operation_type='SELL').order_by('-created_at')

        # Считаем суммы (cost - это рубли)
        buy_sum = buy_orders.aggregate(Sum('cost'))['cost__sum'] or 0
        sell_sum = sell_orders.aggregate(Sum('cost'))['cost__sum'] or 0

        # Сериализация для таблицы
        def serialize(qs):
            res = []
            for o in qs:
                display_id = o.external_id if o.external_id else str(o.id)
                res.append({
                    'id': o.id,
                    'external_id': display_id,
                    'price': float(o.price),   # Курс
                    'amount': float(o.amount), # USDT (Объем)
                    'cost': float(o.cost),     # RUB (Стоимость)
                    'date': o.created_at.strftime('%d.%m.%Y %H:%M')
                })
            return res

        return JsonResponse({
            'buy_sum': buy_sum,
            'buy_count': buy_orders.count(),
            'sell_sum': sell_sum,
            'sell_count': sell_orders.count(),
            'total_sum': buy_sum + sell_sum,
            'profit': "{:.2f}".format(sell_sum - buy_sum),
            'buy_orders': serialize(buy_orders),
            'sell_orders': serialize(sell_orders),
        })

    # === ЛОГИКА 3: ПРОСТО ОТДАТЬ СТРАНИЦУ ===
    return render(request, 'custom_admin/turnover_control.html')





@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_statistics_24h(request):
    f_user = request.GET.get('user', '')
    f_exchange = request.GET.get('exchange', '') # Bybit, MEXC, или пусто
    f_type = request.GET.get('type', '')
    f_limit = request.GET.get('limit', '50')

    is_update_action = bool(request.GET)
    orders = []

    if is_update_action:
        # --- 1. ОПРЕДЕЛЯЕМ ПОЛЬЗОВАТЕЛЕЙ ---
        if f_user and f_user.isdigit():
            # Если выбран конкретный пользователь
            users_qs = User.objects.filter(id=f_user)
        else:
            # Если "Все", берем тех, у кого есть хоть какие-то ключи
            users_qs = User.objects.filter(
                Q(bybit_api_key__isnull=False) & ~Q(bybit_api_key='') |
                Q(mexc_api_key__isnull=False) & ~Q(mexc_api_key='')
            )

        filters = {'type': f_type, 'exchange': f_exchange}
        
        # --- 2. СБОР ДАННЫХ ---
        
        # BYBIT
        # Запускаем, если выбран "Bybit" или "Все источники"
        if f_exchange == 'Bybit' or f_exchange == '' or f_exchange == '1':
            # Фильтруем юзеров, у кого есть ключи Bybit, чтобы не гонять пустышки
            bybit_users = users_qs.exclude(bybit_api_key__isnull=True).exclude(bybit_api_key='')
            bybit_data = get_bybit_orders(bybit_users, filters)
            orders.extend(bybit_data)

        # MEXC
        # Запускаем, если выбран "MEXC" или "Все источники"
        if f_exchange == 'MEXC' or f_exchange == '':
            # Фильтруем юзеров, у кого есть ключи MEXC
            mexc_users = users_qs.exclude(mexc_api_key__isnull=True).exclude(mexc_api_key='')
            mexc_data = get_mexc_orders_parallel(mexc_users, filters)
            orders.extend(mexc_data)

        # --- 3. ФИНАЛЬНАЯ СОРТИРОВКА ---
        # Сортируем общий список по дате (свежие сверху)
        orders.sort(key=lambda x: x.created_at, reverse=True)

    # Статистика "на лету"
    total_count = len(orders)
    buy_count = sum(1 for x in orders if x.operation_type == 'BUY')
    sell_count = sum(1 for x in orders if x.operation_type == 'SELL')
    total_sum = sum(x.cost for x in orders)

    # Пагинация
    try:
        limit = int(f_limit)
    except:
        limit = 50
        
    paginator = Paginator(orders, limit)
    page_obj = paginator.get_page(request.GET.get('page'))

    all_users = User.objects.all().order_by('username')

    context = {
        'orders': page_obj,
        'users': all_users,
        'current_user': int(f_user) if f_user.isdigit() else '',
        'current_exchange': f_exchange,
        'current_type': f_type,
        'current_limit': f_limit,
        'total_orders': total_count,
        'buy_count': buy_count,
        'sell_count': sell_count,
        'total_amount': total_sum,
        'is_search': is_update_action
    }

    return render(request, 'custom_admin/statistics_24.html', context)


# === API (только для админов) ===
@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def api_search_users(request):
    query = request.GET.get('q', '').strip()
    if len(query) < 1: return JsonResponse([], safe=False)
    
    users = User.objects.filter(Q(username__icontains=query) | Q(id__icontains=query))[:10]
    return JsonResponse([{'id': u.id, 'username': u.username} for u in users], safe=False)


@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def api_get_turnover(request, user_id):
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    orders = Order.objects.filter(user_id=user_id, created_at__date__range=[start_date, end_date]).order_by('-created_at')
    
    buy_orders = orders.filter(operation_type='BUY')
    sell_orders = orders.filter(operation_type='SELL')
    
    buy_sum = buy_orders.aggregate(Sum('amount'))['amount__sum'] or 0
    sell_sum = sell_orders.aggregate(Sum('amount'))['amount__sum'] or 0
    
    def fmt(val): return f"{val:,.2f}".replace(',', ' ').replace('.', ',')
    def serialize(qs): 
        return [{'id': o.pk, 'amount': float(o.amount), 'cost': float(o.cost) if hasattr(o,'cost') else 0, 'date': o.created_at.strftime('%d.%m %H:%M')} for o in qs]

    data = {
        'buy_sum': fmt(buy_sum), 'buy_count': buy_orders.count(),
        'sell_sum': fmt(sell_sum), 'sell_count': sell_orders.count(),
        'total_sum': fmt(buy_sum + sell_sum), 'profit': fmt(sell_sum - buy_sum),
        'buy_orders': serialize(buy_orders), 'sell_orders': serialize(sell_orders),
    }
    return JsonResponse(data)



@user_passes_test(lambda u: u.is_superuser)
def export_screenshots_view(request):
    # 1. Получаем параметры
    user_id = request.GET.get('user_id')
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    bank_id = request.GET.get('bank')
    op_type = request.GET.get('type')

    if not user_id:
        return HttpResponse("User ID required", status=400)

    # 2. Базовая фильтрация
    orders = Order.objects.filter(user_id=user_id).exclude(screenshot='').exclude(screenshot__isnull=True)

    # 3. Фильтры
    if start_str:
        start_date = parse_date(start_str)
        if start_date:
            start_dt = timezone.make_aware(datetime.combine(start_date, time.min))
            orders = orders.filter(created_at__gte=start_dt)
    
    if end_str:
        end_date = parse_date(end_str)
        if end_date:
            end_dt = timezone.make_aware(datetime.combine(end_date, time.max))
            orders = orders.filter(created_at__lte=end_dt)

    if op_type and op_type in ['BUY', 'SELL']:
        orders = orders.filter(operation_type=op_type)

    if bank_id and bank_id.isdigit():
        orders = orders.filter(bank_detail_id=int(bank_id))

    orders_list = list(orders)
    if not orders_list:
        return HttpResponse("Нет скриншотов по выбранным критериям", content_type="text/plain; charset=utf-8")

    # 4. Скачивание файлов из S3 — ПАРАЛЛЕЛЬНО (тот же приём, что и в
    # bybit_service.get_orders_parallel — ThreadPoolExecutor, 10 потоков).
    # Раньше файлы читались строго по одному, друг за другом — при большой
    # выборке это N последовательных сетевых запросов к Yandex S3, каждый
    # держит воркер gunicorn занятым, а админ не видит вообще ничего, пока
    # не соберётся весь архив целиком.
    def _fetch_screenshot(order):
        if not (order.screenshot and order.screenshot.storage.exists(order.screenshot.name)):
            return None
        original_name = os.path.basename(order.screenshot.name)
        ext = os.path.splitext(original_name)[1] or '.png'
        label = order.external_id if order.external_id else f"manual_{order.id}"
        filename = f"{label}{ext}"
        with order.screenshot.open('rb') as f:
            data = f.read()
        return (filename, data)

    fetched = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        future_to_order = {executor.submit(_fetch_screenshot, o): o for o in orders_list}
        for future in concurrent.futures.as_completed(future_to_order):
            order = future_to_order[future]
            try:
                result = future.result()
                if result:
                    fetched[order.id] = result
            except Exception as e:
                print(f"Ошибка с файлом ордера {order.id}: {e}")
                continue

    # 5. Сборка архива — уже из скачанных байт, чисто в памяти, быстро.
    # Порядок сохраняем как в исходной выборке (fetched заполнялся
    # параллельно, вразнобой).
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for o in orders_list:
            entry = fetched.get(o.id)
            if entry:
                filename, data = entry
                zip_file.writestr(filename, data)

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    
    # === ИСПРАВЛЕНИЕ ОШИБКИ DATETIME ===
    # Используем timezone.now() вместо datetime.now()
    filename_zip = f"screens_{user_id}_{timezone.now().strftime('%Y%m%d')}.zip"
    
    response['Content-Disposition'] = f'attachment; filename="{filename_zip}"'
    
    return response

@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_manual_entry_save(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    import json, re
    try:
        body = json.loads(request.body)
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    user_id = body.get('user_id')
    month   = body.get('month')

    if not user_id or not month:
        return JsonResponse({'error': 'user_id и month обязательны'}, status=400)

    if not re.match(r'^\d{4}-\d{2}$', str(month)):
        return JsonResponse({'error': 'Формат месяца: YYYY-MM'}, status=400)

    try:
        entry, created = MonthlyManualEntry.objects.update_or_create(
            user_id=user_id,
            month=month,
            defaults={
                'gross':           body.get('gross', 0),
                'buy_qty':         body.get('buy_qty', 0),
                'buy_cost':        body.get('buy_cost', 0),
                'buy_comm':        body.get('buy_comm', 0),
                'sell_qty':        body.get('sell_qty', 0),
                'sell_cost':       body.get('sell_cost', 0),
                'sell_comm':       body.get('sell_comm', 0),
                'exch_comm_rub':   body.get('exch_comm_rub', 0),
                'remainder_qty':   body.get('remainder_qty', 0),
                'remainder_price': body.get('remainder_price', 0),
                'remainder_cost':  body.get('remainder_cost', 0),
                'source':          body.get('source', 'manual'),
            }
        )
        return JsonResponse({
            'ok':      True,
            'created': created,
            'id':      entry.id,
            'month':   entry.month,
            'gross':   float(entry.gross),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_manual_entry_list(request, user_id):
    """
    Возвращает все ручные записи прибыли для пользователя.
    GET /p2p-admin/manual-entries/<user_id>/
    """
    entries = MonthlyManualEntry.objects.filter(user_id=user_id).order_by('-month')

    months_ru = {
        '01': 'Январь',  '02': 'Февраль', '03': 'Март',
        '04': 'Апрель',  '05': 'Май',      '06': 'Июнь',
        '07': 'Июль',    '08': 'Август',   '09': 'Сентябрь',
        '10': 'Октябрь', '11': 'Ноябрь',  '12': 'Декабрь',
    }

    data = []
    for e in entries:
        month_num  = e.month.split('-')[1]
        month_name = months_ru.get(month_num, '')
        year       = e.month.split('-')[0]
        data.append({
            'id':              e.id,
            'month':           e.month,
            'month_name':      f"{month_name} {year}",
            'source':          getattr(e, 'source', 'manual'),
            # Покупки
            'buy_qty':         float(getattr(e, 'buy_qty',  0) or 0),
            'buy_cost':        float(getattr(e, 'buy_cost', 0) or 0),
            'buy_comm':        float(getattr(e, 'buy_comm', 0) or 0),
            # Продажи
            'sell_qty':        float(getattr(e, 'sell_qty',  0) or 0),
            'sell_cost':       float(getattr(e, 'sell_cost', 0) or 0),
            'sell_comm':       float(getattr(e, 'sell_comm', 0) or 0),
            # Комиссия биржи
            'exch_comm_rub':   float(getattr(e, 'exch_comm_rub', 0) or 0),
            # Остаток
            'remainder_qty':   float(getattr(e, 'remainder_qty',   0) or 0),
            'remainder_price': float(getattr(e, 'remainder_price', 0) or 0),
            'remainder_cost':  float(getattr(e, 'remainder_cost',  0) or 0),
            # Прибыль
            'gross':           float(e.gross),
        })

    return JsonResponse({'entries': data})

@login_required(login_url='admin_login')
@user_passes_test(lambda u: u.is_superuser, login_url='admin_login')
def admin_manual_entry_upload_excel(request):
    """
    Массовая загрузка Excel файлов за период.
    Имя файла = username пользователя (simonenko.xlsx).
    Читает лист 'Табл.№1', вытаскивает итоговые строки и строки ордеров.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST only'}, status=405)

    import openpyxl
    from django.contrib.auth import get_user_model
    User = get_user_model()

    month = request.POST.get('month', '').strip()  # '2026-01'
    files = request.FILES.getlist('excel_files')
    # username'ы, для которых админ явно подтвердил перезапись поверх расхождения
    force_usernames = {
        u.strip().lower() for u in request.POST.get('force_usernames', '').split(',') if u.strip()
    }

    if not month or not files:
        return JsonResponse({'error': 'month и файлы обязательны'}, status=400)

    results = []

    MISMATCH_FIELDS = [
        ('buy_qty',         'Кол-во покупок'),
        ('buy_cost',        'Стоимость покупок'),
        ('buy_comm',        'Комиссия банка (покупки)'),
        ('sell_qty',        'Кол-во продаж'),
        ('sell_cost',       'Стоимость продаж'),
        ('sell_comm',       'Комиссия банка (продажи)'),
        ('exch_comm_rub',   'Комиссия биржи, ₽'),
        ('remainder_qty',   'Остаток, кол-во'),
        ('remainder_price', 'Остаток, курс'),
        ('remainder_cost',  'Остаток, стоимость'),
        ('gross',           'Прибыль'),
    ]
    MISMATCH_TOLERANCE = 1.0  # ₽ / единицы — меньше считается округлением, не расхождением

    for f in files:
        username = f.name.rsplit('.', 1)[0].lower()  # 'simonenko.xlsx' -> 'simonenko'
        result   = {'file': f.name, 'username': username}

        # Ищем пользователя
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            result['status'] = 'error'
            result['message'] = f'Пользователь "{username}" не найден'
            results.append(result)
            continue

        # Читаем Excel
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

            # Ищем лист Табл.№1
            sheet = None
            for name in wb.sheetnames:
                if 'табл' in name.lower() and '1' in name:
                    sheet = wb[name]
                    break

            if sheet is None:
                result['status'] = 'error'
                result['message'] = 'Лист "Табл.№1" не найден'
                results.append(result)
                continue

            # Читаем все строки
            rows = list(sheet.iter_rows(values_only=True))

            buy_qty = 0.0; buy_cost = 0.0; buy_comm = 0.0
            sell_qty = 0.0; sell_cost = 0.0; sell_comm = 0.0
            exch_comm_rub = 0.0
            remainder_qty = 0.0; remainder_price = 0.0; remainder_cost = 0.0
            gross = 0.0

            for row in rows:
                if not row or len(row) < 4:
                    continue

                label = str(row[0] or '').lower().strip()

                # Итоговые строки
                if 'торговый результат' in label:
                    buy_qty   = float(row[4] or 0)
                    buy_cost  = float(row[6] or 0)
                    buy_comm  = float(row[7] or 0)
                    sell_qty  = float(row[8] or 0)
                    sell_cost = float(row[10] or 0)
                    sell_comm = float(row[11] or 0)
                    continue

                if 'остаток' in label and 'нереализован' in label:
                    remainder_qty   = float(row[4] or 0)
                    remainder_price = float(row[5] or 0)
                    remainder_cost  = float(row[6] or 0)
                    continue

                if 'прибыль' in label and 'реализован' not in label:
                    gross = float(row[4] or 0)
                    continue

                # Строки ордеров — считаем комиссию биржи в рублях (M * J)
                # row[0] = №, row[8] = sell_qty, row[9] = sell_price, row[12] = exch_comm_usdt
                try:
                    num = row[0]
                    if num and str(num).strip().isdigit():
                        sell_price_row    = float(row[9]  or 0)
                        exch_comm_usdt_row = float(row[12] or 0)
                        if exch_comm_usdt_row > 0 and sell_price_row > 0:
                            exch_comm_rub += exch_comm_usdt_row * sell_price_row
                except Exception:
                    pass

            new_values = {
                'buy_qty':         buy_qty,
                'buy_cost':        buy_cost,
                'buy_comm':        buy_comm,
                'sell_qty':        sell_qty,
                'sell_cost':       sell_cost,
                'sell_comm':       sell_comm,
                'exch_comm_rub':   round(exch_comm_rub, 2),
                'remainder_qty':   remainder_qty,
                'remainder_price': remainder_price,
                'remainder_cost':  remainder_cost,
                'gross':           gross,
            }

            # Сверяем с уже сохранёнными данными (могли быть вручную поправлены) —
            # если расходятся и это не подтверждено явно, блокируем перезапись.
            existing = MonthlyManualEntry.objects.filter(user=user, month=month).first()
            diffs = []
            if existing:
                for field, field_label in MISMATCH_FIELDS:
                    old_v = float(getattr(existing, field, 0) or 0)
                    new_v = float(new_values[field])
                    if abs(old_v - new_v) > MISMATCH_TOLERANCE:
                        diffs.append({
                            'field': field, 'label': field_label,
                            'old': old_v, 'new': new_v,
                        })

            if diffs and username not in force_usernames:
                result['status']  = 'mismatch'
                result['message'] = f'Расхождение с уже сохранёнными данными по {len(diffs)} полям — сохранение заблокировано'
                result['diffs']   = diffs
                result['data'] = {
                    'buy_cost':  buy_cost,
                    'sell_cost': sell_cost,
                    'gross':     gross,
                    'remainder': f'{remainder_qty} USDT × {remainder_price} ₽',
                }
                results.append(result)
                continue

            # Сохраняем
            entry, created = MonthlyManualEntry.objects.update_or_create(
                user=user,
                month=month,
                defaults={
                    **new_values,
                    'source': 'excel',
                }
            )

            result['status']  = 'ok'
            result['message'] = f'{"Создано" if created else "Обновлено"} — gross: {gross:,.2f} ₽'
            result['data'] = {
                'buy_cost':  buy_cost,
                'sell_cost': sell_cost,
                'gross':     gross,
                'remainder': f'{remainder_qty} USDT × {remainder_price} ₽',
            }

        except Exception as e:
            result['status']  = 'error'
            result['message'] = str(e)

        results.append(result)

    return JsonResponse({'results': results})
