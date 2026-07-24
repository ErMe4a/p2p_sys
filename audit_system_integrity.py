# audit_system_integrity.py
# READ-ONLY (кроме опционального Блока 10/13 — см. предупреждение там).
# Полный аудит целостности данных и рисков в отчётности на основе AUDIT_PLAN.md.
# В отличие от audit_orders_2026.py (аномалии курса/суммы/ручного ввода за год),
# этот скрипт проверяет СТРУКТУРНЫЕ риски: дубли, неверифицированные данные
# в отчётах, гонки при фискализации, расхождение "живых" и "сохранённых" данных,
# смещение часового пояса, конфликты ручных записей.
#
# Запуск:  python manage.py shell < audit_system_integrity.py
# Результат: /tmp/audit_system_integrity.xlsx

import time as _time
from collections import defaultdict
from datetime import datetime, timezone as dt_timezone
from decimal import Decimal
from zoneinfo import ZoneInfo

from django.conf import settings
from django.contrib.auth import get_user_model

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from orders.models import Order, MonthlyManualEntry, DocumentSubmission

User = get_user_model()
MSK = ZoneInfo("Europe/Moscow")
SYSTEM_START = datetime(2026, 2, 1, 0, 0, 0, tzinfo=dt_timezone.utc)  # как в views.py
OUTPUT_PATH = "/tmp/audit_system_integrity.xlsx"

# ── Опциональные проверки, требующие обращения к API бирж ──────────────────
# По умолчанию ВЫКЛЮЧЕНЫ: они делают реальные сетевые запросы под ключами
# пользователей (лимиты запросов бирж) и включают в отчёт данные, которых
# нет в БД. Блок 13 (exchange_api.get_order_from_exchange) — ЕДИНСТВЕННАЯ
# часть скрипта, которая не гарантированно read-only: при ошибке ключа
# биржи она, как и в проде, помечает ключ пользователя невалидным в БД
# (см. exchange_api._mark_bybit_invalid). Включайте осознанно и не часто.
RUN_LIVE_RECONCILIATION = False   # Блок 10: live 24ч (bybit_service/mexc_service) vs БД
RUN_LIVE_SPOT_CHECK = False       # Блок 13: точечная сверка price/cost/amount с биржей
LIVE_SPOT_CHECK_SAMPLE = 5        # сколько ордеров сверять в Блоке 13


def f(x):
    try:
        return float(x or 0)
    except Exception:
        return 0.0


def fmt_dt(o):
    return o.created_at.astimezone(MSK).strftime("%d.%m.%Y %H:%M")


print("Загружаю ордера из БД...")
orders = list(Order.objects.all().order_by("user_id", "created_at"))
uid_to_name = {u.id: u.username for u in User.objects.all()}
print(f"Загружено ордеров: {len(orders)} (без ограничения по году — проверяем весь массив)")


def base_row(o, note=""):
    return [
        o.id, o.external_id or "", uid_to_name.get(o.user_id, o.user_id),
        o.operation_type, o.currency, round(f(o.amount), 4), round(f(o.price), 4),
        round(f(o.cost), 2), o.exchange_type, fmt_dt(o),
        o.is_verified, o.is_manual, note,
    ]


BASE_HEADER = ["ID системы", "ID на бирже", "Владелец", "Тип", "Валюта",
               "Кол-во", "Курс", "Сумма (руб)", "Биржа", "Дата",
               "Верифицирован", "Ручной ввод", "Диагноз"]


# ============================================================
# 1. ДУБЛИ ордеров: (user, external_id, exchange_type)
#    Order.Meta не имеет unique_together (в отличие от UnprocessedOrder/
#    IgnoredOrder) — дубль удваивает сумму во всех отчётах и может
#    привести к повторной фискализации.
# ============================================================
by_key = defaultdict(list)
for o in orders:
    ext = o.external_id or ""
    if ext:
        by_key[(o.user_id, ext, o.exchange_type)].append(o)

duplicates = []
for key, group in by_key.items():
    if len(group) > 1:
        for o in group:
            duplicates.append((o, f"Дубль: {len(group)} ордеров с одинаковым ID биржи {key[1]!r} ({key[2]})"))


# ============================================================
# 2. НЕВЕРИФИЦИРОВАННЫЕ/ПРОВАЛЕННЫЕ ордера с ненулевой суммой,
#    которые ничем не отфильтрованы в export_excel_report /
#    admin_profit_view / nds_generator / uvedomlenie_generator —
#    они попадают в цифры отчётности и уведомлений как есть.
# ============================================================
unverified_in_reports = []
for o in orders:
    if o.is_manual:
        continue  # осознанный ручной ввод — не "непроверенные" данные
    receipt = o.receipt if isinstance(o.receipt, dict) else {}
    is_failed = bool(receipt.get("verification_failed"))
    if (not o.is_verified or is_failed) and f(o.cost) != 0:
        reason = "verification_failed" if is_failed else "is_verified=False"
        unverified_in_reports.append(
            (o, f"Непроверенные данные ({reason}) участвуют в отчёте с суммой {f(o.cost):.2f} руб")
        )


# ============================================================
# 3. НУЛЕВЫЕ / ОТРИЦАТЕЛЬНЫЕ суммы — evotor_atol.py при amount<=0 или
#    cost<=0 подставляет 1.0 вместо отказа, т.е. это не просто "плохая
#    строка в отчёте", а риск фиктивной позиции в реальном чеке.
# ============================================================
bad_values = []
for o in orders:
    price, amount, cost = f(o.price), f(o.amount), f(o.cost)
    if price <= 0 or amount <= 0 or cost <= 0:
        bad_values.append((o, f"price={price} amount={amount} cost={cost} — есть поле <= 0"))


# ============================================================
# 4. РАСХОЖДЕНИЕ price × amount vs cost (>2%) — три поля должны быть
#    консистентны; большое расхождение обычно значит, что поля пришли
#    из разных источников/моментов времени (например, частичная
#    ручная правка).
# ============================================================
PRICE_AMOUNT_COST_TOLERANCE = 0.02
cost_mismatch = []
for o in orders:
    if o.price is None or o.amount is None or o.cost is None:
        continue
    price, amount, cost = Decimal(o.price), Decimal(o.amount), Decimal(o.cost)
    if price <= 0 or amount <= 0 or cost <= 0:
        continue
    expected = price * amount
    diff = abs(expected - cost)
    if cost and (diff / cost) > Decimal(str(PRICE_AMOUNT_COST_TOLERANCE)):
        dev_pct = float(diff / cost * 100)
        cost_mismatch.append(
            (o, f"Курс×Кол-во = {float(expected):.2f}, а Сумма = {float(cost):.2f} (расхождение {dev_pct:.0f}%)")
        )


# ============================================================
# 5. ВАЛЮТА вне допустимого списка (USDT/TON) — такие ордера молча
#    выпадают из _calc_currency/_calc_currency_from_orders, но
#    продолжают существовать в БД и в отчётах без фильтра по валюте
#    (например api_get_turnover).
# ============================================================
bad_currency = []
for o in orders:
    if o.currency not in ("USDT", "TON"):
        bad_currency.append((o, f"Валюта {o.currency!r} не входит в USDT/TON — не попадёт в помесячный расчёт прибыли"))


# ============================================================
# 6. ПРАВКИ ПОСЛЕ ФИСКАЛИЗАЦИИ ЧЕКА — receipt_service.py замораживает
#    price/sum/amount в order.receipt в момент отправки чека. Order не
#    имеет поля updated_at, поэтому сравнение с этим "снимком" — ЕДИНСТВЕННЫЙ
#    способ обнаружить постфактум-правку из самой БД. Расхождение значит:
#    цифра в уже отправленном фискальном чеке НЕ совпадает с тем, что
#    сейчас покажет отчёт по этому ордеру.
# ============================================================
post_fiscal_edits = []
for o in orders:
    receipt = o.receipt if isinstance(o.receipt, dict) else {}
    uuid_ = receipt.get("uuid")
    if not uuid_:
        continue
    try:
        r_price = float(receipt.get("price") or 0)
        r_sum = float(receipt.get("sum") or 0)
        r_amount = float(receipt.get("amount") or 0)
    except (TypeError, ValueError):
        continue
    if (abs(r_price - f(o.price)) > 0.01
            or abs(r_sum - f(o.cost)) > 0.01
            or abs(r_amount - f(o.amount)) > 0.00000001):
        note = (f"Чек {uuid_}: было price={r_price} sum={r_sum} amount={r_amount}; "
                f"сейчас price={f(o.price)} cost={f(o.cost)} amount={f(o.amount)}")
        post_fiscal_edits.append((o, note))


# ============================================================
# 7. НАРУШЕНИЕ ПРИНЦИПА "не фискализировать неверифицированные данные" —
#    прямая проверка: чек отправлен (есть uuid), а ордер при этом не
#    is_verified и не is_manual. В норме таких быть не должно вообще.
# ============================================================
invariant_violations = []
for o in orders:
    receipt = o.receipt if isinstance(o.receipt, dict) else {}
    if receipt.get("uuid") and not o.is_verified and not o.is_manual:
        invariant_violations.append(
            (o, f"Чек {receipt.get('uuid')} отправлен без верификации и без пометки is_manual")
        )


# ============================================================
# 8. ЧАСОВОЙ ПОЯС СЕРВЕРА — bybit_api.py/mexc_api.py/bybit_service.py/
#    mexc_service.py конвертируют ms-таймстемп биржи через
#    datetime.fromtimestamp(ms/1000) (наивное, в ЛОКАЛЬНОЙ TZ ОС) + make_aware()
#    (просто подписывает эту наивную дату как TIME_ZONE='Europe/Moscow',
#    БЕЗ пересчёта). Если TZ ОС сервера — не Europe/Moscow, все даты
#    created_at систематически смещены → ордера у границы месяца попадают
#    не в тот отчётный период (не тот квартал уведомления/НДС).
# ============================================================
os_utcoffset = datetime.now().astimezone().utcoffset()
os_offset_hours = os_utcoffset.total_seconds() / 3600 if os_utcoffset else None
os_tzname = _time.tzname
tz_mismatch = os_offset_hours is None or abs(os_offset_hours - 3.0) > 0.01

tz_note = (
    f"OS UTC-смещение = {os_offset_hours:+.1f}ч (tzname={os_tzname}); "
    f"DJANGO TIME_ZONE={settings.TIME_ZONE} USE_TZ={settings.USE_TZ}. "
    + ("НЕСОВПАДЕНИЕ: сервер не в MSK(+3) — created_at синхронизированных ордеров, "
       "вероятно, смещены." if tz_mismatch else "Совпадает с MSK(+3) — риска смещения по этой причине нет.")
)
print("\n[Блок 8] " + tz_note)


# ============================================================
# 9. ОРДЕРА У ГРАНИЦЫ МЕСЯЦА (+/- ~4ч от 1-го числа) — кандидаты на
#    попадание не в тот отчётный период, если Блок 8 показал несовпадение.
#    Справочный список для ручной сверки со временем на самой бирже.
# ============================================================
month_boundary_orders = []
for o in orders:
    d = o.created_at.astimezone(MSK)
    if (d.day == 1 and d.hour < 4) or (d.day >= 28 and d.hour >= 20):
        month_boundary_orders.append((o, f"Дата у границы месяца ({d.strftime('%d.%m %H:%M')} МСК) — риск смещения периода при TZ-баге"))


# ============================================================
# 10. LIVE (админ. "24ч") vs БД — bybit_service/mexc_service читают
#     биржу напрямую и ничего не сохраняют; отчёты читают Order,
#     заполненный периодической синхронизацией (bybit_api/mexc_api).
#     Это два независимых парсера — расхождение не гарантированно нулевое.
#     ОПЦИОНАЛЬНО (см. RUN_LIVE_RECONCILIATION) — делает реальные запросы
#     к Bybit/MEXC под ключами пользователей.
# ============================================================
live_reconciliation_rows = []
if RUN_LIVE_RECONCILIATION:
    from django.utils import timezone as dj_timezone
    from datetime import timedelta
    from orders.bybit_service import get_orders_parallel
    from orders.mexc_service import get_mexc_orders_parallel

    since = dj_timezone.now() - timedelta(hours=24)

    bybit_users = User.objects.exclude(bybit_api_key__isnull=True).exclude(bybit_api_key="")
    live_bybit = get_orders_parallel(bybit_users)
    db_bybit = [o for o in orders if o.exchange_type == "Bybit" and o.created_at >= since]
    live_reconciliation_rows.append((
        "Bybit", len(live_bybit), sum(f(x.cost) for x in live_bybit),
        len(db_bybit), sum(f(x.cost) for x in db_bybit),
    ))

    mexc_users = User.objects.exclude(mexc_api_key__isnull=True).exclude(mexc_api_key="")
    live_mexc = get_mexc_orders_parallel(mexc_users)
    db_mexc = [o for o in orders if o.exchange_type == "MEXC" and o.created_at >= since]
    live_reconciliation_rows.append((
        "MEXC", len(live_mexc), sum(f(x.cost) for x in live_mexc),
        len(db_mexc), sum(f(x.cost) for x in db_mexc),
    ))

    for exch, lc, ls, dc, ds in live_reconciliation_rows:
        print(f"[Блок 10] {exch}: live count={lc} sum={ls:.2f} | БД count={dc} sum={ds:.2f} | "
              f"diff count={lc - dc} diff sum={ls - ds:.2f}")
else:
    print("\n[Блок 10] Пропущен (RUN_LIVE_RECONCILIATION=False) — не делает сетевых запросов к биржам.")


# ============================================================
# 11. MonthlyManualEntry vs реальные ордера за тот же месяц —
#     _calc_month_profit_from_orders использует ручную запись, только
#     если по ордерам активности НЕТ. Если ордера появились ПОСЛЕ того,
#     как ручная запись уже была создана (например, досинхронизировались
#     позже), система молча переключится на расчёт по ордерам — а ручная
#     цифра могла уже уйти в поданное уведомление/НДС.
# ============================================================
order_months_by_user = defaultdict(set)
for o in orders:
    d = o.created_at.astimezone(MSK)
    order_months_by_user[o.user_id].add((d.year, d.month))

manual_conflicts = []
for m in MonthlyManualEntry.objects.all():
    try:
        year, month = (int(x) for x in m.month.split("-"))
    except ValueError:
        continue
    if (year, month) in order_months_by_user.get(m.user_id, set()):
        manual_conflicts.append((
            uid_to_name.get(m.user_id, m.user_id), m.month, m.source, float(m.gross),
            "У пользователя есть реальные ордера за этот месяц — ручная запись, скорее всего, больше не используется в расчёте",
        ))


# ============================================================
# 12. DocumentSubmission — toggle_document_submission захардкожен на
#     username == 'maks' (orders/views.py). Для всех остальных
#     пользователей отметка "Отправлено" не сохраняется. Проверяем,
#     что фактически ни у кого, кроме maks, нет отметок — это ожидаемый
#     симптом бага, а не "пользователи забыли отметить".
# ============================================================
doc_submission_rows = []
for u in User.objects.exclude(username="maks").order_by("username"):
    cnt = DocumentSubmission.objects.filter(user=u, submitted=True).count()
    doc_submission_rows.append((u.username, cnt))
nonzero_doc_submissions = [r for r in doc_submission_rows if r[1] > 0]


# ============================================================
# 13. (ОПЦИОНАЛЬНО, НЕ read-only) Точечная сверка price/cost/amount с
#     биржей через exchange_api.get_order_from_exchange для небольшой
#     выборки верифицированных ордеров Bybit/MEXC. ВНИМАНИЕ: при ошибке
#     ключа эта функция, как и в проде, помечает ключ пользователя
#     невалидным в БД — это единственная часть скрипта с побочным эффектом.
# ============================================================
spot_check_rows = []
if RUN_LIVE_SPOT_CHECK:
    from orders.exchange_api import get_order_from_exchange

    sample = [o for o in orders if o.exchange_type in ("Bybit", "MEXC") and o.is_verified][-LIVE_SPOT_CHECK_SAMPLE:]
    for o in sample:
        fresh = get_order_from_exchange(user=o.user, order_id=o.external_id, exchange_name=o.exchange_type)
        spot_check_rows.append((
            o.id, o.external_id, o.exchange_type,
            f(o.price), f(o.amount), f(o.cost),
            fresh.get("price") if fresh else None,
            fresh.get("amount") if fresh else None,
            fresh.get("cost") if fresh else None,
        ))
        print(f"[Блок 13] order {o.id}: БД price={f(o.price)} amount={f(o.amount)} cost={f(o.cost)} | "
              f"биржа сейчас = {fresh}")
else:
    print("[Блок 13] Пропущен (RUN_LIVE_SPOT_CHECK=False) — не делает сетевых запросов к биржам.")


# ============================================================
# ЗАПИСЬ В EXCEL
# ============================================================
print("\nФормирую Excel-отчёт...")
wb = Workbook()
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CRIT_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
CRIT_FONT = Font(bold=True, color="FFFFFF")
RED = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
YELLOW = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
GREY = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")

RISK_FILL = {"Критично": CRIT_FILL, "Высокий": RED, "Средний": YELLOW, "Низкий": GREEN, "—": GREY}
RISK_FONT = {"Критично": CRIT_FONT}


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, 10), 60)


def write_order_sheet(name, rows_with_notes):
    ws = wb.create_sheet(name)
    ws.append(BASE_HEADER)
    style_header(ws)
    for o, note in rows_with_notes:
        ws.append(base_row(o, note))
    autosize(ws)
    return ws


# ── Лист 0: ИТОГ ──
ws0 = wb.active
ws0.title = "ИТОГ"
ws0.append(["#", "Категория", "Кол-во", "Риск", "Что делать"])
style_header(ws0)

tz_risk = "Критично" if tz_mismatch else "Низкий"
summary = [
    ["1", "Дубли ордеров (user + ID биржи + биржа)", len(duplicates), "Критично",
     "Нет unique-constraint в Order.Meta — удвоение суммы в отчётах и риск повторной фискализации. См. AUDIT_PLAN.md п.1"],
    ["2", "Неверифицированные/проваленные ордера в отчётах", len(unverified_in_reports), "Критично",
     "Ни один отчёт (Excel/НДС/уведомление) не фильтрует is_verified — данные из расширения без проверки биржей идут в цифры. См. AUDIT_PLAN.md п.2"],
    ["3", "Нулевые/отрицательные price/amount/cost", len(bad_values), "Высокий",
     "evotor_atol.py подставит 1.0 вместо отказа — риск фиктивной позиции в чеке. См. AUDIT_PLAN.md п.3"],
    ["4", "Курс×Кол-во не сходится с Суммой (>2%)", len(cost_mismatch), "Высокий",
     "Поля заполнены из разных источников/моментов — проверить вручную. См. AUDIT_PLAN.md п.4"],
    ["5", "Валюта вне USDT/TON", len(bad_currency), "Средний",
     "Выпадает из помесячного расчёта прибыли молча. См. AUDIT_PLAN.md п.5"],
    ["6", "Правки ордера ПОСЛЕ отправки чека", len(post_fiscal_edits), "Критично",
     "Цифра в уже отправленном чеке разошлась с текущим отчётом — комплаенс-инцидент, не баг-тикет. См. AUDIT_PLAN.md п.6"],
    ["7", "Чек отправлен без verified/manual", len(invariant_violations), "Критично",
     "Прямое нарушение принципа 'не фискализировать неверифицированное'. См. AUDIT_PLAN.md п.7"],
    ["8", "Часовой пояс сервера", "—", tz_risk, tz_note],
    ["9", "Ордера у границы месяца (риск смещения периода)", len(month_boundary_orders),
     "Средний" if not tz_mismatch else "Критично",
     "Сверить вручную со временем на самой бирже, особенно если Блок 8 = Критично. См. AUDIT_PLAN.md п.9"],
    ["10", "Live (24ч) vs БД — Bybit/MEXC",
     (f"{len(live_reconciliation_rows)} бирж" if RUN_LIVE_RECONCILIATION else "не запускалось"),
     "Высокий" if RUN_LIVE_RECONCILIATION else "—",
     "Включить RUN_LIVE_RECONCILIATION=True для сверки живых данных с БД. См. AUDIT_PLAN.md п.10 и консоль"],
    ["11", "MonthlyManualEntry конфликтует с реальными ордерами", len(manual_conflicts), "Высокий",
     "Ручная цифра, возможно уже поданная в ФНС, молча перестала использоваться. См. AUDIT_PLAN.md п.11"],
    ["12", "Отметки 'Отправлено' у пользователей кроме maks", len(nonzero_doc_submissions), "Средний",
     "0 — ожидаемый симптом бага (гейт захардкожен на username=='maks'); >0 — исключения, разобрать вручную. См. AUDIT_PLAN.md п.12"],
    ["13", "Точечная сверка с биржей (price/amount/cost)",
     (f"{len(spot_check_rows)} ордеров" if RUN_LIVE_SPOT_CHECK else "не запускалось"),
     "—", "Включить RUN_LIVE_SPOT_CHECK=True (НЕ read-only — см. предупреждение в коде). См. AUDIT_PLAN.md примечания"],
    ["—", "Дублирующиеся чеки Evotor (по external_id ord_<id>_...)", "N/A", "—",
     "Нельзя проверить из БД — Order.receipt хранит только последний uuid. Сверять в личном кабинете Evotor/ОФД. См. AUDIT_PLAN.md примечания"],
]
for row in summary:
    ws0.append(row)
    risk = row[3]
    fill = RISK_FILL.get(risk, GREY)
    font = RISK_FONT.get(risk)
    for cell in ws0[ws0.max_row]:
        cell.fill = fill
        if font:
            cell.font = font
autosize(ws0)
ws0.column_dimensions["B"].width = 48
ws0.column_dimensions["E"].width = 70

# ── Остальные листы ──
write_order_sheet("1. Дубли ордеров", duplicates)
write_order_sheet("2. Неверифицир. в отчётах", unverified_in_reports)
write_order_sheet("3. Нулевые-отрицательные", bad_values)
write_order_sheet("4. Курс×Кол-во vs Сумма", cost_mismatch)
write_order_sheet("5. Валюта вне USDT-TON", bad_currency)
write_order_sheet("6. Правки после чека", post_fiscal_edits)
write_order_sheet("7. Чек без верификации", invariant_violations)
write_order_sheet("9. Граница месяца", month_boundary_orders)

# ── Лист 11: конфликты ручных записей ──
ws11 = wb.create_sheet("11. Конфликты ручных записей")
ws11.append(["Пользователь", "Месяц", "Источник", "Gross в ручной записи (руб)", "Диагноз"])
style_header(ws11)
for row in manual_conflicts:
    ws11.append(row)
autosize(ws11)

# ── Лист 12: DocumentSubmission ──
ws12 = wb.create_sheet("12. Отметки ФНС кроме maks")
ws12.append(["Пользователь", "Кол-во отметок 'Отправлено'"])
style_header(ws12)
for username, cnt in doc_submission_rows:
    ws12.append([username, cnt])
    if cnt > 0:
        for cell in ws12[ws12.max_row]:
            cell.fill = YELLOW
autosize(ws12)

# ── Лист 13: точечная сверка с биржей (если запускалась) ──
if RUN_LIVE_SPOT_CHECK:
    ws13 = wb.create_sheet("13. Сверка с биржей")
    ws13.append(["ID системы", "ID биржи", "Биржа", "БД: курс", "БД: кол-во", "БД: сумма",
                 "Биржа: курс", "Биржа: кол-во", "Биржа: сумма"])
    style_header(ws13)
    for row in spot_check_rows:
        ws13.append(list(row))
    autosize(ws13)

wb.save(OUTPUT_PATH)
print(f"\nГотово! Отчёт сохранён: {OUTPUT_PATH}")
print("БД не изменялась" + (" (кроме Блока 13, если он был включён)." if RUN_LIVE_SPOT_CHECK else "."))
print("\nСВОДКА:")
for row in summary:
    print(f"  [{row[0]}] {row[1]}: {row[2]}  (риск: {row[3]})")
