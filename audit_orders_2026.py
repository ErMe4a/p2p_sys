# audit_full_system.py
# READ-ONLY. Полный аудит ордеров за 2026 год -> Excel-отчёт.
# Ничего не меняет в БД, только читает и пишет .xlsx на диск.
#
# Запуск:  python manage.py shell < audit_full_system.py
# Результат: /tmp/audit_full_system.xlsx

from collections import defaultdict
from datetime import datetime
from statistics import median

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo

from orders.models import Order
from django.contrib.auth import get_user_model

User = get_user_model()
MSK = ZoneInfo("Europe/Moscow")
YEAR = 2026
year_start = datetime(YEAR, 1, 1, tzinfo=MSK)
year_end = datetime(YEAR + 1, 1, 1, tzinfo=MSK)
OUTPUT_PATH = "/tmp/audit_full_system.xlsx"

def f(x):
    try:
        return float(x or 0)
    except Exception:
        return 0.0

def fmt_dt(o):
    return o.created_at.astimezone(MSK).strftime("%d.%m.%Y %H:%M")

print("Загружаю ордера из БД...")
orders = list(Order.objects.filter(created_at__gte=year_start, created_at__lt=year_end).order_by("created_at"))
uid_to_name = {u.id: u.username for u in User.objects.all()}
print(f"Загружено: {len(orders)}")


def base_row(o, note=""):
    """Единый формат строки ордера для любой вкладки: ID системы, ID биржи,
    владелец, тип, валюта, кол-во, курс, сумма, биржа, дата, диагноз."""
    return [
        o.id, o.external_id or "", uid_to_name.get(o.user_id, o.user_id),
        o.operation_type, o.currency, round(f(o.amount), 4), round(f(o.price), 4),
        round(f(o.cost), 2), o.exchange_type, fmt_dt(o), note,
    ]

BASE_HEADER = ["ID системы", "ID на бирже", "Владелец", "Тип", "Валюта",
               "Кол-во", "Курс", "Сумма (руб)", "Биржа", "Дата", "Диагноз"]

def is_manual_order(o):
    exch = (getattr(o, "exchange_type", "") or "").lower()
    ext = getattr(o, "external_id", "") or ""
    if "telegram" in exch:
        return True
    if ext.startswith("OB-") or ext.startswith("OS-"):
        return True
    return False


# ============================================================
# 1. ДУБЛИ по external_id (100% совпадение ID биржи у одного юзера)
# ============================================================
by_extid = defaultdict(list)
for o in orders:
    ext = getattr(o, "external_id", "") or ""
    if ext:
        by_extid[(o.user_id, ext)].append(o)

duplicates = []
for (uid, ext), group in by_extid.items():
    if len(group) > 1:
        for o in group:
            duplicates.append((o, f"Дубль: {len(group)} ордеров с одинаковым ID биржи {ext!r}"))


# ============================================================
# 2. КУРС: медиана дня СВОЯ ДЛЯ КАЖДОЙ (валюта, биржа) - не мешаем MEXC с Bybit
# ============================================================
price_by_day_cur_exch = defaultdict(list)
for o in orders:
    price = f(getattr(o, "price", 0))
    if price <= 0:
        continue
    day = o.created_at.astimezone(MSK).date()
    price_by_day_cur_exch[(o.currency, o.exchange_type, day)].append(price)

median_price = {k: median(v) for k, v in price_by_day_cur_exch.items() if len(v) >= 3}

PRICE_DEVIATION_PCT = 30  # порог выше, т.к. база теперь точнее (своя биржа+валюта+день)
price_outliers = []
for o in orders:
    price = f(getattr(o, "price", 0))
    if price <= 0:
        continue
    day = o.created_at.astimezone(MSK).date()
    med = median_price.get((o.currency, o.exchange_type, day))
    if med is None:
        continue
    dev_pct = abs(price - med) / med * 100
    if dev_pct > PRICE_DEVIATION_PCT:
        note = f"Курс {price:.2f} отличается от курса дня на этой же бирже ({med:.2f}) на {dev_pct:.0f}%"
        price_outliers.append((o, note, dev_pct))


# ============================================================
# 3. КОЛИЧЕСТВО: percentile-метод (>10x от p95 юзера), искл. нач. остаток.
# ВАЖНО: p95 считается LEAVE-ONE-OUT - БЕЗ учёта самого проверяемого
# ордера. Без этого при малом кол-ве сделок у юзера сам выброс попадает
# в топ-5% и становится собственным порогом сравнения ("самомаскировка") -
# проверено на синтетическом тесте: при n=16 выброс 15000 на фоне
# обычных ~20-30 НЕ ловился без этой поправки.
# ============================================================
amounts_by_user_cur = defaultdict(list)
for o in orders:
    if (getattr(o, "exchange_type", "") or "") == "Остаток (до 1 фев)":
        continue
    amt = f(o.amount)
    if amt > 0:
        amounts_by_user_cur[(o.user_id, o.currency)].append(amt)

AMOUNT_MULTIPLIER = 10
amount_outliers = []
for o in orders:
    if (getattr(o, "exchange_type", "") or "") == "Остаток (до 1 фев)":
        continue
    amt = f(o.amount)
    if amt <= 0:
        continue
    key = (o.user_id, o.currency)
    all_vals = amounts_by_user_cur.get(key, [])
    if len(all_vals) < 11:  # нужно >=10 ОСТАВШИХСЯ значений после исключения текущего
        continue
    # leave-one-out: убираем ОДНО вхождение текущего значения (не все совпадающие)
    others = list(all_vals)
    others.remove(amt)
    s = sorted(others)
    p95 = s[int(len(s) * 0.95)]
    if p95 <= 0:
        continue
    if amt > p95 * AMOUNT_MULTIPLIER:
        ratio = amt / p95
        note = f"Кол-во в {ratio:.0f} раз больше типичной сделки этого юзера ({p95:.2f}, без учёта этого ордера)"
        amount_outliers.append((o, note, ratio))


# ============================================================
# 4. РУЧНОЙ ВВОД (Telegram/OB-/OS-): классификация по типу поломки поля
# ============================================================
pattern_A, pattern_B, pattern_C, pattern_D = [], [], [], []
NORMAL_ROUNDING_CEILING = 25  # P2P округление до 25% - норма, не показываем

for o in orders:
    if not is_manual_order(o):
        continue
    amount = f(o.amount); price = f(getattr(o, "price", 0)); cost = f(o.cost)
    if amount <= 0 or price <= 0 or cost <= 0:
        continue
    expected = amount * price
    if expected <= 0:
        continue
    dev_pct = (cost - expected) / expected * 100
    if abs(dev_pct) <= NORMAL_ROUNDING_CEILING:
        continue  # обычное округление P2P-сделки, не аномалия

    if abs(price - cost) < max(1, cost * 0.01):
        note = f"Похоже, в поле 'Курс' вписана сумма сделки ({cost:.2f}) вместо реального курса"
        pattern_A.append((o, note, expected))
        continue
    if abs(cost - amount) < max(1, amount * 0.02):
        note = f"Похоже, в поле 'Сумма' вписано количество крипты ({amount:.4f}) вместо рублей"
        pattern_B.append((o, note, expected))
        continue

    matched = False
    for shift in [10, 100, 1000, 0.1, 0.01, 0.001]:
        ratio = cost / expected
        if abs(ratio - shift) / shift < 0.05:
            note = f"Сумма отличается от ожидаемой ровно в {shift}x - похоже на 'забытый ноль' при вводе"
            pattern_C.append((o, note, expected, shift))
            matched = True
            break
    if matched:
        continue

    note = f"Сумма {cost:.2f} не сходится с курс×кол-во ({expected:.2f}), отклонение {dev_pct:+.0f}%, причина неочевидна"
    pattern_D.append((o, note, expected, dev_pct))


# ============================================================
# 5. Подозрительно большая комиссия банка
# ============================================================
COMMISSION_PERCENT_CEILING = 5
FIXED_COMMISSION_RATIO_CEILING = 0.10
big_bank_comm = []
for o in orders:
    comm = f(getattr(o, "commission", 0))
    cost = f(o.cost)
    ct = getattr(o, "commission_type", None)
    if ct == "PERCENT" and 0 < comm <= 100 and comm > COMMISSION_PERCENT_CEILING:
        big_bank_comm.append((o, f"Комиссия банка {comm}% - подозрительно много для банковской комиссии"))
    elif ct in ("MONEY", "RUB", "FIX") and cost > 0 and comm > cost * FIXED_COMMISSION_RATIO_CEILING:
        big_bank_comm.append((o, f"Комиссия банка {comm:.0f}руб = {comm/cost*100:.0f}% от суммы сделки - подозрительно много"))


# ============================================================
# 6. amount == price == cost (битый/тестовый ввод)
# ============================================================
triple_equal = []
for o in orders:
    amount = f(o.amount); price = f(getattr(o, "price", 0)); cost = f(o.cost)
    if amount > 0 and amount == price == cost:
        triple_equal.append((o, "Курс, кол-во и сумма - одно и то же число: явно тестовый/битый ввод"))


# ============================================================
# 7. СПРАВОЧНО: устойчивое системное отклонение курса по юзеру
# (возможный легитимный отдельный канал закупки, требует ручной оценки)
# ============================================================
user_currency_devs = defaultdict(list)
for o, note, dev in price_outliers:
    user_currency_devs[(o.user_id, o.currency)].append(dev)

stable_offset_users = []
for (uid, cur), devs in user_currency_devs.items():
    if len(devs) >= 10:  # устойчивое отклонение на многих ордерах
        stable_offset_users.append((uid, cur, len(devs), median(devs)))


# ============================================================
# ЗАПИСЬ В EXCEL
# ============================================================
print("Формирую Excel-отчёт...")
wb = Workbook()
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RED = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
YELLOW = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

def style_header(ws):
    row = ws[1]
    for cell in row:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, 10), 60)

def write_sheet(name, rows_with_notes):
    ws = wb.create_sheet(name)
    ws.append(BASE_HEADER)
    style_header(ws)
    for item in rows_with_notes:
        o, note = item[0], item[1]
        ws.append(base_row(o, note))
    autosize(ws)
    return ws

# ── Лист 0: ИТОГ ──
ws0 = wb.active
ws0.title = "ИТОГ"
ws0.append(["Категория", "Кол-во ордеров", "Риск", "Что делать"])
style_header(ws0)
summary = [
    ["1. Дубли по ID биржи (external_id)", len(duplicates), "Высокий", "Оставить один из дублей, остальные удалить/пометить"],
    ["2. Курс сильно отличается от курса дня НА ТОЙ ЖЕ бирже", len(price_outliers), "Высокий", f">{PRICE_DEVIATION_PCT}% откл, база - своя биржа+валюта+день"],
    ["3. Аномальное количество (экстрим для этого юзера)", len(amount_outliers), "Средний", f">{AMOUNT_MULTIPLIER}x от p95 юзера, искл. нач. остаток"],
    ["4А. Ручной ввод: Курс=Сумма (поле 'курс' сломано)", len(pattern_A), "Высокий", "Исправить курс, сумма верна"],
    ["4Б. Ручной ввод: Сумма=Кол-во (поле 'сумма' сломано)", len(pattern_B), "Высокий", "Исправить сумму = кол-во х курс"],
    ["4В. Ручной ввод: сдвиг на порядок в сумме", len(pattern_C), "Высокий", "Сдвинуть запятую в сумме на нужный порядок"],
    ["4Г. Ручной ввод: непонятное расхождение >25%", len(pattern_D), "Средний", "Смотреть глазами / спросить пользователя"],
    ["5. Подозрительно большая комиссия банка", len(big_bank_comm), "Средний", f">{COMMISSION_PERCENT_CEILING}% или >{int(FIXED_COMMISSION_RATIO_CEILING*100)}% от суммы"],
    ["6. Битый/тестовый ввод (курс=кол-во=сумма)", len(triple_equal), "Высокий", "Проверить и удалить, если тестовые"],
    ["7. Юзеры со стабильным отклонением курса (справочно)", len(stable_offset_users), "Низкий / не проблема", "Похоже на легитимный отдельный канал закупки - см. вкладку"],
]
for row in summary:
    ws0.append(row)
    fill = RED if row[2] == "Высокий" else (YELLOW if row[2] == "Средний" else GREEN)
    for cell in ws0[ws0.max_row]:
        cell.fill = fill
autosize(ws0)
ws0.column_dimensions["A"].width = 55
ws0.column_dimensions["D"].width = 45

# ── Остальные листы ──
write_sheet("1. Дубли ордеров", duplicates)
write_sheet("2. Курс - выбросы", [(o, n) for o, n, d in price_outliers])
write_sheet("3. Кол-во - выбросы", [(o, n) for o, n, r in amount_outliers])
write_sheet("4А. Ручной - Курс=Сумма", [(o, n) for o, n, e in pattern_A])
write_sheet("4Б. Ручной - Сумма=Кол-во", [(o, n) for o, n, e in pattern_B])
write_sheet("4В. Ручной - сдвиг порядка", [(o, n) for o, n, e, s in pattern_C])
write_sheet("4Г. Ручной - непонятно", [(o, n) for o, n, e, d in pattern_D])
write_sheet("5. Большая комиссия банка", big_bank_comm)
write_sheet("6. Битый ввод", triple_equal)

# ── Справочный лист: устойчивые отклонения по юзеру ──
ws7 = wb.create_sheet("7. Справка - стаб. отклонения")
ws7.append(["Пользователь", "Валюта", "Кол-во ордеров с откл.", "Медианное отклонение %", "Комментарий"])
style_header(ws7)
for uid, cur, cnt, dev in sorted(stable_offset_users, key=lambda x: -x[2]):
    ws7.append([uid_to_name.get(uid, uid), cur, cnt, round(dev, 0),
                "Устойчиво на многих ордерах - вероятно легитимный канал (не опечатка)"])
autosize(ws7)

wb.save(OUTPUT_PATH)
print(f"\nГотово! Отчёт сохранён: {OUTPUT_PATH}")
print("БД не изменялась.")
print("\nСВОДКА:")
for row in summary:
    print(f"  {row[0]}: {row[1]}")
