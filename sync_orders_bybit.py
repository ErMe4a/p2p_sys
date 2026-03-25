"""
Запуск рядом с manage.py:

    python sync_orders_bybit.py

После завершения сохраняет файл:
    bybit_diff_YYYY-MM-DD.xlsx

Залей на git и скачай.

Точность сравнения:
    price, cost  — до 2 знаков
    amount       — до 3 знаков
    operation_type — точное совпадение

Порог: показывает только расхождения > 1.
Что НЕ трогает: комиссию, банк, скриншот, чек, дату.
"""

import os
import sys
import django
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")

try:
    django.setup()
except Exception as e:
    print(f"Ошибка инициализации Django: {e}")
    print("Проверь DJANGO_SETTINGS_MODULE (смотри в manage.py)")
    sys.exit(1)

import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bybit_p2p import P2P
from orders.models import Order
from django.contrib.auth import get_user_model

User = get_user_model()

DATE_FROM_MS   = int(datetime(2026, 2, 1).timestamp() * 1000)
DIFF_THRESHOLD = Decimal("1")
PRECISION = {
    "price":  Decimal("0.01"),
    "cost":   Decimal("0.01"),
    "amount": Decimal("0.001"),
}

FIELD_LABELS = {
    "price":          "Курс",
    "cost":           "Сумма (руб)",
    "amount":         "Кол-во (крипта)",
    "operation_type": "Тип (BUY/SELL)",
}


# ── Bybit API ─────────────────────────────────────────────────────────────────

def fetch_all_bybit_orders(api) -> dict:
    result = {}
    page = 1

    while True:
        try:
            response = api.get_orders(
                page=page,
                size=30,
                status=50,
                beginTime=DATE_FROM_MS,
            )
        except Exception as e:
            print(f"  [!] Ошибка запроса страницы {page}: {e}")
            break

        ret_code = response.get("ret_code")
        if ret_code is None:
            ret_code = response.get("retCode")

        if ret_code != 0:
            msg = response.get("ret_msg") or response.get("retMsg") or "unknown"
            print(f"  [!] Bybit API ошибка (стр.{page}): code={ret_code} msg={msg}")
            if page == 1:
                return None
            break

        items = (
            response.get("result", {}).get("items")
            or response.get("result", {}).get("list")
            or []
        )

        if not items:
            break

        for item in items:
            order_id = str(item.get("id") or item.get("orderId") or "").strip()
            if order_id:
                result[order_id] = _parse_item(item)

        print(f"  Страница {page}: {len(items)} ордеров  (всего: {len(result)})")

        last_ms = int(items[-1].get("createDate") or items[-1].get("createTime") or 0)
        if len(items) < 30 or last_ms < DATE_FROM_MS:
            break

        page += 1
        time.sleep(0.3)

    return result


def _parse_item(item: dict) -> dict:
    side_val       = str(item.get("side") or "")
    operation_type = "SELL" if side_val == "1" else "BUY"
    price          = _to_decimal(item.get("price"))
    crypto_amount  = _to_decimal(item.get("notifyTokenQuantity") or item.get("quantity"))
    fiat_amount    = _to_decimal(item.get("amount"))

    if crypto_amount == Decimal("0") and price > 0 and fiat_amount > 0:
        crypto_amount = (fiat_amount / price).quantize(Decimal("0.00000001"), rounding=ROUND_DOWN)

    return {
        "price":          price,
        "amount":         crypto_amount,
        "cost":           fiat_amount,
        "operation_type": operation_type,
    }


# ── Утилиты ───────────────────────────────────────────────────────────────────

def _to_decimal(value) -> Decimal:
    try:
        return Decimal(str(float(value or 0)))
    except Exception:
        return Decimal("0")


def _round(value, precision: Decimal) -> Decimal:
    try:
        return Decimal(str(value or 0)).quantize(precision, rounding=ROUND_HALF_UP)
    except Exception:
        return Decimal("0")


def fields_differ(order: Order, api_data: dict) -> list:
    diffs = []
    for field in ("price", "cost", "amount"):
        precision = PRECISION[field]
        db_val  = _round(getattr(order, field), precision)
        api_val = _round(api_data[field], precision)
        if abs(db_val - api_val) > DIFF_THRESHOLD:
            diffs.append((field, getattr(order, field), api_data[field]))
    if order.operation_type != api_data["operation_type"]:
        diffs.append(("operation_type", order.operation_type, api_data["operation_type"]))
    return diffs


# ── Excel ─────────────────────────────────────────────────────────────────────

def create_excel(rows: list) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расхождения Bybit"

    # Стили
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1F4E79")
    red_fill       = PatternFill("solid", fgColor="FFE0E0")
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left",   vertical="center")
    thin           = Side(style="thin", color="CCCCCC")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Пользователь",
        "ID ордера",
        "Поле",
        "В базе",
        "На Bybit",
        "Разница",
        "Дата ордера",
    ]

    ws.row_dimensions[1].height = 22

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    # Ширина колонок
    col_widths = [20, 24, 18, 18, 18, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # Данные
    for row_idx, row in enumerate(rows, 2):
        username, order_id, field, db_val, api_val, diff, order_date = row

        values = [username, order_id, FIELD_LABELS.get(field, field),
                  str(db_val), str(api_val), str(diff), order_date]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = red_fill
            cell.border    = border
            cell.alignment = center if col_idx != 1 else left

    # Итого внизу
    last_row = len(rows) + 3
    ws.cell(row=last_row, column=1, value=f"Всего расхождений: {len(rows)}").font = Font(bold=True)

    filename = f"bybit_diff_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(filepath)
    return filepath


# ── Основная логика ───────────────────────────────────────────────────────────

def main():
    print("=" * 80)
    print("  РЕЖИМ: ДИАГНОСТИКА + EXCEL")
    print(f"  Порог расхождения: > {DIFF_THRESHOLD}  |  Ордера с 01.02.2026")
    print("=" * 80)

    users = User.objects.filter(
        order__exchange_type__in=["Bybit", "BYBIT"]
    ).distinct()

    if not users.exists():
        print("\nНет пользователей с Bybit ордерами.")
        return

    total_checked    = 0
    total_differ     = 0
    total_not_in_api = 0
    total_skipped    = 0

    excel_rows = []  # Все строки для Excel

    for user in users:
        has_keys = bool(
            getattr(user, "bybit_api_key", None)
            and getattr(user, "bybit_api_secret", None)
        )

        orders = (
            Order.objects
            .filter(user=user, exchange_type__in=["Bybit", "BYBIT"])
            .exclude(external_id="")
            .exclude(external_id__isnull=True)
            .order_by("-created_at")
        )

        count = orders.count()
        print(f"\nПользователь: {user.username}  "
              f"[{'ключи есть' if has_keys else 'КЛЮЧЕЙ НЕТ'}]  "
              f"ордеров в БД: {count}")

        if not has_keys:
            total_skipped += count
            print("  Пропускаем (нет API ключей Bybit)")
            continue

        try:
            api = P2P(
                testnet=False,
                api_key=user.bybit_api_key,
                api_secret=user.bybit_api_secret,
            )
        except Exception as e:
            print(f"  [!] Ошибка инициализации: {e}")
            total_skipped += count
            continue

        print("  Загружаем ордера с Bybit...")
        api_orders = fetch_all_bybit_orders(api)

        if api_orders is None:
            print("  [!] Не удалось загрузить данные — пропускаем")
            total_skipped += count
            continue

        print(f"  Получено с Bybit API: {len(api_orders)} ордеров")
        print(f"  Сверяем с БД...")

        user_diffs = 0

        for order in orders:
            total_checked += 1

            api_data = api_orders.get(order.external_id)
            if api_data is None:
                total_not_in_api += 1
                continue

            diffs = fields_differ(order, api_data)
            if not diffs:
                continue

            total_differ += 1
            user_diffs   += 1

            order_date = order.created_at.strftime("%d.%m.%Y") if order.created_at else ""

            print(f"  {order.external_id}  ->  РАСХОЖДЕНИЕ:")
            for field, db_val, api_val in diffs:
                prec = PRECISION.get(field, Decimal("0.01"))
                diff = abs(_round(db_val, prec) - _round(api_val, prec))
                print(f"      {FIELD_LABELS.get(field, field)}: {db_val}  =>  {api_val}  (разница: {diff})")

                excel_rows.append((
                    user.username,
                    order.external_id,
                    field,
                    db_val,
                    api_val,
                    diff,
                    order_date,
                ))

        if user_diffs == 0:
            print("  Расхождений нет.")

    # ── Итог ─────────────────────────────────────────────────────────────────
    print("\n" + "=" * 80)
    print(f"  Проверено:           {total_checked}")
    print(f"  Расхождений > 1:     {total_differ}")
    print(f"  Нет в API (старые):  {total_not_in_api}")
    print(f"  Пропущено:           {total_skipped}")
    print("=" * 80)

    if excel_rows:
        filepath = create_excel(excel_rows)
        print(f"\n  Excel сохранён: {filepath}")
        print(f"  Запушь на git и скачай.\n")
    else:
        print("\n  Расхождений не найдено — Excel не создавался.\n")


if __name__ == "__main__":
    main()